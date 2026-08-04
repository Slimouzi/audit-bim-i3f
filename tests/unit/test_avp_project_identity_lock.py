"""Aucun repli ne peut nommer un livrable client.

Le pack a livré ``260803 MCP_Audit 7427L AVP - Contrôle Maquettes.xlsx`` :
``project_name`` n'avait pas été passé, et la résolution est retombée sur le nom
du **projet BIMData**. Or ce nom est celui d'un espace de travail, choisi par
celui qui l'a créé — ``MCP_Audit`` désigne l'outil, pas un chantier.

Trois sources se disputaient l'identité, aucune n'est fiable :

- le projet BIMData → un espace de travail ;
- l'entête d'un classeur MOA → le chantier de **référence** (Tarare 0546L),
  déjà neutralisé, revérifié ici ;
- un libellé générique passé en paramètre (« Projet », « I3F »).

Le tool doit donc refuser et **proposer**, à partir du nom de la maquette, qui
est posé par l'équipe projet.
"""

from __future__ import annotations

import pytest

from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.mcp.session import _Session, current_session
from audit_bim.profiles.i3f.tools_reporting import generate_avp_i3f_pack as tr_generate_avp_i3f_pack

MODELE = "DIEPPE-7427L-BATA-ARCHI-APD (3).ifc"


def _espace(uuid, nom, aire):
    """Espace porteur de sa surface : sans quantité, le tool réclamerait un .ifc
    avant d'en arriver à l'identité, et le test ne prouverait rien."""
    return {
        "uuid": uuid,
        "type": "IfcSpace",
        "name": nom,
        "longname": nom,
        "property_sets": [
            {
                "name": "Qto_SpaceBaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": aire}],
            }
        ],
    }


@pytest.fixture
def session(tmp_path, monkeypatch):
    """Session dont le projet BIMData porte un nom d'espace de travail."""
    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    sess = _Session()
    sess.snapshot = ModelSnapshot(
        project={"name": "MCP_Audit"},
        model={"name": MODELE},
        zones=[{"uuid": "Z1", "type": "IfcZone", "name": "LOGEMENT 01", "space_uuids": ["SP1"]}],
        spaces=[_espace("SP1", "SEJOUR", 25.4)],
    ).index()
    token = current_session.set(sess)
    try:
        yield sess
    finally:
        current_session.reset(token)


def _generer(**kw):
    params = {
        "phase": "APD",
        "auditor_name": "Stanislas Limouzi",
        "export_pdf": False,
    }
    params.update(kw)
    return tr_generate_avp_i3f_pack(**params)


def _question(res, key):
    return next(q for q in res["questions"] if q["key"] == key)


def test_missing_name_does_not_fall_back_on_the_bimdata_project(session):
    """Le cas exact qui a produit « 260803 MCP_Audit 7427L AVP - … »."""
    res = _generer(project_code="7427L")

    assert res["status"] == "needs_context"
    assert "project_name" in res["missing"]
    assert "output_dir" not in res


def test_refusal_suggests_the_name_read_from_the_model(session):
    """La suggestion vient de la maquette, pas de l'espace BIMData."""
    res = _generer(project_code="7427L")

    q = _question(res, "project_name")
    assert q["suggestion"] == "DIEPPE"
    assert "MCP_Audit" not in q["question"]


@pytest.mark.parametrize("nom", ["I3F", "Projet", "MCP_Audit", "Tarare", "mcp audit", "PROJET"])
def test_generic_names_are_refused(session, nom):
    """Un libellé générique est refusé au même titre qu'un paramètre absent.

    Casse et séparateurs neutralisés : ``mcp audit`` ne doit pas passer là où
    ``MCP_Audit`` est refusé.
    """
    res = _generer(project_name=nom, project_code="7427L")

    assert res["status"] == "needs_context"
    assert "project_name" in res["missing"]
    q = _question(res, "project_name")
    assert q["rejected"] == nom
    assert res.get("output_dir") is None


def test_generic_code_is_refused(session):
    """Le code aussi nomme le fichier : même exigence."""
    res = _generer(project_name="Dieppe", project_code="Projet")

    assert res["status"] == "needs_context"
    assert "project_code" in res["missing"]


def test_explicit_real_identity_is_accepted(session):
    """Nom et code réels : la génération est autorisée."""
    res = _generer(project_name="Dieppe", project_code="7427L")

    assert res.get("status") != "needs_context", res
    assert res["project_name"] == "Dieppe"
    assert res["project_code"] == "7427L"


def test_moa_template_header_can_never_name_the_deliverables(session, tmp_path, monkeypatch):
    """Même désigné explicitement, un classeur MOA ne fournit pas l'identité.

    Son entête ne peut qu'alimenter une **suggestion** ; elle ne doit jamais
    devenir la valeur retenue, sans quoi un pack Dieppe sortirait sous le nom du
    chantier de référence.
    """
    import openpyxl

    from audit_bim.profiles.i3f import tools_reporting

    modele = tmp_path / "controle.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grille de contrôle"
    ws["B2"], ws["C2"] = "Projet", "Tarare"
    ws["B3"], ws["C3"] = "ESI", "0546L"
    wb.save(modele)
    monkeypatch.setattr(tools_reporting, "_auto_controle_xlsx", lambda: None)

    res = _generer(controle_xlsx=str(modele))

    assert res["status"] == "needs_context"
    assert {"project_name", "project_code"} <= set(res["missing"])
    # L'entête ne s'impose pas : la maquette reste la source de la suggestion.
    assert _question(res, "project_name")["suggestion"] == "DIEPPE"


def test_moa_header_never_even_suggests_an_identity(tmp_path, monkeypatch):
    """Refuser une mauvaise valeur ne suffit pas : il faut cesser de la proposer.

    Le gabarit porte ici l'identité d'un chantier **réel mais étranger**, absente
    de toute liste de libellés génériques — c'est le cas que le filtrage par
    denylist ne pourrait pas attraper. Aucune suggestion ne doit en sortir : une
    proposition affichée finit recopiée.
    """
    import openpyxl

    from audit_bim.profiles.i3f import tools_reporting

    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    monkeypatch.setattr(tools_reporting, "_auto_controle_xlsx", lambda: None)

    modele = tmp_path / "controle.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grille de contrôle"
    ws["B2"], ws["C2"] = "Projet", "Liffré"
    ws["B3"], ws["C3"] = "ESI", "0912K"
    wb.save(modele)

    sess = _Session()
    # Maquette sans identité lisible : aucune suggestion ne peut en venir, donc
    # ce qui resterait affiché viendrait forcément de l'entête.
    sess.snapshot = ModelSnapshot(project={"name": "MCP_Audit"}, model={"name": "M.ifc"}).index()
    token = current_session.set(sess)
    try:
        res = _generer(controle_xlsx=str(modele))
    finally:
        current_session.reset(token)

    assert res["status"] == "needs_context"
    for cle in ("project_name", "project_code"):
        q = _question(res, cle)
        assert "suggestion" not in q, q
        assert "Liffré" not in q["question"]
        assert "0912K" not in q["question"]


def test_confirm_context_never_bypasses_identity(session):
    """``confirm_context`` couvre l'auteur du contrôle, jamais ce qui nomme."""
    res = _generer(confirm_context=True)

    assert res["status"] == "needs_context"
    assert {"project_name", "project_code"} <= set(res["missing"])
    assert "OBLIGATOIRES" in res["next_step"]


# ── la phase aussi nomme le fichier ────────────────────────────────────


def _session_sans_phase(tmp_path, monkeypatch, modele=MODELE):
    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    sess = _Session()
    sess.snapshot = ModelSnapshot(
        project={"name": "MCP_Audit"},
        model={"name": modele},
        zones=[{"uuid": "Z1", "type": "IfcZone", "name": "LOGEMENT 01", "space_uuids": ["SP1"]}],
        spaces=[_espace("SP1", "SEJOUR", 25.4)],
    ).index()
    sess.phase = None
    return sess


def _gabarit_moa_avp(tmp_path):
    """Classeur de contrôle du chantier de référence, en phase AVP."""
    import openpyxl

    chemin = tmp_path / "controle.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grille de contrôle"
    ws["B2"], ws["C2"] = "Projet", "Tarare"
    ws["B3"], ws["C3"] = "ESI", "0546L"
    ws["B4"], ws["C4"] = "Phase", "AVP"
    wb.save(chemin)
    return str(chemin)


def test_moa_header_phase_never_names_the_deliverables(tmp_path, monkeypatch):
    """Un gabarit AVP ne doit pas nommer « AVP » un pack dont la phase est inconnue.

    Même famille que le nom et le code, en plus discret : une phase reste
    toujours plausible, donc rien ne signale l'erreur à la relecture. Le pack
    sortirait « <Projet> <Code> AVP - … » sur une opération en APD.
    """
    from audit_bim.profiles.i3f import tools_reporting

    monkeypatch.setattr(tools_reporting, "_auto_controle_xlsx", lambda: None)
    sess = _session_sans_phase(tmp_path, monkeypatch)
    token = current_session.set(sess)
    try:
        res = tr_generate_avp_i3f_pack(
            project_name="Dieppe",
            project_code="7427L",
            auditor_name="Stanislas Limouzi",
            controle_xlsx=_gabarit_moa_avp(tmp_path),
            export_pdf=False,
        )
    finally:
        current_session.reset(token)

    assert res["status"] == "needs_context"
    assert "project_phase" in res["missing"]
    assert res.get("phase") is None
    assert not list(tmp_path.glob("avp_pack_*"))


def test_phase_is_not_bypassable_by_confirm_context(tmp_path, monkeypatch):
    """Elle nomme le fichier : même statut que ``project_name`` / ``project_code``."""
    from audit_bim.profiles.i3f import tools_reporting

    monkeypatch.setattr(tools_reporting, "_auto_controle_xlsx", lambda: None)
    sess = _session_sans_phase(tmp_path, monkeypatch)
    token = current_session.set(sess)
    try:
        res = tr_generate_avp_i3f_pack(
            project_name="Dieppe",
            project_code="7427L",
            auditor_name="Stanislas Limouzi",
            controle_xlsx=_gabarit_moa_avp(tmp_path),
            confirm_context=True,
            export_pdf=False,
        )
    finally:
        current_session.reset(token)

    assert res["status"] == "needs_context"
    assert "project_phase" in res["missing"]
    assert "OBLIGATOIRES" in res["next_step"]


def test_confirmed_session_phase_is_still_a_valid_source(tmp_path, monkeypatch):
    """``_State.phase`` a été confirmée par l'auditeur : ce n'est pas un repli."""
    from audit_bim.requirements.models import BIMPhase

    sess = _session_sans_phase(tmp_path, monkeypatch)
    sess.phase = BIMPhase.AVP
    token = current_session.set(sess)
    try:
        res = tr_generate_avp_i3f_pack(
            project_name="Dieppe",
            project_code="7427L",
            auditor_name="Stanislas Limouzi",
            export_pdf=False,
        )
    finally:
        current_session.reset(token)

    assert res.get("status") != "needs_context", res
    assert res["phase"] == "AVP"
