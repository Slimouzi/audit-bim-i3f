"""Tests de :mod:`audit_bim.doe.address` (auto-extraction adresse DOE)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from audit_bim.doe.address import (
    extract_address_from_doe,
    extract_address_from_text,
)
from audit_bim.enrichment.models import ProjectAddress

# ── extract_address_from_text ───────────────────────────────────────────


class TestExtractFromText:
    def test_full_match_street_cp_town(self):
        text = (
            "DOSSIER DES OUVRAGES EXÉCUTÉS\n"
            "Opération : Résidence Le Belvédère\n"
            "10 Rue de Rivoli\n"
            "75004 PARIS\n"
            "Maître d'ouvrage : I3F\n"
        )
        addr = extract_address_from_text(text)
        assert isinstance(addr, ProjectAddress)
        assert addr.source == "doe"
        assert addr.postal_code == "75004"
        assert addr.town == "PARIS"
        assert addr.address_lines and "Rivoli" in addr.address_lines[0]
        assert addr.address_lines[0].lower().startswith("10")

    def test_match_cp_only_when_no_street(self):
        text = "Adresse projet : 13001 MARSEILLE\nMOA : Habitat 13"
        addr = extract_address_from_text(text)
        assert addr is not None
        assert addr.postal_code == "13001"
        assert addr.town == "MARSEILLE"
        assert addr.address_lines == []

    def test_returns_none_without_cp(self):
        text = "Pas d'adresse ici, juste du blabla."
        assert extract_address_from_text(text) is None

    def test_returns_none_on_empty(self):
        assert extract_address_from_text("") is None
        assert extract_address_from_text(None) is None  # type: ignore[arg-type]

    def test_handles_comma_separator(self):
        text = "Site : 42, Boulevard Voltaire — 75011 PARIS\n"
        addr = extract_address_from_text(text)
        assert addr is not None
        assert addr.postal_code == "75011"
        assert addr.town == "PARIS"
        assert addr.address_lines and "Voltaire" in addr.address_lines[0]

    def test_handles_bis_ter(self):
        text = "Implantation\n5 bis avenue de la République\n69100 VILLEURBANNE\n"
        addr = extract_address_from_text(text)
        assert addr is not None
        assert addr.postal_code == "69100"
        assert addr.town == "VILLEURBANNE"
        assert "5 bis" in addr.address_lines[0]

    def test_picks_street_close_to_cp_not_random_number(self):
        # Un n° parasite (bât. 12) loin de l'adresse, mais une vraie
        # adresse à proximité du CP doit gagner.
        text = (
            "Bâtiment 12\n"
            "Centre de soins — équipement médical\n"
            "...\n"
            "...\n"
            "Localisation du site :\n"
            "8 rue du Faubourg Saint-Antoine\n"
            "75011 PARIS\n"
        )
        addr = extract_address_from_text(text)
        assert addr is not None
        assert addr.postal_code == "75011"
        assert "Faubourg" in addr.address_lines[0]


# ── extract_address_from_doe (dispatch par extension) ───────────────────


class TestExtractFromDoe:
    def test_returns_none_on_missing_file(self, tmp_path):
        assert extract_address_from_doe(tmp_path / "ghost.xlsx") is None

    def test_returns_none_on_unsupported_extension(self, tmp_path):
        p = tmp_path / "doc.txt"
        p.write_text("10 rue de Rivoli 75004 PARIS", encoding="utf-8")
        # .txt n'est pas dispatché → None (même si le contenu contient
        # une adresse, on ne lit pas les .txt).
        assert extract_address_from_doe(p) is None

    def test_extracts_from_xlsx_header(self, tmp_path: Path):
        # On colle l'adresse dans l'en-tête (lignes 1-5), pas dans le
        # tableau d'équipements en dessous.
        p = tmp_path / "doe.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "DOSSIER DES OUVRAGES EXÉCUTÉS"
        ws["A2"] = "Opération : Résidence Le Belvédère"
        ws["A3"] = "10 Rue de Rivoli"
        ws["A4"] = "75004 PARIS"
        ws["A5"] = ""
        # Zone tabulaire en dessous (ne doit PAS perturber l'extraction)
        ws["A7"] = "UUID"
        ws["B7"] = "Nom"
        ws["A8"] = "abc-123"
        ws["B8"] = "Porte coupe-feu"
        wb.save(p)

        addr = extract_address_from_doe(p)
        assert addr is not None
        assert addr.source == "doe"
        assert addr.postal_code == "75004"
        assert addr.town == "PARIS"
        assert "Rivoli" in addr.address_lines[0]

    def test_xlsx_without_address_returns_none(self, tmp_path: Path):
        p = tmp_path / "doe.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "UUID"
        ws["B1"] = "Nom"
        ws["A2"] = "abc-123"
        ws["B2"] = "Porte coupe-feu"
        wb.save(p)
        assert extract_address_from_doe(p) is None


# ── Intégration avec resolve_project_address (fallback DOE) ─────────────


class TestEnrichmentFallbackToDoe:
    def test_doe_fallback_when_ifc_empty(self, tmp_path: Path):
        from types import SimpleNamespace

        from audit_bim.enrichment.address import resolve_project_address

        p = tmp_path / "doe.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Implantation : 5 bis avenue de la République"
        ws["A2"] = "69100 VILLEURBANNE"
        wb.save(p)

        snap = SimpleNamespace(buildings=[], sites=[])
        addr = resolve_project_address(snap, doe_path=str(p))
        assert addr.source == "doe"
        assert addr.town == "VILLEURBANNE"

    def test_ifc_wins_over_doe(self, tmp_path: Path):
        from types import SimpleNamespace

        from audit_bim.enrichment.address import resolve_project_address

        p = tmp_path / "doe.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "10 Rue de Rivoli 75004 PARIS"
        wb.save(p)

        snap = SimpleNamespace(
            buildings=[
                {
                    "BuildingAddress": {
                        "AddressLines": ["1 boulevard du Test"],
                        "PostalCode": "13001",
                        "Town": "MARSEILLE",
                    }
                }
            ],
            sites=[],
        )
        # IFC prioritaire — DOE non utilisé même si fourni.
        addr = resolve_project_address(snap, doe_path=str(p))
        assert addr.source == "ifc_building"
        assert addr.town == "MARSEILLE"

    def test_raises_when_nothing_anywhere(self, tmp_path: Path):
        from types import SimpleNamespace

        from audit_bim.enrichment.address import resolve_project_address

        # Pas d'IFC, pas de DOE → erreur explicite mentionnant le DOE.
        p = tmp_path / "doe.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "pas d'adresse ici"
        wb.save(p)

        snap = SimpleNamespace(buildings=[], sites=[])
        with pytest.raises(ValueError, match="DOE"):
            resolve_project_address(snap, doe_path=str(p))
