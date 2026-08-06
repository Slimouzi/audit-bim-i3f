"""Inventaire de la table des attributs MRN — compteurs figés, éprouvés en CI.

Le fichier du maître d'ouvrage vit hors du dépôt. Une fixture synthétique
reproduit donc les compteurs **par construction**, pour que la CI prouve ce
qu'elle affirme : les tests sur le fichier local restent, mais ne sont plus les
seuls garants.

Une exigence atomique est **une ligne de propriété attendue**. Les 4 620
cellules d'applicabilité sont un diagnostic — les appeler « exigences »
multiplierait le référentiel par quatre.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from audit_bim.profiles.bim_in_motion.mrn.attributes import (
    EXPECTED_PHASES,
    SHEET_LAYOUTS,
    parse_mrn_attribute_table,
)

TABLE = Path(
    "/Users/stani/code/MCP/Documents maître d'ouvrage/Documents BIM in Motion/"
    "Données d'entrée/MRN_CHARTE_Annexe_Table des attributs_03082025.xlsx"
)

needs_table = pytest.mark.skipif(not TABLE.is_file(), reason=f"table MRN absente : {TABLE}")

#: Compteurs arrêtés après mesure, feuille par feuille.
EXPECTED = {
    "Généralités": (80, 769),
    "Gros Oeuvre - CEA": (270, 1389),
    "CVC-PLB-SSI-ELEC": (532, 1987),
    "VRD-Extérieur": (131, 475),
}
EXPECTED_ROWS = 1013
EXPECTED_CELLS = 4620

#: Layout **indépendant** de la production, relevé à la main sur le classeur du
#: maître d'ouvrage. Construire la fixture depuis ``SHEET_LAYOUTS`` rendait le
#: garde-fou circulaire : le parseur était validé contre lui-même, et remettre
#: ``Gros Oeuvre`` en colonne D laissait passer les compteurs inchangés.
#:
#: (header_row, property_column, carrier_columns, phase_columns)
FROZEN_LAYOUT = {
    "Généralités": (3, 4, (8, 14), (15, 22)),
    "Gros Oeuvre - CEA": (3, 7, None, (10, 17)),
    "CVC-PLB-SSI-ELEC": (3, 7, None, (10, 17)),
    "VRD-Extérieur": (2, 7, None, (10, 17)),
}


def test_the_production_layout_matches_the_independently_frozen_one():
    """Le contrôle que la fixture ne peut pas rendre : les deux sources doivent coïncider.

    Sans lui, remettre une feuille technique en colonne ``D`` passerait — la
    fixture serait construite avec la même erreur qu'elle est censée détecter.
    """
    for name, (header, prop, carriers, phases) in FROZEN_LAYOUT.items():
        layout = SHEET_LAYOUTS[name]
        assert layout.header_row == header, name
        assert layout.property_column == prop, name
        assert layout.carrier_columns == carriers, name
        assert layout.phase_columns == phases, name
    assert set(SHEET_LAYOUTS) == set(FROZEN_LAYOUT)


def test_the_technical_sheets_never_read_the_type_object_column():
    """Colonne ``D`` = ``IfcTypeObject`` sur les feuilles techniques.

    L'y lire donnait 178 exigences au lieu de 1 013. Le contrôle nomme l'erreur
    plutôt que de la laisser réapparaître sous un compteur juste.
    """
    for name in ("Gros Oeuvre - CEA", "CVC-PLB-SSI-ELEC", "VRD-Extérieur"):
        assert SHEET_LAYOUTS[name].property_column == 7, f"{name} doit lire G, pas D"
        assert SHEET_LAYOUTS[name].ifc_type_object_column == 4


# ── Fixture synthétique : les compteurs tournent sans le fichier client ──


def _build_synthetic_table(
    path: Path,
    *,
    broken_phase_headers: bool = False,
    formula_without_value: bool = False,
) -> Path:
    """Reconstruit une table à la forme du fichier réel, aux mêmes compteurs.

    Les croix sont distribuées de façon déterministe pour retomber exactement
    sur les totaux mesurés — dont l'unique ``x+`` de ``CVC-PLB-SSI-ELEC``.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for name, (rows, cells) in EXPECTED.items():
        header_row, property_column, carriers, phase_span = FROZEN_LAYOUT[name]
        carrier_span = range(carriers[0], carriers[1] + 1) if carriers else range(0)
        phase_span = range(phase_span[0], phase_span[1] + 1)
        sheet = workbook.create_sheet(name)

        phases = list(EXPECTED_PHASES)
        if broken_phase_headers and name == "VRD-Extérieur":
            # Décalage d'une colonne : le bloc déclaré tomberait sur « Exemple
            # de valeur », exactement le faux positif que le refus doit voir.
            phases = ["Exemple de valeur", *phases[:-1]]
        for index, col in enumerate(phase_span):
            sheet.cell(header_row, col).value = phases[index]
        for index, col in enumerate(carrier_span):
            sheet.cell(header_row, col).value = f"MAQ{index}"

        # Colonnes métier de VRD qu'un parseur par mots-clés confondrait.
        if name == "VRD-Extérieur":
            sheet.cell(header_row, 6).value = "Nom de la propriété"
            sheet.cell(header_row, 7).value = "Attribut/Propriétés"
            sheet.cell(header_row, 9).value = "Exemple de valeur"

        applicable = list(carrier_span) + list(phase_span)
        first = header_row + 1

        # Toutes les lignes d'abord : le nombre d'exigences ne dépend pas du
        # nombre de croix, et les mêler ferait varier l'un avec l'autre.
        for offset in range(rows):
            row = first + offset
            sheet.cell(row, property_column).value = f"Prop_{name[:3]}_{offset}"
            if offset == 0:
                sheet.cell(row, 1).value = "Objet fusionné"
                sheet.cell(row, 3).value = "IfcWall"

        # Puis les croix, distribuées colonne par colonne jusqu'au total exact.
        placed = 0
        while placed < cells:
            row = first + (placed // len(applicable)) % rows
            col = applicable[placed % len(applicable)]
            sheet.cell(row, col).value = "x"
            placed += 1

        # L'unique croix nuancée du classeur réel.
        if name == "CVC-PLB-SSI-ELEC":
            sheet.cell(first, phase_span[0]).value = "x+"

        if formula_without_value and name == "Généralités":
            sheet.cell(first, phase_span[0]).value = '=IF(1=1,"x","")'

    workbook.create_sheet("Liste des pièces").sheet_state = "hidden"
    workbook.save(path)
    return path


@pytest.fixture(scope="module")
def synthetic(tmp_path_factory):
    return parse_mrn_attribute_table(
        _build_synthetic_table(tmp_path_factory.mktemp("mrn") / "table.xlsx")
    )


def test_the_synthetic_fixture_reproduces_the_counters(synthetic):
    """Le contrôle qui tourne réellement en CI."""
    summary = synthetic.summary()
    assert summary["requirement_rows"] == EXPECTED_ROWS
    assert summary["applicability_cells_effective"] == EXPECTED_CELLS
    for name, (rows, cells) in EXPECTED.items():
        assert summary["per_sheet"][name]["requirement_rows"] == rows, name
        assert summary["per_sheet"][name]["applicability_cells"] == cells, name


def test_the_nuanced_marker_is_counted_without_being_flattened(synthetic):
    """``x+`` est une croix, mais pas une croix ordinaire.

    L'aplatir perdrait la nuance ; l'ignorer perdrait une exigence visible dans
    le classeur.
    """
    assert synthetic.marker_variants == {"x": 4619, "x+": 1}
    nuanced = [req for req in synthetic.requirements if req.nuanced_marks]
    assert len(nuanced) == 1, "une seule croix nuancée dans tout le classeur"
    mark = nuanced[0].nuanced_marks[0]
    assert (mark.marker, mark.marker_kind, mark.axis) == ("x+", "applicable_with_note", "phase")
    # La nuance doit survivre jusqu'à l'exigence : sans elle, PR 3 ne saurait
    # plus laquelle des 1013 portait un x+.
    assert mark.label in nuanced[0].applicable_phases


def test_no_applicability_comes_from_a_formula(synthetic):
    """Constat, pas hypothèse.

    L'écart de 4 cellules observé pendant la mesure avait été attribué aux
    formules. C'était **faux** : il n'y en a aucune. ``formula_value_missing``
    est donc un garde-fou préventif, pas l'explication d'un delta constaté.
    """
    assert synthetic.formula_applicability_cells == 0


# ── Refus explicites ──────────────────────────────────────────────────


def test_a_shifted_phase_block_is_refused(tmp_path):
    """Un bloc décalé tomberait sur « Exemple de valeur ».

    Sans ce refus, une colonne métier serait comptée comme de l'applicabilité —
    et le total resterait plausible.
    """
    path = _build_synthetic_table(tmp_path / "decale.xlsx", broken_phase_headers=True)
    with pytest.raises(ValueError, match="bloc de phases"):
        parse_mrn_attribute_table(path)


def test_a_formula_without_cached_value_is_refused(tmp_path):
    """``formula_value_missing`` doit savoir se déclencher.

    Il ne le fait pas sur le fichier réel — d'où ce test. Un garde-fou qu'aucun
    cas n'exerce affirme sans mesurer.
    """
    path = _build_synthetic_table(tmp_path / "formule.xlsx", formula_without_value=True)
    with pytest.raises(ValueError, match="formula_value_missing"):
        parse_mrn_attribute_table(path)


def test_a_missing_requirement_sheet_is_refused(tmp_path):
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active.title = "Généralités"
    path = tmp_path / "incomplet.xlsx"
    workbook.save(path)

    with pytest.raises(ValueError, match="absentes"):
        parse_mrn_attribute_table(path)


# ── VRD : les colonnes métier ne sont jamais de l'applicabilité ───────


def test_vrd_never_counts_its_business_columns(synthetic):
    """F, G et I portent « propriété » et « exemple », pas des phases."""
    layout = SHEET_LAYOUTS["VRD-Extérieur"]
    assert layout.header_row == 2
    assert layout.phase_columns == (10, 17)  # J:Q
    assert layout.carrier_columns is None
    for forbidden in (6, 7, 9):  # F, G, I
        assert forbidden not in layout.phase_range()


def test_a_keyword_parser_would_have_been_wrong(tmp_path):
    """Non-vacuité : la détection par mots-clés compte à tort F, G et I.

    C'est cette approche qui attribuait trois colonnes de trop à VRD —
    « Nom de la **pro**priété » et « **Exe**mple de valeur » contiennent des
    libellés de phase par coïncidence.
    """
    import openpyxl

    path = _build_synthetic_table(tmp_path / "vrd.xlsx")
    sheet = openpyxl.load_workbook(path)["VRD-Extérieur"]

    words = {"ESQ", "DIAG", "APS", "APD", "PRO", "DCE", "EXE", "DOE", "EXPL", "MAINT"}
    naive = [
        col
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(2, col).value
        and any(word in str(sheet.cell(2, col).value).upper() for word in words)
    ]
    assert {6, 7, 9} <= set(naive), "le parseur naïf doit bien capter F, G et I"
    assert set(naive) != set(SHEET_LAYOUTS["VRD-Extérieur"].phase_range())


def test_carrier_scope_is_never_inherited(synthetic):
    """Une feuille sans colonnes maquette ne dit rien de ses porteurs.

    L'inventer depuis ``Généralités`` attribuerait des exigences à des maquettes
    que le document ne désigne pas.
    """
    for name in ("Gros Oeuvre - CEA", "CVC-PLB-SSI-ELEC", "VRD-Extérieur"):
        for req in synthetic.by_sheet(name):
            assert req.carrier_models == []
            assert req.carrier_scope == "non_specifie"
    for req in synthetic.by_sheet("Généralités"):
        assert req.carrier_scope == "declare"


def test_merged_cells_are_forward_filled(synthetic):
    """Un objet déclaré une fois vaut pour les lignes qu'il coiffe."""
    rows = synthetic.by_sheet("Généralités")
    assert rows[0].object_name == "Objet fusionné"
    assert rows[5].object_name == "Objet fusionné", "la valeur doit se propager"


# ── Fichier réel — conservé, mais pas seul garant ─────────────────────


@needs_table
def test_the_real_table_matches_the_frozen_counters():
    table = parse_mrn_attribute_table(TABLE)
    summary = table.summary()
    assert summary["requirement_rows"] == EXPECTED_ROWS
    assert summary["applicability_cells_effective"] == EXPECTED_CELLS
    assert summary["applicability_marker_variants"] == {"x": 4619, "x+": 1}
    assert summary["formula_applicability_cells"] == 0
    for name, (rows, cells) in EXPECTED.items():
        assert summary["per_sheet"][name] == {
            "requirement_rows": rows,
            "applicability_cells": cells,
        }


@needs_table
def test_the_real_table_hides_its_reference_lists():
    """Listes, classifications et dictionnaires sont masqués, pas des exigences."""
    table = parse_mrn_attribute_table(TABLE)
    assert len(table.hidden_sheets) == 8
    assert set(SHEET_LAYOUTS).isdisjoint(table.hidden_sheets)


def test_the_attribute_parser_never_imports_the_i3f_profile():
    import ast

    from audit_bim.profiles.bim_in_motion.mrn import attributes as module

    tree = ast.parse(Path(module.__file__).read_text(encoding="utf-8"))
    for node in ast.walk(tree):
        modules = []
        if isinstance(node, ast.Import):
            modules = [alias.name for alias in node.names]
        elif isinstance(node, ast.ImportFrom) and node.module:
            modules = [node.module] + [f"{node.module}.{a.name}" for a in node.names]
        assert not [m for m in modules if "i3f" in m]


@needs_table
def test_the_real_nuanced_mark_is_pinned_to_its_requirement():
    """L'unique ``x+`` du classeur réel, à sa place exacte.

    Le test synthétique prouve que la nuance survit au parcours ; il ne dit rien
    de l'endroit où elle se trouve dans le document du maître d'ouvrage. Sans ce
    contrôle, l'affirmation « la croix nuancée est figée » dépassait la mesure —
    et un déplacement de cette exigence passerait inaperçu.
    """
    table = parse_mrn_attribute_table(TABLE)
    nuanced = [req for req in table.requirements if req.nuanced_marks]

    assert len(nuanced) == 1
    requirement = nuanced[0]
    assert requirement.sheet == "CVC-PLB-SSI-ELEC"
    assert requirement.row == 443
    assert requirement.property_name == "Rearmement"

    mark = requirement.nuanced_marks[0]
    assert (mark.axis, mark.label, mark.marker, mark.marker_kind) == (
        "phase",
        "APD",
        "x+",
        "applicable_with_note",
    )
    assert "APD" in requirement.applicable_phases
