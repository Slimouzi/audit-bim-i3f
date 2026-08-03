"""Résolution de l'identité projet du pack AVP I3F (tool MCP) + phase unique.

Couvre les correctifs de revue :
- P1 : l'entête « Projet » du classeur de contrôle ne nomme **jamais** les
  livrables — ce classeur est le plus souvent un template MOA de référence
  (Tarare 0546L), et son identité nommait les packs d'après un autre chantier
  que celui audité. Elle n'en est même plus une **suggestion** : le produit ne
  doit pas seulement refuser les mauvaises valeurs, il doit cesser de les
  proposer. Ordre strict : paramètre explicite > on demande.
- P2a : phase absente → ``needs_context`` (project_phase), jamais de défaut
  silencieux « AVP ».
- P2b : ``project_context_questions`` pose une **unique** question de phase
  (clé ``project_phase``, aide loi MOP), sans suggestion « PRO » codée.
"""

from __future__ import annotations

import openpyxl
import pytest

from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.mcp import server as mcp_server
from audit_bim.mcp.session import _Session, current_session
from audit_bim.requirements.models import BIMPhase


@pytest.fixture
def _isolated(tmp_path, monkeypatch):
    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    # Mode input permissif (pas d'AUDIT_INPUT_DIR) → tmp_path autorisé.
    monkeypatch.delenv("AUDIT_INPUT_DIR", raising=False)
    sess = _Session()
    token = current_session.set(sess)
    try:
        yield sess, tmp_path
    finally:
        current_session.reset(token)


def _controle_xlsx(path, *, projet="Tarare", esi="0546L", phase="AVP"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grille de contrôle"
    # Bloc entête : label en col B, valeur en col C.
    rows = [("Projet", projet), ("ESI", esi)]
    if phase is not None:
        rows.append(("Phase", phase))
    r = 1
    for label, val in rows:
        ws.cell(r, 2, label)
        ws.cell(r, 3, val)
        r += 1
    # Grille minimale.
    ws.cell(r, 1, "CODE 3F")
    ws.cell(r, 2, "POINTS DE CONTROLE")
    wb.save(path)
    return str(path)


def _attach_minimal_snapshot(sess):
    sess.snapshot = ModelSnapshot(project={"name": "I3F"}, model={"name": "M.ifc"}).index()


def test_generate_pack_requires_snapshot(_isolated):
    """Le flux MCP AVP est maquette-first : sans snapshot, on demande
    l'extraction avant de générer."""
    _sess, tmp_path = _isolated
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx")

    res = mcp_server.generate_avp_i3f_pack(controle_xlsx=ctrl, auditor="AMO BIM", export_pdf=False)

    assert res.get("status") == "needs_context"
    assert "snapshot" in res["missing"]


def test_control_header_never_names_the_deliverables(_isolated):
    """P1 : l'entête du classeur de contrôle ne nomme JAMAIS les livrables.

    Le classeur fourni est celui du projet de référence MOA (Tarare 0546L) ;
    la maquette auditée est une autre. Générer un pack « Tarare » serait livrer
    au client des fichiers au nom d'un autre chantier.
    """
    sess, tmp_path = _isolated
    sess.snapshot = ModelSnapshot(project={"name": "I3F"}, model={"name": "M.ifc"}).index()
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx")

    res = mcp_server.generate_avp_i3f_pack(controle_xlsx=ctrl, auditor="AMO BIM", export_pdf=False)

    # « I3F » est un libellé générique et le contexte modèle n'est plus une
    # source d'identité : nom ET code sont demandés.
    assert res.get("status") == "needs_context"
    assert {"project_name", "project_code"} <= set(res["missing"])
    assert res.get("project_name") != "Tarare"

    # L'entête n'est même plus une SUGGESTION : une valeur proposée finit
    # recopiée, et « Tarare / 0546L » nommerait un pack d'après un autre chantier.
    for cle in ("project_name", "project_code"):
        q = next(q for q in res["questions"] if q["key"] == cle)
        assert "suggestion" not in q, q
        assert "Tarare" not in q["question"]
        assert "0546L" not in q["question"]


def test_auto_discovered_control_template_never_suggests_its_identity(_isolated):
    """Un classeur AUTO-DÉCOUVERT ne suggère rien : l'utilisateur ne l'a pas désigné."""
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    _controle_xlsx(tmp_path / "controle maquettes.xlsx")
    monkey_dir = tmp_path

    import os

    os.environ["AUDIT_INPUT_DIR"] = str(monkey_dir)
    try:
        res = mcp_server.generate_avp_i3f_pack(auditor="AMO BIM", export_pdf=False)
    finally:
        os.environ.pop("AUDIT_INPUT_DIR", None)

    assert res.get("status") == "needs_context"
    q = next(q for q in res["questions"] if q["key"] == "project_code")
    assert "suggestion" not in q
    assert "Tarare" not in q["question"] and "0546L" not in q["question"]


def test_missing_phase_asks_instead_of_defaulting_avp(_isolated):
    """P2a : aucune phase (ni paramètre, ni ``_State.phase``) → needs_context
    sur project_phase, pas de défaut « AVP ».

    Le classeur est ici sans phase, mais cela ne change plus rien : même avec une
    entête « AVP », le gabarit ne fournit pas la phase (elle nomme le fichier)."""
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx", phase=None)

    res = mcp_server.generate_avp_i3f_pack(controle_xlsx=ctrl, auditor="AMO BIM", export_pdf=False)

    assert res.get("status") == "needs_context"
    assert "project_phase" in res["missing"]
    q = next(q for q in res["questions"] if q["key"] == "project_phase")
    assert "aide_lecture_loi_mop" in q


def test_explicit_phase_param_used(_isolated):
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx", phase=None)
    res = mcp_server.generate_avp_i3f_pack(
        controle_xlsx=ctrl,
        phase="PRO",
        auditor="AMO BIM",
        export_pdf=False,
        project_name="Dieppe",
        project_code="7427L",
    )
    assert res.get("status") != "needs_context"
    assert res["phase"] == "PRO"


def test_confirmed_audit_phase_is_a_valid_source(_isolated):
    """``_State.phase`` a été confirmée par l'auditeur : ce n'est pas un repli.

    C'est la seule source admise en dehors du paramètre explicite.
    """
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    sess.phase = BIMPhase.DCE
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx", phase=None)
    res = mcp_server.generate_avp_i3f_pack(
        controle_xlsx=ctrl,
        auditor="AMO BIM",
        export_pdf=False,
        project_name="Dieppe",
        project_code="7427L",
    )
    assert res.get("status") != "needs_context"
    assert res["phase"] == "DCE"


def test_missing_code_asks(_isolated):
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx", esi="")
    res = mcp_server.generate_avp_i3f_pack(controle_xlsx=ctrl, auditor="AMO BIM", export_pdf=False)
    assert res.get("status") == "needs_context"
    assert "project_code" in res["missing"]


def test_auteur_controle_asked_when_missing(_isolated):
    """P2 : ni auteur_controle ni auditor fournis → demandé explicitement
    (pas de « AMO BIM » générique par défaut)."""
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx")  # nom/code/phase OK
    res = mcp_server.generate_avp_i3f_pack(controle_xlsx=ctrl, export_pdf=False)
    assert res.get("status") == "needs_context"
    assert "auditor_name" in res["missing"]  # clé = paramètre du tool


def test_auteur_controle_from_auditor(_isolated):
    """auditor fourni → pas de question auteur (auteur = auditor)."""
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx")
    res = mcp_server.generate_avp_i3f_pack(
        controle_xlsx=ctrl,
        auditor="CdP BIM 3F",
        export_pdf=False,
        project_name="Dieppe",
        project_code="7427L",
        # La phase vient du paramètre : l'entête AVP du gabarit MOA ne la
        # fournit plus, elle nomme le fichier client.
        phase="AVP",
    )
    assert res.get("status") != "needs_context"


def test_auteur_controle_bypass_with_confirm(_isolated):
    """confirm_context=True → génère malgré l'auteur manquant (repli AMO BIM)."""
    sess, tmp_path = _isolated
    _attach_minimal_snapshot(sess)
    ctrl = _controle_xlsx(tmp_path / "ctrl.xlsx")
    res = mcp_server.generate_avp_i3f_pack(
        controle_xlsx=ctrl,
        confirm_context=True,
        export_pdf=False,
        project_name="Dieppe",
        project_code="7427L",
        phase="AVP",
    )
    assert res.get("status") != "needs_context"


# ── project_context_questions : phase unique alignée sur le contrat ─────


def test_project_context_questions_single_phase_key(_isolated):
    """P2b : clé ``project_phase`` (pas ``phase``), aide loi MOP, pas de
    suggestion « PRO » codée en dur."""
    sess, _ = _isolated
    res = mcp_server.project_context_questions()
    phase_qs = [q for q in res["questions"] if q["key"] == "project_phase"]
    assert len(phase_qs) == 1
    q = phase_qs[0]
    assert "aide_lecture_loi_mop" in q
    assert "PRO (cas le plus fréquent" not in q.get("suggestion", "")
    # Plus d'ancienne clé "phase" côté question.
    assert not any(q["key"] == "phase" for q in res["questions"])
    assert "project_phase" in res["missing"]


def test_project_context_questions_proposes_detected_phase(_isolated):
    sess, _ = _isolated
    sess.snapshot = ModelSnapshot(
        project={"name": "X", "phase": "APD"}, model={"name": "M.ifc"}
    ).index()
    res = mcp_server.project_context_questions()
    q = next(q for q in res["questions"] if q["key"] == "project_phase")
    # APD (loi MOP) → proposition AVP.
    assert q.get("suggested_value") == "AVP"
