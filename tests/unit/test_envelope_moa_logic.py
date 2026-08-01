"""Logique MOA « Extraction surface enveloppe » : consommation d'envelope.json
(par_type), colonnes IFC OpenShell (sans Solibri), hors_filtre hors total."""

from __future__ import annotations

import json

import openpyxl
import pytest

from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.mcp import server as mcp_server
from audit_bim.mcp import tools_reporting as mcp_reporting
from audit_bim.mcp.session import _Session, current_session
from audit_bim.reporting.avp_i3f import (
    AvpMeta,
    _build_enveloppe_xlsx,
    write_avp_i3f_report_pack,
)
from audit_bim.reporting.avp_sources import AvpSources, read_envelope_json

# 8 types de façade sommant 2071,18 m² ; 2 types hors filtre (982,31) → total brut 3053,49.
_PAR_TYPE = [
    {"type": "ME_36", "etages": ["R+1", "R+2"], "net_side_area_m2": 300.0, "n": 40},
    {"type": "ME_30", "etages": "R+1", "netsidearea_m2": 300.0, "nombre": 38},
    {"type": "ME_25", "etages": "RDC", "netsidearea_m2": 300.0, "nombre": 35},
    {"type": "ME_20", "etages": "RDC", "netsidearea_m2": 300.0, "nombre": 30},
    {"type": "MR_vitre", "etages": "R+1", "netsidearea_m2": 250.0, "nombre": 12},
    {"type": "ME_pignon", "etages": "R+2", "netsidearea_m2": 250.0, "nombre": 20},
    {"type": "ME_soubassement", "etages": "RDC", "netsidearea_m2": 200.0, "nombre": 15},
    {"type": "ME_acrotere", "etages": "TOITURE", "netsidearea_m2": 171.18, "nombre": 8},
]
_ENVELOPE_DOC = {
    "par_type": _PAR_TYPE,
    "hors_filtre_type": [
        {"type": "MUR INT", "etages": "", "netsidearea_m2": 500.0, "nombre": 120},
        {"type": "MUR TECH", "etages": "", "netsidearea_m2": 482.31, "nombre": 30},
    ],
    "superficie_facades_m2": 2521.88,
    "superficie_facades_nette_m2": 2071.18,
    "superficie_calque_total_m2": 3053.49,
    "superficie_menuiseries_m2": 450.70,
    "superficie_menuiseries_fenetres_m2": 300.0,
    "superficie_menuiseries_portes_m2": 150.70,
    "shab_m2": 2164.98,
    "ratio_fac_shab": 0.9568,
    "seuil_i3f": 0.9,
}

_SOLIBRI_COLS = ("Surface Solibri", "Solibri Surface des Fenêtres", "Solibri Surface des Portes")
_OPENSHELL_COLS = (
    "Surface IFC OpenShell",
    "IFC OpenShell Surface des Fenêtres",
    "IFC OpenShell Surface des Portes",
)


def _write_json(tmp_path):
    p = tmp_path / "mn_bat_envelope.json"
    p.write_text(json.dumps(_ENVELOPE_DOC), encoding="utf-8")
    return str(p)


def _env_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return grid


def _all_text(grid):
    return "\n".join(str(c) for row in grid for c in row if c is not None)


# ── read_envelope_json ──────────────────────────────────────────────────


def test_read_envelope_json_one_row_per_type(tmp_path):
    src = read_envelope_json(_write_json(tmp_path))
    assert len(src.table.rows) == 8  # 8 lignes métier, pas 484
    assert src.sheet_title == "TDB 2022 04.2 - Extraction s..."
    assert all(r[0] == "Mur" for r in src.table.rows)
    by_type = {r[1]: r for r in src.table.rows}
    assert by_type["ME_36"][2] == "R+1, R+2"
    assert by_type["ME_36"][8] == 40
    d_sum = round(sum(r[3] for r in src.table.rows), 2)
    assert d_sum == pytest.approx(2071.18)  # Σ NetSideArea filtré
    assert src.ratio_fac_shab == pytest.approx(0.9568)
    assert len(src.hors_filtre_type) == 2  # diagnostic conservé à part


# ── onglet MOA généré ───────────────────────────────────────────────────


def _build(tmp_path):
    src = read_envelope_json(_write_json(tmp_path))
    out = tmp_path / "enveloppe.xlsx"
    _build_enveloppe_xlsx(out, AvpSources(enveloppe=src), AvpMeta(project_name="MN_BAT"))
    return _env_sheet(out)


def test_moa_columns_present_solibri_absent(tmp_path):
    text = _all_text(_build(tmp_path))
    for col in _OPENSHELL_COLS + ("Archicad BQ NetSideArea",):
        assert col in text, f"colonne MOA manquante : {col}"
    for col in _SOLIBRI_COLS:
        assert col not in text, f"colonne Solibri résiduelle : {col}"
    assert "Solibri" not in text


def test_eight_business_rows_and_d_sum(tmp_path):
    grid = _build(tmp_path)
    # 8 lignes métier (Composant = « Mur »), pas 484.
    assert sum(1 for row in grid for c in row if c == "Mur") == 8
    # Somme colonne D « Archicad BQ NetSideArea » ≈ 2071,18 (pas 3053,49).
    header_row = next(r for r in grid if "Archicad BQ NetSideArea" in [str(c) for c in r])
    d_idx = [str(c) for c in header_row].index("Archicad BQ NetSideArea")
    d_vals = [
        row[d_idx]
        for row in grid
        if row[0] == "Mur" and len(row) > d_idx and isinstance(row[d_idx], (int, float))
    ]
    assert round(sum(d_vals), 2) == pytest.approx(2071.18)
    assert 3053.49 not in d_vals


def test_synthesis_ratio_and_no_client_hors_filtre_note(tmp_path):
    text = _all_text(_build(tmp_path))
    assert "ratio FAC/SHAB" in text and "0.9568" in text
    assert "Superficie des façades" in text
    assert "écart :" in text
    assert "Seuil 3F 2026" in text
    # hors_filtre reste disponible dans la source, mais pas dans le classeur MOA client.
    assert "Hors filtre" not in text and "exclu du total façade" not in text


def test_tarare_layout_coordinates_and_formulas(tmp_path):
    src = read_envelope_json(_write_json(tmp_path))
    out = tmp_path / "enveloppe.xlsx"
    _build_enveloppe_xlsx(out, AvpSources(enveloppe=src), AvpMeta(project_name="MN_BAT"))
    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb.active
    try:
        assert ws.freeze_panes is None
        assert ws.max_row == 18
        assert ws["A1"].value == "Composant"
        assert ws["A2"].value == "Mur"
        assert ws["A10"].value is None
        assert ws["C11"].value == "Superficie des façades : "
        assert ws["D11"].value == "=SUM(D2:D10)"
        assert ws["E11"].value == "=SUM(E2:E10)"
        assert ws["F11"].value == "=SUM(F2:F9)"
        assert ws["D12"].value == "écart : "
        assert ws["E12"].value == "=E11/D11-1"
        assert ws["C14"].value == "Superficie des menuiseries : "
        assert ws["C16"].value == "SHAB : "
        assert ws["D17"].value == "=D11/D16"
        assert ws["C18"].value == "Seuil 3F 2026 : "
    finally:
        wb.close()


def test_envelope_json_keeps_priority_when_snapshot_exists(tmp_path):
    src = read_envelope_json(_write_json(tmp_path))
    snap = ModelSnapshot(
        elements=[
            {
                "uuid": "W1",
                "type": "IfcWall",
                "Name": "Mur elementaire",
                "layers": [{"name": "221 - MURS - Extérieurs périphériques.Exndo"}],
                "property_sets": [
                    {
                        "name": "Qto_WallBaseQuantities",
                        "properties": [{"definition": {"name": "NetSideArea"}, "value": 3053.49}],
                    }
                ],
            }
        ]
    ).index()
    pack = write_avp_i3f_report_pack(
        None,
        tmp_path / "pack",
        sources=AvpSources(enveloppe=src),
        snapshot=snap,
        project_name="MCP_Audit",
        project_code="0546L",
        phase="AVP",
        auditor="Stanislas Limouzi",
        date="260801",
        export_pdf=False,
    )
    wb = openpyxl.load_workbook(pack.enveloppe_xlsx, data_only=True)
    ws = wb.active
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    text = _all_text(grid)
    assert ws.title == "TDB 2022 04.2 - Extraction s..."
    assert "Layer" not in text
    assert "Surface IFC OpenShell (m²)" not in text
    assert "Surface IFC OpenShell" in text
    assert "BIMDATA" not in text
    assert sum(1 for row in grid for c in row if c == "Mur") == 8


def test_mcp_auto_uses_unique_envelope_json_from_input_dir(tmp_path, monkeypatch):
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    output_dir.mkdir()
    env_json = input_dir / "250613_MN_BAT_envelope.json"
    env_json.write_text(json.dumps(_ENVELOPE_DOC), encoding="utf-8")
    monkeypatch.setenv("AUDIT_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(output_dir))

    sess = _Session()
    sess.snapshot = ModelSnapshot(
        project={"name": "MCP_Audit"},
        model={"name": "250613_MN_BAT.ifc"},
        elements=[
            {
                "uuid": "W1",
                "type": "IfcWall",
                "Name": "Mur elementaire",
                "layers": [{"name": "221 - MURS - Extérieurs périphériques.Exndo"}],
                "property_sets": [
                    {
                        "name": "Qto_WallBaseQuantities",
                        "properties": [{"definition": {"name": "NetSideArea"}, "value": 3053.49}],
                    }
                ],
            }
        ],
    ).index()
    token = current_session.set(sess)
    try:
        res = mcp_server.generate_avp_i3f_pack(
            project_name="MCP_Audit",
            project_code="0546L",
            phase="PRO",
            auditor="Stanislas Limouzi",
            export_pdf=False,
        )
    finally:
        current_session.reset(token)

    assert res.get("status") != "needs_context"
    assert res["envelope_json_used"] == str(env_json)
    env_path = next(p for p in res["paths"] if "Extraction surface enveloppe" in p)
    wb = openpyxl.load_workbook(env_path, data_only=True)
    ws = wb.active
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    assert ws.title == "TDB 2022 04.2 - Extraction s..."
    assert sum(1 for row in grid for c in row if c == "Mur") == 8


def test_mcp_auto_uses_local_envelope_json_when_input_dir_unset(tmp_path, monkeypatch):
    input_dir = tmp_path / "audit_in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    output_dir.mkdir()
    env_json = input_dir / "250613_MN_BAT_envelope.json"
    env_json.write_text(json.dumps(_ENVELOPE_DOC), encoding="utf-8")
    monkeypatch.delenv("AUDIT_INPUT_DIR", raising=False)
    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(output_dir))
    monkeypatch.setattr(mcp_reporting, "_auto_envelope_roots", lambda: [input_dir])

    sess = _Session()
    sess.snapshot = ModelSnapshot(
        project={"name": "MCP_Audit"},
        model={"name": "250613_MN_BAT.ifc"},
        elements=[
            {
                "uuid": "W1",
                "type": "IfcWall",
                "Name": "Mur elementaire",
                "layers": [{"name": "221 - MURS - Extérieurs périphériques.Exndo"}],
                "property_sets": [
                    {
                        "name": "Qto_WallBaseQuantities",
                        "properties": [{"definition": {"name": "NetSideArea"}, "value": 3053.49}],
                    }
                ],
            }
        ],
    ).index()
    token = current_session.set(sess)
    try:
        res = mcp_server.generate_avp_i3f_pack(
            project_name="MCP_Audit",
            project_code="0546L",
            phase="PRO",
            auditor="Stanislas Limouzi",
            export_pdf=False,
        )
    finally:
        current_session.reset(token)

    assert res.get("status") != "needs_context"
    assert res["envelope_json_used"] == str(env_json)
    env_path = next(p for p in res["paths"] if "Extraction surface enveloppe" in p)
    wb = openpyxl.load_workbook(env_path, data_only=True)
    ws = wb.active
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    assert ws.title == "TDB 2022 04.2 - Extraction s..."
    assert sum(1 for row in grid for c in row if c == "Mur") == 8
