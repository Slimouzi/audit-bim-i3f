"""Pack AVP **sans sources I3F** : extraction depuis le snapshot d'audit.

Couvre l'exigence produit : si les fichiers I3F sont absents, générer SHAB,
Zones/Espaces, Enveloppe et Menuiseries depuis ``AuditResult.snapshot`` — et
interdire un livrable client réduit au seul bandeau (QA gate).
"""

from __future__ import annotations

import openpyxl
import pytest

from audit_bim.audit.engine import AuditResult
from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.reporting import avp_i3f
from audit_bim.reporting.avp_i3f import AvpQaError, write_avp_i3f_report_pack
from audit_bim.reporting.avp_sources import AvpSources
from audit_bim.requirements.models import BIMPhase, RequirementsCatalog


def _catalog() -> RequirementsCatalog:
    return RequirementsCatalog(
        cch_version="3.6",
        cch_source_pdf="/tmp/cch.pdf",
        data_spec_source="/tmp/data.xlsx",
        naming_spec_source="/tmp/naming.xlsx",
        properties=[],
        naming_rules=[],
        storey_names=[],
        zone_specs=[],
        room_specs=[],
    )


def _synthetic_result() -> AuditResult:
    """Snapshot minimal : 1 mur d'enveloppe (layer cible, casse/accents/espaces
    variés), 1 espace LongName='' / Name='CHAMBRE', 1 fenêtre — surfaces
    uniquement en « Superficie calculée » (pas de BaseQuantities)."""
    wall = {
        "uuid": "W1",
        "type": "IfcWall",
        "name": "Mur ext 1",
        # Variante tolérante du layer cible « MURS - Extérieurs périphériques.Exnd ».
        "layers": [{"name": "murs  -  Exterieurs Periphériques.exnd"}],
        "property_sets": [
            {
                "name": "Pset_WallCommon",
                "properties": [{"definition": {"name": "Superficie calculée"}, "value": 42.5}],
            }
        ],
    }
    # Mur-rideau : NE DOIT PAS être compté dans l'enveloppe (décision explicite).
    curtain = {
        "uuid": "CW1",
        "type": "IfcCurtainWall",
        "name": "Façade vitrée",
        "layers": [{"name": "MURS - Extérieurs périphériques.Exnd"}],
        "property_sets": [
            {
                "name": "Pset_CurtainWallCommon",
                "properties": [{"definition": {"name": "Superficie calculée"}, "value": 999.0}],
            }
        ],
    }
    window = {
        "uuid": "WIN1",
        "type": "IfcWindow",
        "name": "F25",
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [
                    {"definition": {"name": "Width"}, "value": 0.6},
                    {"definition": {"name": "Height"}, "value": 1.3},
                ],
            }
        ],
    }
    space = {
        "uuid": "S1",
        "type": "IfcSpace",
        "name": "CHAMBRE",
        "longname": "",  # vide → le libellé exporté doit reprendre Name
        "property_sets": [
            {
                "name": "Pset_SpaceCommon",
                "properties": [{"definition": {"name": "Superficie calculée"}, "value": 12.98}],
            }
        ],
    }
    snap = ModelSnapshot(
        project={"name": "Programme"},
        model={"name": "M.ifc"},
        spaces=[space],
        elements=[wall, curtain, window],
    ).index()
    return AuditResult(phase=BIMPhase.AVP, catalog=_catalog(), snapshot=snap, findings=[])


def _all_text(path) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is not None:
                    parts.append(str(c))
    wb.close()
    return "\n".join(parts)


# ── Extraction snapshot (source absente) ────────────────────────────────


def test_snapshot_fallback_fills_annexes(tmp_path):
    result = _synthetic_result()
    # sources=None → aucune source I3F ; l'extraction snapshot doit prendre
    # le relais (sinon la QA gate lèverait).
    pack = write_avp_i3f_report_pack(
        result, tmp_path / "out", sources=None, project_name="X", project_code="Y", export_pdf=False
    )

    shab = _all_text(pack.shab_xlsx)
    zones = _all_text(pack.zones_espaces_xlsx)
    env = _all_text(pack.enveloppe_xlsx)

    # SHAB / Zones : l'espace apparaît, libellé repris de Name (LongName vide).
    assert "CHAMBRE" in shab
    assert "12.98" in shab
    assert "CHAMBRE" in zones
    # Surface issue de « Superficie calculée » (tracée).
    assert "Superficie calculée" in shab

    # Enveloppe : le mur cible apparaît avec sa surface et la source tracée.
    assert "Mur ext 1" in env
    assert "42.5" in env
    assert "Superficie calculée" in env


def _duplex_result() -> AuditResult:
    """Duplex : une zone « Logement Duplex A101 » traversant deux étages
    (R+1 / R+2), avec un séjour au R+1 et une chambre au R+2."""
    sejour = {
        "uuid": "SP-SEJOUR",
        "type": "IfcSpace",
        "name": "SEJOUR",
        "longname": "Séjour",
        "storey": {"uuid": "ST1", "name": "R+1"},
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": 24.0}],
            }
        ],
    }
    chambre = {
        "uuid": "SP-CH",
        "type": "IfcSpace",
        "name": "CHAMBRE 01",
        "longname": "",  # vide → libellé repris de Name
        "storey": {"uuid": "ST2", "name": "R+2"},
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": 12.0}],
            }
        ],
    }
    storeys = [{"uuid": "ST1", "name": "R+1"}, {"uuid": "ST2", "name": "R+2"}]
    zone = {
        "uuid": "Z-A101",
        "type": "IfcZone",
        "name": "Logement Duplex A101",
        "spaces": ["SP-SEJOUR", "SP-CH"],  # zone traversant les 2 étages
    }
    snap = ModelSnapshot(
        project={"name": "Programme"},
        model={"name": "M.ifc"},
        storeys=storeys,
        spaces=[sejour, chambre],
        zones=[zone],
    ).index()
    return AuditResult(phase=BIMPhase.AVP, catalog=_catalog(), snapshot=snap, findings=[])


def test_shab_export_has_zone_and_storey_columns(tmp_path):
    result = _duplex_result()
    pack = write_avp_i3f_report_pack(
        result, tmp_path / "out", sources=None, project_name="X", project_code="Y", export_pdf=False
    )
    wb = openpyxl.load_workbook(pack.shab_xlsx, data_only=True)
    ws = wb.active
    # Localise l'en-tête et vérifie la présence des colonnes Zone + Étage.
    header = None
    for row in ws.iter_rows(values_only=True):
        if row and "Composant" in [c for c in row if c is not None]:
            header = [c for c in row]
            break
    assert header is not None
    assert "Zone" in header
    assert "Étage" in header
    zone_col = header.index("Zone")
    storey_col = header.index("Étage")

    # Collecte les lignes de données (espaces).
    data = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[1] in ("Séjour", "CHAMBRE 01"):
            data[row[1]] = (row[zone_col], row[storey_col])
    wb.close()

    # Les deux pièces portent la même zone (duplex) et leur étage respectif.
    assert data["Séjour"] == ("Logement Duplex A101", "R+1")
    assert data["CHAMBRE 01"] == ("Logement Duplex A101", "R+2")


def test_shab_space_multiple_storeys_joined(tmp_path):
    """Un espace rattaché à deux étages (cas duplex au niveau pièce) →
    les deux étages sont listés (séparés par « / »)."""
    space = {
        "uuid": "SP-DUP",
        "type": "IfcSpace",
        "name": "SEJOUR DUPLEX",
        "storeys": [{"uuid": "ST1", "name": "R+1"}, {"uuid": "ST2", "name": "R+2"}],
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": 40.0}],
            }
        ],
    }
    snap = ModelSnapshot(
        project={"name": "P"},
        model={"name": "M.ifc"},
        storeys=[{"uuid": "ST1", "name": "R+1"}, {"uuid": "ST2", "name": "R+2"}],
        spaces=[space],
    ).index()
    result = AuditResult(phase=BIMPhase.AVP, catalog=_catalog(), snapshot=snap, findings=[])
    pack = write_avp_i3f_report_pack(
        result, tmp_path / "out", sources=None, project_name="X", project_code="Y", export_pdf=False
    )
    txt = _all_text(pack.shab_xlsx)
    assert "R+1 / R+2" in txt


def test_envelope_excludes_curtain_wall(tmp_path):
    result = _synthetic_result()
    pack = write_avp_i3f_report_pack(
        result, tmp_path / "out", sources=None, project_name="X", project_code="Y", export_pdf=False
    )
    env = _all_text(pack.enveloppe_xlsx)
    # Le mur-rideau (999) ne doit pas être compté dans l'enveloppe.
    assert "999" not in env
    assert "Façade vitrée" not in env


def test_menuiseries_from_snapshot(tmp_path):
    result = _synthetic_result()
    pack = write_avp_i3f_report_pack(
        result, tmp_path / "out", sources=None, project_name="X", project_code="Y", export_pdf=False
    )
    men = _all_text(pack.menuiseries_xlsx)
    assert "F25" in men  # fenêtre extraite
    # Surface = L×H (0.6×1.3 = 0.78) tracée.
    assert "0.78" in men


# ── QA gate anti-livrable vide ──────────────────────────────────────────


def test_qa_gate_raises_when_extraction_empty(tmp_path, monkeypatch):
    """Snapshot avec espaces/murs mais extraction vide (forcée) → AvpQaError."""
    result = _synthetic_result()
    # On neutralise l'extraction snapshot : les annexes sortiront vides
    # alors que le snapshot contient des données → la QA gate doit lever.
    monkeypatch.setattr(avp_i3f, "build_sources_from_snapshot", lambda snap: AvpSources())

    with pytest.raises(AvpQaError) as exc:
        write_avp_i3f_report_pack(
            result,
            tmp_path / "out",
            sources=None,
            project_name="X",
            project_code="Y",
            export_pdf=False,
        )
    # Les 3 exports concernés sont signalés.
    assert set(exc.value.empty) >= {"SHAB", "Zones/Espaces", "Enveloppe"}


def test_pack_without_snapshot_no_gate(tmp_path):
    """result=None (pas de snapshot) → la QA gate ne se déclenche pas
    (rien d'exploitable à comparer)."""
    pack = write_avp_i3f_report_pack(
        None, tmp_path / "out", sources=None, project_name="X", project_code="Y", export_pdf=False
    )
    # Génère quand même le pack (bandeau + NOT_AVAILABLE), sans lever.
    assert pack.shab_xlsx.exists()


def test_tool_returns_error_status_on_empty(tmp_path, monkeypatch):
    """Le tool MCP renvoie un statut d'erreur explicite (pas un fichier vide)
    quand la QA gate échoue."""
    from audit_bim.mcp import server as mcp_server
    from audit_bim.mcp.session import _Session, current_session

    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    monkeypatch.setattr(avp_i3f, "build_sources_from_snapshot", lambda snap: AvpSources())

    sess = _Session()
    sess.result = _synthetic_result()
    token = current_session.set(sess)
    try:
        res = mcp_server.generate_avp_i3f_pack(
            project_name="X", project_code="Y", phase="AVP", export_pdf=False
        )
    finally:
        current_session.reset(token)

    assert res.get("status") == "error"
    assert res.get("error") == "empty_deliverable"
    assert "SHAB" in res["empty_deliverables"]
