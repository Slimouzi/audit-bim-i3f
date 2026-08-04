"""L'annexe Enveloppe ne doit jamais s'ouvrir en erreur chez le client.

Sans aucune ligne de données, le bloc synthèse remonte au contact de l'entête :
``summary_row`` vaut 2 (ligne Excel 3), et la plage ``D2:D3`` du total englobe la
cellule qui la porte. Excel signale alors une **référence circulaire** à
l'ouverture, puis propage ``#VALEUR!`` dans l'écart et le ratio FAC/SHAB —
l'indicateur que le client lit en premier.

Le cas nominal (8 lignes, disposition Tarare) doit rester **strictement**
inchangé : ces coordonnées et ces formules sont le livrable de référence.
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from audit_bim.profiles.i3f.tools_reporting import generate_avp_i3f_pack as tr_generate_avp_i3f_pack
from audit_bim.reporting.avp.models import AvpMeta
from audit_bim.reporting.avp.xlsx_enveloppe import _build_enveloppe_xlsx
from audit_bim.reporting.avp_sources import (
    ENVELOPPE_MOA_HEADERS,
    AvpSources,
    EnveloppeSource,
    SheetTable,
)
from audit_bim.reporting.word_report import NOT_AVAILABLE


@pytest.fixture
def meta():
    return AvpMeta(project_name="Dieppe", project_code="7427L", phase="APD")


def _table(n):
    return SheetTable(
        title="TDB 2022 04.2 - Extraction s...",
        headers=list(ENVELOPPE_MOA_HEADERS),
        rows=[
            ["Mur", f"ME_{i}", "RDC", 100.0 + i, 100.0 + i, 5.0, 3.0, 2.0, 4, None]
            for i in range(n)
        ],
    )


def _formules(ws):
    return {
        c.coordinate: c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    }


def _construire(tmp_path, meta, src):
    chemin = tmp_path / "enveloppe.xlsx"
    _build_enveloppe_xlsx(chemin, AvpSources(enveloppe=src), meta)
    return load_workbook(chemin).active


def test_empty_envelope_writes_no_circular_formula(tmp_path, meta):
    """Zéro ligne : aucune formule, donc aucune plage ne peut s'auto-référencer."""
    ws = _construire(tmp_path, meta, None)

    assert _formules(ws) == {}
    assert ws["D3"].value == NOT_AVAILABLE
    assert ws["E3"].value == NOT_AVAILABLE
    assert ws["F3"].value == NOT_AVAILABLE


def test_empty_envelope_writes_no_ecart_formula(tmp_path, meta):
    """``=E/D-1`` sans total de façade rendrait #DIV/0!."""
    ws = _construire(tmp_path, meta, None)

    assert ws["E4"].value == NOT_AVAILABLE


def test_empty_envelope_without_shab_writes_no_ratio_formula(tmp_path, meta):
    """Diviser par la cellule SHAB, qui porte alors du texte, rendrait #VALEUR!."""
    ws = _construire(tmp_path, meta, None)

    assert ws["D8"].value == NOT_AVAILABLE  # SHAB
    assert ws["D9"].value == NOT_AVAILABLE  # ratio FAC/SHAB


def test_rows_without_shab_still_drops_only_the_ratio(tmp_path, meta):
    """Une SHAB absente ne doit pas emporter les totaux, qui restent calculables."""
    ws = _construire(tmp_path, meta, EnveloppeSource(table=_table(8), shab=None))

    assert ws["D11"].value == "=SUM(D2:D10)"
    assert ws["E12"].value == "=E11/D11-1"
    assert ws["D17"].value == NOT_AVAILABLE


def test_tarare_layout_formulas_are_unchanged(tmp_path, meta):
    """Disposition de référence : coordonnées ET formules figées.

    Ce test est le garde-fou anti-régression du livrable client : toute
    modification du builder qui déplacerait le bloc synthèse le casse.
    """
    ws = _construire(tmp_path, meta, EnveloppeSource(table=_table(8), shab=2164.68))

    assert ws["D11"].value == "=SUM(D2:D10)"
    assert ws["E11"].value == "=SUM(E2:E10)"
    assert ws["E12"].value == "=E11/D11-1"
    assert ws["D17"].value == "=D11/D16"
    assert ws["D16"].value == 2164.68


def _mur(uuid, ifc_type="IfcWall", *, layers=None):
    el = {
        "uuid": uuid,
        "type": ifc_type,
        "name": "Mur de base:MUR ENDUIT 20 mm",
        "property_sets": [
            {
                "name": "Qto_WallBaseQuantities",
                "properties": [{"definition": {"name": "NetSideArea"}, "value": 900.13}],
            }
        ],
    }
    if layers is not None:
        el["layers"] = layers
    return el


def _generer_avec(tmp_path, monkeypatch, elements):
    from audit_bim.extraction.model_data import ModelSnapshot
    from audit_bim.mcp.session import _Session, current_session
    from audit_bim.reporting import avp_i3f

    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    monkeypatch.setattr(avp_i3f, "build_sources_from_snapshot", lambda snap: AvpSources())

    sess = _Session()
    sess.snapshot = ModelSnapshot(
        project={"name": "MCP_Audit"},
        model={"name": "DIEPPE-7427L.ifc"},
        elements=elements,
    ).index()
    token = current_session.set(sess)
    try:
        return tr_generate_avp_i3f_pack(
            project_name="Dieppe",
            project_code="7427L",
            phase="APD",
            auditor_name="Stanislas Limouzi",
            auto_compute_envelope=False,
            export_pdf=False,
        )
    finally:
        current_session.reset(token)


def test_empty_envelope_annex_is_refused_on_revit_without_layer(tmp_path, monkeypatch):
    """Le cas réel : export **Revit**, aucun calque, annexe Enveloppe vide.

    ``count_envelope_walls`` reconnaît l'enveloppe au calque ArchiCAD ; sur une
    maquette Revit il rend 0, et la QA gate se taisait — laissant sortir un pack
    « OK » dont l'annexe Enveloppe était vide. C'est exactement le trou que ce
    test verrouille : **pas de clé ``layers`` du tout** sur le mur.
    """
    res = _generer_avec(tmp_path, monkeypatch, [_mur("W1")])

    assert res.get("status") == "error"
    assert res.get("error") == "empty_deliverable"
    assert "Enveloppe" in res["empty_deliverables"]
    assert res.get("needs_envelope_source") is True
    assert "envelope_json" in res["next_step"]
    assert "envelope_type_pattern" in res["next_step"]


def test_empty_envelope_annex_is_refused_on_archicad_with_layer(tmp_path, monkeypatch):
    """Le chemin ArchiCAD historique reste couvert, en plus du cas Revit."""
    res = _generer_avec(
        tmp_path,
        monkeypatch,
        [_mur("W1", layers=[{"name": "221 - MURS - Extérieurs périphériques.Exndo"}])],
    )

    assert res.get("status") == "error"
    assert res.get("error") == "empty_deliverable"
    assert "Enveloppe" in res["empty_deliverables"]


def test_empty_envelope_annex_is_refused_on_curtain_wall(tmp_path, monkeypatch):
    """Une enveloppe en mur-rideau est de la façade : même refus."""
    res = _generer_avec(tmp_path, monkeypatch, [_mur("CW1", "IfcCurtainWall")])

    assert res.get("status") == "error"
    assert "Enveloppe" in res["empty_deliverables"]


def test_model_without_any_wall_does_not_demand_an_envelope_annex(tmp_path, monkeypatch):
    """Sans aucun mur, l'annexe Enveloppe n'a pas lieu d'être — pas de faux refus."""
    res = _generer_avec(tmp_path, monkeypatch, [{"uuid": "S1", "type": "IfcSlab", "name": "Dalle"}])

    assert "Enveloppe" not in (res.get("empty_deliverables") or [])
