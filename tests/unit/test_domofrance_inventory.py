"""Inventaire Domofrance — la forme est éprouvée en CI, le texte ne l'est pas.

Le classeur du maître d'ouvrage vit hors du dépôt. Sans précaution, tous les
tests qui en dépendent seraient ignorés en CI, et une suite verte ne prouverait
aucun des compteurs que le scope prétend figer — c'est exactement ce qui était
arrivé sur le gabarit MRN.

La suite est donc coupée en deux, et la coupure est explicite :

- **[CI] — éprouvé sans le fichier client.** La structure (413 lignes, 286
  distinctes, duplication, tables de surfaces) est rejouée depuis une fixture de
  *forme* versionnée, extraite du vrai classeur et non retapée : chaque libellé
  y est remplacé par un index, ce qui conserve la duplication et les décomptes
  de valeurs distinctes sans emporter une seule phrase du client. Le mécanisme
  de détection lexicale est éprouvé sur des phrases écrites ici.
- **[CLASSEUR] — ignoré sans le fichier client.** Les *agrégats* lexicaux
  (99 ``needs_bbox``, 77 ``manual_only``…) dépendent du texte réel et ne peuvent
  pas être reconstitués depuis une forme anonymisée. Ils restent sous
  ``needs_workbook``, et ce module le dit plutôt que de le laisser croire
  couvert.

Les deux marqueurs se lisent tels quels dans ``docs/scope-domofrance-controls.md``,
en face de chaque compteur.
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "scripts"))

import inventory_domofrance_controls as inv  # noqa: E402

#: Le classeur du client, hors dépôt. Surchargeable pour ne pas figer le poste
#: de celui qui a écrit le test.
WORKBOOK = os.environ.get(
    "AUDIT_BIM_DOMOFRANCE_WORKBOOK",
    "/Users/stani/code/MCP/Documents maître d'ouvrage/Documents Domofrance/"
    "Données d'entrée/Liste de contrôle.xlsx",
)

needs_workbook = pytest.mark.skipif(
    not Path(WORKBOOK).is_file(),
    reason=(
        "[CLASSEUR] agrégat dépendant du texte réel du maître d'ouvrage — "
        "non reconstituable depuis la forme anonymisée"
    ),
)

SHAPE = "domofrance_controls_shape.json"


def _shape() -> dict:
    return json.loads((Path(__file__).parent / SHAPE).read_text(encoding="utf-8"))


def _build_synthetic_workbook(path: Path) -> Path:
    """Reconstruit un .xlsx depuis la forme anonymisée.

    On passe par un vrai fichier plutôt que par un faux objet feuille : c'est
    ``parse_controls`` lui-même — bornes de lignes, indices de colonnes, saut
    des lignes vides — qui doit être éprouvé, pas une imitation de sa lecture.
    """
    import openpyxl

    shape = _shape()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = inv.CONTROL_SHEET
    for col, name in zip(inv.CONTROL_COLUMN_INDEXES, shape["control_columns"], strict=True):
        sheet.cell(inv.CONTROL_HEADER_ROW, col, name)
    for entry in shape["controls"]:
        for col, value in zip(inv.CONTROL_COLUMN_INDEXES, entry["values"], strict=True):
            if value:
                sheet.cell(entry["row"], col, value)

    surface = wb.create_sheet(inv.SURFACE_SHEET)
    for table in shape["surface"].values():
        first = table["first_col"]
        surface.cell(3, first, table["title_row3"])
        for offset, header in enumerate(table["header"]):
            surface.cell(inv.SURFACE_HEADER_ROW, first + offset, header)
        for index, line in enumerate(table["rows"]):
            for offset, value in enumerate(line):
                if value:
                    surface.cell(inv.SURFACE_FIRST_ROW + index, first + offset, value)

    target = path / "domofrance_shape.xlsx"
    wb.save(target)
    return target


# --------------------------------------------------------------------------
# [CI] Normalisation — éprouvée sur des phrases écrites ici.
# --------------------------------------------------------------------------


def test_apostrophe_typographique_laisse_un_espace():
    """Le bug qui a motivé l'ordre des opérations dans ``_normalize``."""
    assert " d acces " in inv._normalize("rampes d’accès")


def test_normalisation_retire_les_accents_et_la_casse():
    assert inv._normalize("HAUTEUR Sous Plafond").strip() == "hauteur sous plafond"


def test_normalisation_borne_le_texte_par_des_espaces():
    """Sans bornes, un motif comme ``" cm "`` raterait un texte qui finit par là."""
    normalized = inv._normalize("17 cm")
    assert normalized.startswith(" ") and normalized.endswith(" ")
    assert " cm " in normalized


# --------------------------------------------------------------------------
# [CI] Signaux lexicaux — le mécanisme, pas les agrégats.
# --------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("phrase", "attendu"),
    (
        ("La largeur de passage est de 0,90 m", "needs_bbox"),
        ("Ne pas constituer d'obstacle au sens des règles", "needs_collision"),
        ("Vérifiez la présence d'un équipement dans le local", "needs_space_context"),
        ("Il est recommandé de positionner les accès côté rue", "manual_only"),
    ),
)
def test_chaque_signal_reconnait_sa_famille(phrase, attendu):
    assert attendu in inv._detect_signals(inv._normalize(phrase))


def test_un_controle_peut_ne_porter_aucun_signal():
    """Les familles ne sont ni exhaustives ni exclusives — le scope le dit."""
    assert inv._detect_signals(inv._normalize("Le lot est attribué au titulaire")) == frozenset()


@pytest.mark.parametrize(
    "phrase",
    ("largeur de 0,90 m", "hauteur de 17 cm", "superficie de 8 m2", "tous les 20 mètres"),
)
def test_seuil_chiffre_reconnu(phrase):
    assert inv.NUMERIC_THRESHOLD.search(inv._normalize(phrase))


@pytest.mark.parametrize("phrase", ("trois niveaux au maximum", "format A2", "30% des boîtes"))
def test_nombre_sans_unite_n_est_pas_un_seuil(phrase):
    """Un nombre n'est un seuil que s'il porte une unité dimensionnelle."""
    assert inv.NUMERIC_THRESHOLD.search(inv._normalize(phrase)) is None


# --------------------------------------------------------------------------
# [CI] Structure du classeur — rejouée depuis la forme anonymisée.
# --------------------------------------------------------------------------


@pytest.fixture(scope="module")
def synthetic(tmp_path_factory) -> list[inv.Control]:
    workbook = _build_synthetic_workbook(tmp_path_factory.mktemp("domofrance"))
    return inv.parse_controls(str(workbook))


def test_nombre_de_lignes_et_bornes(synthetic):
    assert len(synthetic) == 413
    assert synthetic[0].row == 4
    assert synthetic[-1].row == 416


def test_lignes_distinctes_et_doublons(synthetic):
    from collections import Counter

    groups = {i: n for i, n in Counter(c.identity for c in synthetic).items() if n > 1}
    assert len({c.identity for c in synthetic}) == 286
    assert len(groups) == 82
    assert sum(n - 1 for n in groups.values()) == 127
    assert sum(groups.values()) == 209
    assert max(groups.values()) == 7


def test_un_controle_existe_sous_deux_types_de_logement(synthetic):
    """286 identités, mais 285 contrôles : un seul est écrit sous deux types."""
    assert len({c.identity_without_type for c in synthetic}) == 285


def test_decomptes_par_colonne(synthetic):
    assert len({c.type_logement for c in synthetic}) == 3
    assert len({c.zone for c in synthetic}) == 57
    assert len({c.element for c in synthetic}) == 101
    assert len({c.verification for c in synthetic}) == 198
    assert len({c.description for c in synthetic}) == 251


def test_les_doublons_ne_sont_jamais_ecrases(synthetic):
    """Le parseur rend les lignes du client, pas une vue dédoublonnée."""
    assert len(synthetic) == 413 != len({c.identity for c in synthetic})
    assert len(inv.distinct_controls(synthetic)) == 286


# --------------------------------------------------------------------------
# [CI] Tables de surfaces — lues sur la forme.
# --------------------------------------------------------------------------


@pytest.fixture(scope="module")
def synthetic_tables(tmp_path_factory) -> list[inv.SurfaceTable]:
    workbook = _build_synthetic_workbook(tmp_path_factory.mktemp("domofrance_surface"))
    return inv.parse_surface_tables(str(workbook))


def test_tables_de_surfaces_lues_sur_la_forme(synthetic_tables):
    collectif, individuel = synthetic_tables
    assert collectif.label == "LOGEMENT COLLECTIF"
    assert collectif.typologies == ("T1bis", "T2", "T3", "T4", "T5", "T6")
    assert len(collectif.room_types) == 13
    assert collectif.has_width_column
    assert collectif.numeric_cells == 53

    assert individuel.label == "LOGEMENT INDIVIDUEL"
    assert individuel.typologies == ("T3", "T4", "T5")
    assert len(individuel.room_types) == 13
    assert individuel.numeric_cells == 27


def test_la_ligne_total_n_est_pas_un_type_de_piece(synthetic_tables):
    """Compter « Total » parmi les pièces ferait 14 pièces au lieu de 13."""
    for table in synthetic_tables:
        assert table.total_row == 19
        assert not any(name.lower().startswith("total") for name in table.room_types)


# --------------------------------------------------------------------------
# [CLASSEUR] Agrégats dépendant du texte réel — ignorés sans le fichier client.
# --------------------------------------------------------------------------


@pytest.fixture(scope="module")
def real() -> list[inv.Control]:
    return inv.parse_controls(WORKBOOK)


@needs_workbook
def test_agregats_lexicaux_sur_le_classeur_reel(real):
    counts = {name: sum(1 for c in real if name in c.signals) for name in inv.SIGNALS}
    assert counts == {
        "needs_bbox": 99,
        "needs_collision": 15,
        "needs_space_context": 287,
        "manual_only": 77,
    }
    assert sum(1 for c in real if c.needs_geometry) == 331


@needs_workbook
def test_noyau_outillable_du_scope(real):
    """Le chiffre à annoncer au client : 30 sur 286, pas 331 sur 413.

    L'écart entre les deux est tout l'enjeu du lot. ``needs_geometry`` sature à
    80 % parce que « présence » et « accès » attrapent presque tout le classeur ;
    la conjonction grandeur + seuil + absence d'appréciation retombe à 10 %.
    """
    distinct = inv.distinct_controls(real)
    core = inv.tooling_core(distinct)
    assert len(distinct) == 286
    assert len(core) == 30
    assert len(core) / len(distinct) < 0.15


@needs_workbook
def test_legende_indicative_de_la_table_collectif():
    """La table collectif est légendée « à titre indicatif ».

    Aucun verdict de conformité ne peut donc être rendu sur ces surfaces. La
    légende est absente de la fixture de forme — c'est une phrase du client —
    donc ce test ne peut exister qu'ici.
    """
    collectif, _ = inv.parse_surface_tables(WORKBOOK)
    assert "titre indicatif" in collectif.caption.lower()
    assert "souhaitable" in collectif.caption.lower()
