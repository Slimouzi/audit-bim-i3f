"""Structure du classeur Excel pilotée par le profil — I3F figé à l'octet près.

Un nom d'onglet n'est pas une phrase : c'est une **clé technique**. Un TCD, une
macro ou un rapprochement côté maître d'ouvrage peuvent le référencer par son
nom, et le changer casse un usage aval sans que rien n'échoue de notre côté.
D'où deux exigences opposées, toutes deux testées ici :

- I3F conserve « Référentiel I3F » et « Référence CCH » **exactement** ;
- un profil tiers obtient ses propres libellés, sans héritage.
"""

from __future__ import annotations

from dataclasses import replace

import pytest
from openpyxl import load_workbook

import audit_bim.profiles.registry as reg
from audit_bim.audit.engine import AuditResult
from audit_bim.audit.findings import ErrorType, Finding, Severity, Theme
from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.profiles import get_profile
from audit_bim.profiles.models import ReportStructureSpec
from audit_bim.reporting import xlsx_annex
from audit_bim.requirements.models import BIMPhase

I3F_SHEET_NAME = "Référentiel I3F"
I3F_COLUMN_LABEL = "Référence CCH"


class _Catalog:
    cch_version = "3.6"
    cch_source_pdf = "cch.pdf"
    data_spec_source = "spec.xlsx"
    naming_spec_source = "nom.xlsx"
    properties: list = []
    naming_rules: list = []
    storey_names: list = []
    zone_specs: list = []
    room_specs: list = []


def _result() -> AuditResult:
    snap = ModelSnapshot()
    snap.project = {"name": "P"}
    snap.model = {"name": "M"}
    combos = [
        (Theme.CLASSIFICATION, ErrorType.CLASSIFICATION_MISSING),
        (Theme.NAMING_ZONE, ErrorType.NAMING_MISSING),
    ]
    findings = [
        Finding(
            element_uuid=f"u{i}",
            ifc_type="IfcWall",
            name=f"W{i}",
            theme=th,
            error_type=et,
            severity=Severity.HIGH,
            expected="x",
            actual=None,
            ref_cch="6.3",
        )
        for i, (th, et) in enumerate(combos)
    ]
    return AuditResult(snapshot=snap, catalog=_Catalog(), phase=BIMPhase.PRO, findings=findings)


def _workbook(tmp_path, monkeypatch, *, profile_id=None):
    """Génère le classeur et le relit avec openpyxl (pas xlsxwriter)."""
    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    out = tmp_path / "annex.xlsx"
    xlsx_annex.write_xlsx_annex(_result(), out, profile_id=profile_id)
    assert out.is_file() and out.stat().st_size > 0
    return load_workbook(out)


# ── 1. I3F : les deux chaînes historiques, exactement ─────────────────


def test_i3f_referential_sheet_name_is_exact(tmp_path, monkeypatch):
    assert I3F_SHEET_NAME in _workbook(tmp_path, monkeypatch).sheetnames


def test_i3f_reference_column_label_is_exact(tmp_path, monkeypatch):
    wb = _workbook(tmp_path, monkeypatch)
    headers = [c.value for c in wb["Findings (tous)"][1]]
    assert I3F_COLUMN_LABEL in headers


def test_i3f_sheet_name_is_declared_not_composed():
    """Le profil porte le nom EXACT — pas « Référentiel CCH BIM I3F »."""
    spec = get_profile("i3f").report_structure
    assert spec.referential_sheet_name == I3F_SHEET_NAME
    framework = get_profile("i3f").reference_framework
    assert spec.referential_sheet_name != f"Référentiel {framework.name}"


def test_i3f_result_is_identical_with_or_without_explicit_profile(tmp_path, monkeypatch):
    """Appel sans `profile_id` ou avec « i3f » : même classeur."""
    implicit = _workbook(tmp_path / "a", monkeypatch)
    explicit = _workbook(tmp_path / "b", monkeypatch, profile_id="i3f")
    assert implicit.sheetnames == explicit.sheetnames
    for name in implicit.sheetnames:
        assert [c.value for c in implicit[name][1]] == [c.value for c in explicit[name][1]]


def test_column_template_keeps_order_and_widths():
    """Seul le libellé de référence change ; le gabarit reste figé."""
    base = xlsx_annex.COLUMNS
    for profile_id in (None, "i3f", "bim_in_motion"):
        resolved = xlsx_annex._columns_for(profile_id)
        assert len(resolved) == len(base)
        assert [w for _, w in resolved] == [w for _, w in base]


# ── 2. Profil tiers : aucun héritage I3F ──────────────────────────────

FORBIDDEN = ("I3F", "CCH")


@pytest.fixture
def third_party(monkeypatch):
    tiers = replace(
        reg._BIM_IN_MOTION_PROFILE,
        report_structure=ReportStructureSpec(
            finding_reference_column_label="Référence référentiel client",
            referential_sheet_name="Référentiel BIM in Motion",
        ),
    )
    monkeypatch.setattr(reg, "_PROFILES", (reg._I3F_PROFILE, tiers))
    return "bim_in_motion"


def test_third_party_sheet_names_and_headers_have_no_i3f(tmp_path, monkeypatch, third_party):
    wb = _workbook(tmp_path, monkeypatch, profile_id=third_party)
    surface = list(wb.sheetnames)
    for name in wb.sheetnames:
        surface += [c.value for c in wb[name][1] if c.value]
    joined = " | ".join(str(s) for s in surface)
    for term in FORBIDDEN:
        assert term not in joined, f"{term!r} dans les onglets ou en-têtes d'un AMO tiers"


def test_third_party_gets_its_own_labels(tmp_path, monkeypatch, third_party):
    wb = _workbook(tmp_path, monkeypatch, profile_id=third_party)
    assert "Référentiel BIM in Motion" in wb.sheetnames
    assert "Référence référentiel client" in [c.value for c in wb["Findings (tous)"][1]]


def test_profile_without_structure_falls_back_to_neutral(tmp_path, monkeypatch):
    """`bim_in_motion` nu : replis neutres, jamais les libellés d'I3F."""
    wb = _workbook(tmp_path, monkeypatch, profile_id="bim_in_motion")
    assert xlsx_annex.DEFAULT_REFERENTIAL_SHEET_NAME in wb.sheetnames
    assert I3F_SHEET_NAME not in wb.sheetnames
    headers = [c.value for c in wb["Findings (tous)"][1]]
    assert xlsx_annex.DEFAULT_REFERENCE_COLUMN_LABEL in headers
    assert I3F_COLUMN_LABEL not in headers


# ── 3. Smoke : le classeur reste réellement ouvrable ──────────────────


def test_workbook_is_readable_and_complete(tmp_path, monkeypatch):
    wb = _workbook(tmp_path, monkeypatch)
    assert "Synthèse" in wb.sheetnames
    assert "Findings (tous)" in wb.sheetnames
    findings_sheet = wb["Findings (tous)"]
    assert findings_sheet.max_row >= 2, "aucune ligne de finding écrite"
    assert findings_sheet.max_column == len(xlsx_annex.COLUMNS) + 2  # + suggestions


def test_sheet_names_stay_within_excel_limit(tmp_path, monkeypatch, third_party):
    """Excel refuse au-delà de 31 caractères — le tronquage doit tenir."""
    long_spec = replace(
        reg._BIM_IN_MOTION_PROFILE,
        report_structure=ReportStructureSpec(
            finding_reference_column_label="X",
            referential_sheet_name="Référentiel " + "T" * 60,
        ),
    )
    monkeypatch.setattr(reg, "_PROFILES", (reg._I3F_PROFILE, long_spec))
    wb = _workbook(tmp_path, monkeypatch, profile_id="bim_in_motion")
    assert all(len(n) <= 31 for n in wb.sheetnames)
