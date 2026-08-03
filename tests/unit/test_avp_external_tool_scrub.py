"""Aucun outil tiers hérité du classeur MOA ne doit atteindre le client.

Les classeurs MOA de référence sont recyclés comme **gabarit de mise en forme**.
Ils portent l'évaluation d'un autre chantier, dont le logiciel avec lequel chaque
point a été contrôlé (``Solibri``, ``BimCollabZoom``). Recopiées, ces mentions
attribuent le contrôle à un outillage que la chaîne BIMData n'emploie pas, et
trahissent le projet dont le template a été repris.

Trois classes de contamination, observées sur le classeur Tarare réel :

1. colonne « Outil utilisé » de la grille (``D``) ;
2. colonne « Commentaires CdP Bim » (``F``), déjà purgée historiquement ;
3. libellés d'**instruction** des onglets de statistiques
   (« zone de copie de la liste BimCollabZoom »), qui vivent hors de toute
   colonne d'évaluation et échappaient donc à la purge de grille.
"""

from __future__ import annotations

import zipfile

import openpyxl
import pytest

from audit_bim.reporting.avp.pack import _qa_external_tool_mentions
from audit_bim.reporting.avp.xlsx_controle import (
    _clear_template_grid_assessments,
    _scrub_external_tool_mentions,
)


@pytest.fixture
def grille_moa(tmp_path):
    """Grille MOA recyclée, avec l'évaluation d'un chantier précédent."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grille de contrôle"
    ws.append(
        [
            "CODE 3F",
            "POINTS DE CONTROLE",
            "EXIGENCE CCH BIM 3F",
            "Outil utilisé",
            "EVALUATION",
            "Commentaires CdP Bim",
        ]
    )
    ws.append(
        [
            "5.4",
            "Dimensionnement des pièces",
            "Exigence CCH",
            "Solibri",
            "Conforme",
            "RAS sur Tarare",
        ]
    )
    ws.append(
        [
            "6.5",
            "Absence de conflits majeurs",
            "Exigence CCH",
            "BimCollabZoom",
            "Non conforme",
            "Les logiciels tels que BIMcollabZOOM ne suffisent pas.",
        ]
    )
    onglet = wb.create_sheet("Zones Nommage")
    onglet["B10"] = "zone de copie de la liste BimCollabZoom"
    chemin = tmp_path / "grille.xlsx"
    wb.save(chemin)
    return chemin


def test_grid_purge_clears_the_tool_column(grille_moa):
    """La colonne 4 est une évaluation, pas le référentiel : elle doit sauter.

    Les colonnes 1-3 (code, point de contrôle, exigence CCH) sont le référentiel
    I3F et doivent survivre — sans elles, la grille n'a plus d'objet.
    """
    wb = openpyxl.load_workbook(grille_moa)
    ws = wb["Grille de contrôle"]

    _clear_template_grid_assessments(ws)

    assert [ws.cell(2, c).value for c in (4, 5, 6)] == [None, None, None]
    assert [ws.cell(3, c).value for c in (4, 5, 6)] == [None, None, None]
    assert ws.cell(2, 1).value == "5.4"
    assert ws.cell(2, 2).value == "Dimensionnement des pièces"
    assert ws.cell(2, 3).value == "Exigence CCH"


def test_scrub_reaches_instruction_labels_outside_the_grid(grille_moa):
    """Le libellé « zone de copie… » ne vit dans aucune colonne d'évaluation.

    C'est précisément par là que ``BimCollabZoom`` survivait à la purge de grille.
    """
    wb = openpyxl.load_workbook(grille_moa)

    _clear_template_grid_assessments(wb["Grille de contrôle"])
    n = _scrub_external_tool_mentions(wb)

    libelle = wb["Zones Nommage"]["B10"].value
    assert "BimCollab" not in libelle
    assert libelle == "zone de copie de la liste BIMData / IFC OpenShell"
    assert n >= 1


def test_scrub_leaves_formulas_untouched(tmp_path):
    """Réécrire une formule casserait le classeur ; elles ne portent pas de marque."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "=SUM(D2:D7)"

    assert _scrub_external_tool_mentions(wb) == 0
    assert wb.active["A1"].value == "=SUM(D2:D7)"


class _FauxPack:
    def __init__(self, chemins):
        self._chemins = chemins

    def paths(self):
        return self._chemins


def _xlsx_avec(tmp_path, nom, valeur):
    wb = openpyxl.Workbook()
    wb.active["A1"] = valeur
    chemin = tmp_path / nom
    wb.save(chemin)
    return chemin


def test_qa_gate_flags_any_surviving_mention(tmp_path):
    """Filet final : on inspecte le XML réel, sans hypothèse sur l'emplacement."""
    sale = _xlsx_avec(tmp_path, "sale.xlsx", "Contrôlé sous Solibri")
    propre = _xlsx_avec(tmp_path, "propre.xlsx", "Contrôlé sous BIMData / IFC OpenShell")

    contamines = _qa_external_tool_mentions(_FauxPack([sale, propre]))

    assert contamines == ["sale.xlsx"]


def test_qa_gate_catches_mentions_outside_cells(tmp_path):
    """Une marque peut vivre hors des cellules — ici, dans les propriétés du document."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Rien à signaler"
    wb.properties.creator = "Solibri Model Checker"
    chemin = tmp_path / "meta.xlsx"
    wb.save(chemin)

    assert _qa_external_tool_mentions(_FauxPack([chemin])) == ["meta.xlsx"]


def test_qa_gate_ignores_unreadable_artifacts(tmp_path):
    """Un PDF (best-effort) n'est pas une archive : ce n'est pas une contamination."""
    faux = tmp_path / "rapport.pdf"
    faux.write_bytes(b"%PDF-1.4 pas une archive zip")

    assert _qa_external_tool_mentions(_FauxPack([faux])) == []


def test_qa_gate_reads_docx_parts(tmp_path):
    """Le Word est couvert au même titre que les Excel."""
    chemin = tmp_path / "rapport.docx"
    with zipfile.ZipFile(chemin, "w") as z:
        z.writestr("word/document.xml", "<w:p>Contrôle réalisé sous BimCollab Zoom</w:p>")

    assert _qa_external_tool_mentions(_FauxPack([chemin])) == ["rapport.docx"]


def test_word_drops_template_instruction_rows():
    """Le rapport client n'explique pas comment se remplir.

    « zone de copie… » / « coller ici » pilotent la saisie dans l'Excel MOA. Repris
    dans le .docx, ils produisent un livrable qui décrit son propre mode d'emploi.
    """
    from audit_bim.reporting.avp.docx_analyse import _is_placeholder_row

    assert _is_placeholder_row(["", "zone de copie de la liste BimCollabZoom", ""])
    assert _is_placeholder_row(["Coller ici les zones", "", ""])
    assert _is_placeholder_row(["5.4", "Dimensionnement", "", "Solibri", "", ""])
    assert not _is_placeholder_row(["5.4", "Dimensionnement des pièces", "Exigence CCH"])
