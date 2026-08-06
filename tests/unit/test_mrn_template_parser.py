"""Inventaire du gabarit MRN — compteurs figés sur le fichier réel.

Ces tests ne valident pas un parseur contre un fichier fabriqué : ils le
valident contre **le gabarit livré par le maître d'ouvrage**, et figent ce qu'il
contient. Un gabarit qui change n'est pas un test qui casse par accident, c'est
un livrable dont la structure a bougé — et il faut le savoir avant de générer.

Le fichier source vit hors du dépôt. Les tests qui en dépendent sont ignorés
s'il est absent, mais le sont **explicitement** : un ignorer silencieux ferait
passer une suite verte pour une couverture, alors que rien n'aurait été vérifié.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from audit_bim.profiles.bim_in_motion.mrn import parse_mrn_template
from audit_bim.profiles.bim_in_motion.mrn.template import (
    MODEL_COLUMNS,
    NON_MODEL_COLUMN,
)

TEMPLATE = Path(
    "/Users/stani/code/MCP/Documents maître d'ouvrage/Documents BIM in Motion/"
    "Livrable/MRN_GRILLE_CONTROLE_MNEM_22062026.xlsx"
)

needs_template = pytest.mark.skipif(
    not TEMPLATE.is_file(), reason=f"gabarit MRN absent : {TEMPLATE}"
)


SHAPE = Path(__file__).parent / "fixtures" / "mrn_template_shape.json"


def _build_synthetic_template(path: Path, *, decoy_validation: bool = False) -> Path:
    """Reconstruit un gabarit à la **forme** du fichier réel, sans son contenu.

    Le fichier du maître d'ouvrage vit hors du dépôt : les tests qui en
    dépendent sont donc ignorés en CI. Une suite verte ne prouvait alors aucun
    des six compteurs — exactement ce qu'elle prétendait figer.

    Cette fixture rejoue la structure relevée sur le vrai gabarit (numéros de
    ligne, identifiants, marqueurs d'outillage), libellés remplacés. Le désordre
    des sections en fait partie : c'est ce qu'il faut éprouver.
    """
    import json

    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    shape = json.loads(SHAPE.read_text(encoding="utf-8"))
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "1. Informations générales"

    if decoy_validation:
        # Piège : une liste sur une colonne SANS rapport avec les statuts,
        # déclarée en premier. Un parseur qui prend « la première liste »
        # servirait ces valeurs-là.
        decoy = DataValidation(type="list", formula1='"APS,APD,PRO,EXPL"')
        sheet.add_data_validation(decoy)
        # ``AG`` contient la lettre G : une comparaison textuelle la prendrait
        # pour une colonne de statut. C'est le cas dangereux, pas ``B``.
        decoy.sqref = "B8:B134 AG8:AG134"

    status = DataValidation(
        type="list", formula1='"Conforme, Partiellement conforme, Non conforme, N/A"'
    )
    sheet.add_data_validation(status)
    status.sqref = "G8:I112 G116:I134"

    for row, number, marker in shape:
        if number:
            sheet.cell(row, 4).value = number
            sheet.cell(row, 5).value = f"Libellé {number}"
        if marker:
            sheet.cell(row, 6).value = marker

    # Note d'addendum après le dernier contrôle : c'est ce qui doit être
    # classé ``trailing``, et non traité comme une ligne de contrôle.
    sheet.cell(136, 5).value = "Ajout du 05/04/2026"

    categories = workbook.create_sheet("CAT_SSCAT")
    categories.append(["Catégorie", "Sous-catégorie", "Code", "Classe IFC"])
    categories.append(["CVC", "Ventilation", "CV01", "IfcAirTerminal"])

    workbook.save(path)
    return path


@pytest.fixture(scope="module")
def synthetic(tmp_path_factory):
    return parse_mrn_template(
        _build_synthetic_template(tmp_path_factory.mktemp("mrn") / "gabarit.xlsx")
    )


# ── Les six compteurs, éprouvés SANS le fichier local ─────────────────


def test_the_arbitrated_counters_hold_on_the_synthetic_template(synthetic):
    """Le contrôle qui tourne réellement en CI.

    Les tests sur le fichier du maître d'ouvrage sont ignorés là où il est
    absent : sans ce doublon synthétique, la CI verte ne prouvait aucun des six
    compteurs qu'elle était censée figer.
    """
    summary = synthetic.summary()
    assert summary["control_rows"] == 96
    assert summary["section_rows"] == 26
    assert summary["chapter_header_rows"] == 1
    assert summary["distinct_chapters"] == 2
    assert summary["distinct_chapter_ids"] == ["1", "2"]
    assert summary["last_control_row"] == 134
    assert summary["sections_are_ordered"] is False


def test_the_section_disorder_is_reproduced_by_the_fixture(synthetic):
    """La fixture doit porter le désordre, sinon elle n'éprouve rien."""
    order = list(synthetic.sections)
    assert order.index("2.13") < order.index("2.1")
    assert order.index("2.4") > order.index("2.8")
    for control in synthetic.controls:
        assert control.control_id.startswith(f"{control.section_id}.")


def test_the_tool_markers_are_frozen(synthetic):
    """PR 4 devra neutraliser **exactement** ces lignes.

    Un compte flou laisserait passer un marqueur oublié ou ajouté sans que rien
    ne casse — et la purge annoncée serait partielle.
    """
    assert len(synthetic.tool_markers) == 31
    assert min(synthetic.tool_markers) == 5
    assert max(synthetic.tool_markers) == 131
    assert sum(1 for value in synthetic.tool_markers.values() if "ITO" in value) == 2


def test_the_trailing_note_is_classified(synthetic):
    assert synthetic.last_control_row == 134
    assert synthetic.trailing_rows == [136]


def test_a_decoy_list_validation_never_becomes_the_statuses(tmp_path):
    """Le piège du P3, rejoué.

    Une liste déclarée avant celle des statuts, sur une colonne sans rapport,
    serait servie par un parseur qui prend « la première liste trouvée ». Le
    classeur accepterait alors des statuts que le parseur ignorerait.
    """
    template = parse_mrn_template(
        _build_synthetic_template(tmp_path / "piege.xlsx", decoy_validation=True)
    )
    assert template.status_values == [
        "Conforme",
        "Partiellement conforme",
        "Non conforme",
        "N/A",
    ]
    assert "APS" not in template.status_values


@pytest.fixture(scope="module")
def template():
    return parse_mrn_template(TEMPLATE)


# ── Les six compteurs arbitrés ────────────────────────────────────────


@needs_template
def test_the_arbitrated_counters_hold(template):
    """Les six compteurs validés, mesurés et non recopiés."""
    summary = template.summary()
    assert summary["control_rows"] == 96
    assert summary["section_rows"] == 26
    assert summary["chapter_header_rows"] == 1
    assert summary["distinct_chapters"] == 2
    assert summary["distinct_chapter_ids"] == ["1", "2"]
    assert summary["last_control_row"] == 134
    assert summary["sections_are_ordered"] is False


@needs_template
def test_one_header_row_does_not_mean_one_chapter(template):
    """La distinction qui a failli être perdue.

    Le gabarit ne porte qu'une ligne d'en-tête de chapitre, mais couvre deux
    chapitres via la numérotation. Figer ``chapter_header_rows == 1`` sans ce
    contrôle laisserait croire à un document mono-chapitre, et un générateur
    n'écrirait la moitié de la grille sans que rien ne le signale.
    """
    assert len(template.chapter_header_rows) == 1
    assert template.distinct_chapters == 2
    assert template.distinct_chapter_ids == ["1", "2"]
    assert {control.chapter_id for control in template.controls} == {"1", "2"}


@needs_template
def test_the_two_expected_sheets_are_present(template):
    assert template.sheet_names == ["1. Informations générales", "CAT_SSCAT"]
    assert template.categories, "l'onglet CAT_SSCAT doit être lu, pas seulement présent"


# ── La section vient du numéro, jamais de la position ─────────────────


@needs_template
def test_the_section_of_a_control_is_derived_from_its_number(template):
    """Règle centrale du parseur.

    Un parcours séquentiel qui mémorise « la dernière section vue » rattacherait
    des contrôles à la mauvaise section, silencieusement — et le classeur
    produit paraîtrait normal.
    """
    for control in template.controls:
        assert control.control_id.startswith(f"{control.section_id}.")
        assert control.section_id.startswith(f"{control.chapter_id}.")


@needs_template
def test_the_section_disorder_is_preserved(template):
    """Le désordre appartient au document du maître d'ouvrage.

    ``2.13`` apparaît au milieu du chapitre 1, et la suite n'est pas monotone.
    Le réordonner produirait un livrable qui ne correspond plus à la grille
    fournie — ce test est là pour qu'on ne « corrige » pas le client.
    """
    order = list(template.sections)
    assert template.sections_are_ordered is False
    assert order.index("2.13") < order.index("2.1"), "2.13 précède bien 2.1"
    assert order.index("2.4") > order.index("2.8"), "2.4 suit bien 2.8"


@needs_template
def test_a_control_of_chapter_two_appears_inside_chapter_one(template):
    """Non-vacuité du désordre : il traverse les chapitres, pas seulement l'ordre."""
    first_chapter_two = min(c.row for c in template.controls if c.chapter_id == "2")
    last_chapter_one = max(c.row for c in template.controls if c.chapter_id == "1")
    assert first_chapter_two < last_chapter_one


# ── Colonnes ──────────────────────────────────────────────────────────


@needs_template
def test_the_column_roles_are_the_arbitrated_ones():
    """G n'est pas une colonne maquette, malgré la validation qu'elle porte."""
    assert NON_MODEL_COLUMN == 7  # G — contrôles hors maquette numérique
    assert MODEL_COLUMNS == (8, 9)  # H, I
    assert NON_MODEL_COLUMN not in MODEL_COLUMNS


@needs_template
def test_the_status_values_are_read_from_the_workbook(template):
    """Les recopier en dur les figerait à côté du document.

    Si le maître d'ouvrage ajoute un statut, le gabarit l'accepterait et un
    parseur à liste figée l'ignorerait.
    """
    assert template.status_values == [
        "Conforme",
        "Partiellement conforme",
        "Non conforme",
        "N/A",
    ]


# ── Marqueurs d'outillage et lignes de fin ────────────────────────────


@needs_template
def test_tool_markers_are_located_not_erased(template):
    """PR 1 inventorie ; la neutralisation est le lot de génération.

    Les localiser d'abord permet de vérifier plus tard qu'ils ont tous été
    traités — supprimer sans inventaire ne laisse rien à contrôler.
    """
    assert len(template.tool_markers) == 31
    markers = set(template.tool_markers.values())
    assert "x" in markers
    assert any("ITO" in marker for marker in markers)
    assert all(row >= 1 for row in template.tool_markers)


@needs_template
def test_the_rows_after_the_last_control_are_trailing_notes(template):
    """Rien ne doit être écrit après 134 tant que ces lignes ne sont pas classées."""
    assert template.last_control_row == 134
    assert template.trailing_rows == [136]
    assert all(row > template.last_control_row for row in template.trailing_rows)


# ── Refus explicites ──────────────────────────────────────────────────


def test_a_missing_template_is_refused(tmp_path):
    with pytest.raises(FileNotFoundError):
        parse_mrn_template(tmp_path / "absent.xlsx")


def test_a_workbook_without_the_control_sheet_is_refused(tmp_path):
    """Mieux vaut refuser qu'inventorier zéro contrôle.

    Un gabarit sans onglet de contrôle produirait un inventaire vide, qui
    passerait pour « une grille sans contrôles » au lieu d'un mauvais fichier.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active.title = "Autre chose"
    path = tmp_path / "faux.xlsx"
    workbook.save(path)

    with pytest.raises(ValueError, match="Informations générales"):
        parse_mrn_template(path)


def test_the_parser_never_imports_the_i3f_profile():
    """Le livrable MRN appartient à BIM in Motion, pas à I3F.

    Les deux référentiels sont distincts ; les faire communiquer ferait entrer
    les règles de l'un dans le livrable de l'autre.
    """
    import ast

    from audit_bim.profiles.bim_in_motion.mrn import template as module

    source = Path(module.__file__).parent
    for path in sorted(source.glob("*.py")):
        tree = ast.parse(path.read_text(encoding="utf-8"))
        for node in ast.walk(tree):
            modules = []
            if isinstance(node, ast.Import):
                modules = [alias.name for alias in node.names]
            elif isinstance(node, ast.ImportFrom) and node.module:
                # Les alias comptent : ``from audit_bim.profiles import i3f``
                # n'a pas « i3f » dans son module, seulement dans ses noms.
                modules = [node.module] + [f"{node.module}.{alias.name}" for alias in node.names]
            assert not [m for m in modules if "profiles.i3f" in m or "i3f" in m], path.name


def test_the_anti_i3f_guard_sees_an_aliased_import():
    """Non-vacuite : le controle doit reconnaitre les formes qu'il interdit.

    Sans le cas des alias, ``from audit_bim.profiles import i3f`` passerait —
    le garde-fou dirait plus que ce qu'il mesure.
    """
    import ast

    def _modules(source: str) -> list[str]:
        found = []
        for node in ast.walk(ast.parse(source)):
            if isinstance(node, ast.Import):
                found += [alias.name for alias in node.names]
            elif isinstance(node, ast.ImportFrom) and node.module:
                found += [node.module] + [f"{node.module}.{a.name}" for a in node.names]
        return found

    for source in (
        "from audit_bim.profiles.i3f import tools_audit\n",
        "from audit_bim.profiles import i3f\n",
        "import audit_bim.profiles.i3f.tools_query\n",
    ):
        assert [m for m in _modules(source) if "i3f" in m], source

    assert not [m for m in _modules("from audit_bim.mcp.session import _State\n") if "i3f" in m]
