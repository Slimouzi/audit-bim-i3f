"""Tests du parseur d'annexe « Spécification des données » format CCH 2026.

Les fixtures sont des classeurs synthétiques minimaux reproduisant la
structure Vbéta2 (Objet en B, Classe IFC en C, phases en D..J, propriété en M,
localisation en N) — l'annexe réelle est propriétaire et n'est pas embarquée.
"""

from __future__ import annotations

import openpyxl
import pytest

from audit_bim.requirements.data_spec_parser import parse_data_spec
from audit_bim.requirements.data_spec_parser_2026 import (
    is_2026_format,
    parse_data_spec_2026,
)
from audit_bim.requirements.models import BIMPhase


def _write_row(ws, row_idx: int, values: dict[int, object]) -> None:
    for col, value in values.items():
        ws.cell(row=row_idx, column=col + 1, value=value)  # openpyxl 1-based


@pytest.fixture()
def annexe_2026(tmp_path):
    """Classeur synthétique au format Vbéta2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ANNEXE 2 "

    # En-tête de bloc (détection du format : Objet en B, Classe IFC en C)
    _write_row(
        ws,
        3,
        {
            1: "Objet",
            2: "Classe IFC",
            10: "Attendu à partir \nde",
            12: "Propriétés de l'objet",
            13: "Localisation dans l'IFC",
        },
    )
    # Section documents projet
    _write_row(ws, 4, {1: "Projet", 2: "IfcProject", 12: "Documents"})
    _write_row(ws, 5, {3: "X", 4: "X", 5: "X", 12: "Planning"})
    # Bloc pièce : propriété LongName exigée dès APS (X sur D..J partiels)
    _write_row(ws, 6, {1: "Espace \n(Pièces)", 2: "IfcSpace"})
    _write_row(
        ws,
        7,
        {
            3: "X",
            4: "X",
            5: "X",
            6: "X",
            10: "APS / PC",
            12: "Nom 3F de la Pièce",
            13: "IfcLongName",
            15: "liste de données \n(CCH)",
        },
    )
    # Propriété portée par le pset propriétaire (« Pset I3F » → Pset_3F)
    _write_row(ws, 8, {6: "X", 12: "Lot 3F", 13: "Pset I3F"})
    # Quantité BaseQuantities
    _write_row(ws, 9, {3: "X", 12: "Surface", 13: "BaseQuantites/NetArea"})
    # Bloc multi-classes + sous-section Documents rattachée à l'objet
    _write_row(ws, 10, {1: "Réseaux aéraulique", 2: "IfcDuctFittingType\nIfcDuctSegmentType"})
    _write_row(ws, 11, {8: "X", 12: "Matériaux", 13: "IfcMaterial"})
    _write_row(ws, 12, {12: "Documents"})
    _write_row(ws, 13, {8: "X", 12: "Fiches techniques"})
    # Ligne sans X mais avec « Attendu à partir de » → DOE..GESTION
    _write_row(ws, 14, {1: "Pompe", 2: "IfcPumpType"})
    _write_row(
        ws, 15, {10: "DOE", 12: "Référence commerciale", 13: "Pset_ManufacturerTypeInformation"}
    )

    path = tmp_path / "annexe_2026.xlsx"
    wb.save(path)
    return path


@pytest.fixture()
def annexe_v37(tmp_path):
    """Classeur synthétique au format V3.7 (Objet en B, Classe IFC en D)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_row(
        ws,
        1,
        {
            0: "Thème",
            1: "Objet",
            2: "Définition",
            3: "Classe IFC",
            4: "Propriété",
            5: "Pset",
            6: "APS",
            7: "AVP",
            8: "PRO",
            9: "DCE",
            10: "EXE",
            11: "DOE",
            12: "GESTION",
        },
    )
    _write_row(
        ws,
        2,
        {0: "Générale", 1: "Pièce", 3: "IfcSpace", 4: "Nom de la pièce", 5: "LongName", 6: "X"},
    )
    path = tmp_path / "annexe_v37.xlsx"
    wb.save(path)
    return path


def test_detection_2026(annexe_2026, annexe_v37):
    assert is_2026_format(annexe_2026) is True
    assert is_2026_format(annexe_v37) is False


def test_parse_2026_proprietes(annexe_2026):
    specs = parse_data_spec_2026(annexe_2026)

    longname = [s for s in specs if s.property_name == "Nom 3F de la Pièce"]
    assert len(longname) == 1
    s = longname[0]
    assert s.ifc_class == "IfcSpace"
    assert s.pset_or_attribute == "LongName"
    assert s.kind == "property"
    assert BIMPhase.APS in s.required_phases
    assert BIMPhase.DCE in s.required_phases

    pset_3f = [s for s in specs if s.property_name == "Lot 3F"]
    assert pset_3f[0].pset_or_attribute == "Pset_3F"

    quantity = [s for s in specs if s.property_name == "Surface"]
    assert quantity[0].kind == "quantity"
    assert quantity[0].pset_or_attribute == "BaseQuantities/NetArea"


def test_parse_2026_multi_classes_et_documents(annexe_2026):
    specs = parse_data_spec_2026(annexe_2026)

    materiaux = [s for s in specs if s.property_name == "Matériaux"]
    assert {s.ifc_class for s in materiaux} == {"IfcDuctFittingType", "IfcDuctSegmentType"}

    # Documents projet → IfcProject ; documents objet → classes de l'objet
    planning = [s for s in specs if s.property_name == "Planning"]
    assert planning[0].ifc_class == "IfcProject"
    assert planning[0].kind == "document"
    fiches = [s for s in specs if s.property_name == "Fiches techniques"]
    assert {s.ifc_class for s in fiches} == {"IfcDuctFittingType", "IfcDuctSegmentType"}
    assert all(s.kind == "document" for s in fiches)


def test_parse_2026_attendu_a_partir_de(annexe_2026):
    specs = parse_data_spec_2026(annexe_2026)
    ref = [s for s in specs if s.property_name == "Référence commerciale"]
    assert ref[0].required_phases == [BIMPhase.DOE, BIMPhase.GESTION]
    assert ref[0].pset_or_attribute == "Pset_ManufacturerTypeInformation"


def test_dispatch_automatique(annexe_2026, annexe_v37):
    """parse_data_spec route vers le bon parseur selon le format détecté."""
    specs_2026 = parse_data_spec(annexe_2026)
    assert any(s.pset_or_attribute == "Pset_3F" for s in specs_2026)

    specs_v37 = parse_data_spec(annexe_v37)
    assert len(specs_v37) == 1
    assert specs_v37[0].ifc_class == "IfcSpace"
    assert specs_v37[0].pset_or_attribute == "LongName"
