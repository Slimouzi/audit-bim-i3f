"""Extracteur DOE Excel — convertit un xlsx d'équipements en ``DoeRecord``.

Conventions souples : on inspecte la première ligne pour repérer les
colonnes connues (UUID / Tag / Mark / Nom / Type / Étage / Zone), puis on
traite chaque ligne suivante comme un équipement. Toute colonne non
reconnue est considérée comme une propriété à appliquer.

Convention de propriété : si l'en-tête contient `.` ou `/`, le Pset est
extrait (ex: « Pset_3F.Fabricant » → Pset=`Pset_3F`, prop=`Fabricant`).
Sinon, la propriété va dans `Pset_DOE` par défaut.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from ...requirements._openpyxl_compat import patch_openpyxl
from ..models import DoeRecord

patch_openpyxl()


# Mapping en-têtes connues (insensible à la casse, accent-tolérant)
_HEADER_ALIASES = {
    "uuid": ("uuid", "globalid", "global id", "ifc guid", "ifcguid"),
    "tag": ("tag", "mark", "numero", "numéro", "code", "identifiant"),
    "name": ("nom", "libelle", "libellé", "designation", "désignation", "name"),
    "type": ("type", "categorie", "catégorie", "famille"),
    "storey": ("etage", "étage", "niveau", "storey", "level"),
    "zone": ("zone", "local", "piece", "pièce", "logement", "room"),
}

_DEFAULT_PSET = "Pset_DOE"


def _normalize(s: str) -> str:
    import unicodedata

    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


def _detect_header(value) -> tuple[str | None, tuple[str, str] | None]:
    """Pour une cellule d'en-tête, retourne :
    - ('uuid'/'tag'/'name'/'type'/'storey'/'zone', None) si c'est un alias connu
    - (None, (pset, prop)) si c'est une propriété (Pset.prop ou nom seul)
    """
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    n = _normalize(text)
    for slot, aliases in _HEADER_ALIASES.items():
        if n in aliases:
            return slot, None
    # Convention Pset.prop ou Pset/prop
    m = re.match(r"^([A-Za-z][A-Za-z0-9_]+)\s*[./]\s*(.+)$", text)
    if m:
        return None, (m.group(1).strip(), m.group(2).strip())
    return None, (_DEFAULT_PSET, text)


def parse_doe_excel(xlsx_path: str | Path) -> list[DoeRecord]:
    """Parse un fichier DOE Excel multi-feuilles.

    Algorithme :

    1. Itère sur toutes les feuilles du classeur.
    2. Pour chaque feuille, cherche la **ligne d'en-tête** = première
       ligne (parmi les 10 premières) avec ≥ 2 cellules non vides.
    3. Mappe chaque colonne via ``_detect_header`` (slot connu ou
       Pset.Propriété).
    4. Pour chaque ligne suivante non vide, construit un ``DoeRecord``.
    5. Ne garde que les lignes ayant **au moins un identifiant** (uuid /
       tag / name) **et au moins une propriété** — les lignes vides ou
       purement informatives sont filtrées.

    Args:
        xlsx_path: Chemin (str ou Path) vers le xlsx DOE.

    Returns:
        Liste de DoeRecord prête pour le matcher. Toutes feuilles
        confondues, dans l'ordre du classeur.

    Raises:
        FileNotFoundError: Si le fichier n'existe pas.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, data_only=True)
    records: list[DoeRecord] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # Cherche la ligne d'en-tête : 1ère ligne avec ≥ 2 colonnes
            # contenant du texte exploitable.
            header_idx = None
            for i, row in enumerate(rows[:10]):
                non_empty = sum(1 for c in row if c not in (None, ""))
                if non_empty >= 2:
                    header_idx = i
                    break
            if header_idx is None:
                continue
            headers = rows[header_idx]
            # Mapping colonne → slot/(pset,prop)
            col_map: list[tuple[str | None, tuple[str, str] | None]] = [
                _detect_header(h) for h in headers
            ]
            for row_i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                if not any(c not in (None, "") for c in row):
                    continue
                rec = DoeRecord(
                    source=str(path),
                    row_index=row_i,
                    properties={},
                    raw_row={
                        str(h or f"col{i}"): row[i] if i < len(row) else None
                        for i, h in enumerate(headers)
                    },
                )
                for col_i, (slot, pset_prop) in enumerate(col_map):
                    val = row[col_i] if col_i < len(row) else None
                    if val in (None, ""):
                        continue
                    if slot == "uuid":
                        rec.uuid_hint = str(val).strip()
                    elif slot == "tag":
                        rec.tag_hint = str(val).strip()
                    elif slot == "name":
                        rec.name_hint = str(val).strip()
                    elif slot == "type":
                        rec.type_hint = str(val).strip()
                    elif slot == "storey":
                        rec.storey_hint = str(val).strip()
                    elif slot == "zone":
                        rec.zone_hint = str(val).strip()
                    elif pset_prop:
                        pset, prop = pset_prop
                        rec.properties.setdefault(pset, {})[prop] = val
                # Ne garde que les lignes avec au moins un indice + une property
                if (rec.uuid_hint or rec.tag_hint or rec.name_hint) and rec.properties:
                    records.append(rec)
    finally:
        wb.close()
    return records
