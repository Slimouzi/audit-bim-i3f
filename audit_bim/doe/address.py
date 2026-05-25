"""Extraction de l'adresse postale projet à partir d'un fichier DOE.

Le DOE n'a pas de champ structuré « adresse » : l'information est
classiquement dans l'**en-tête** d'une feuille Excel (premières lignes
hors zone tabulaire), la **page de garde** d'un PDF, ou un bloc
d'identité projet sur une image OCRée.

Stratégie : scanner le texte des premières lignes / premières pages
puis appliquer une heuristique regex « numéro de voie + nom de voie
+ code postal + commune ». On accepte un appariement partiel (CP +
commune sans n° de voie) car beaucoup de DOE n'écrivent que la commune.

Utilisé en *fallback* dans :func:`audit_bim.enrichment.address.resolve_project_address`
quand la maquette n'expose pas d'``IfcPostalAddress``.
"""

from __future__ import annotations

import re
from pathlib import Path

from ..enrichment.models import ProjectAddress

# Code postal français : 5 chiffres suivis d'un nom de commune en
# majuscules (au moins première lettre). Tolère espaces, tirets,
# apostrophes, lettres accentuées dans le nom de la commune.
_CP_PATTERN = re.compile(
    r"\b(\d{5})\b\s+([A-ZÉÈÀÂÊÎÔÛÇ][A-Za-zÀ-ÿ' \-]{1,60})",
)

# Numéro de voie + type + nom : « 10 rue de Rivoli », « 1bis avenue X »,
# « 42, Boulevard Voltaire ». Le type de voie est limitatif pour
# minimiser les faux positifs.
_STREET_PATTERN = re.compile(
    r"\b(\d{1,4}(?:\s?(?:bis|ter|quater))?)[,\s]+"
    r"(rue|avenue|av\.?|boulevard|bd\.?|impasse|all[ée]e|place|chemin|route|"
    r"voie|cours|quai|passage|rond[-\s]?point)\s+"
    r"([A-ZÉÈÀÂÊÎÔÛÇa-zà-ÿ' \-]{2,80})",
    re.IGNORECASE,
)

# Marqueurs « adresse » au sens large — utilisés pour booster la zone
# de recherche dans le texte (on accepte du bruit autour).
_ADDR_KEYWORDS = re.compile(
    r"\b(adresse|site|projet|chantier|op[ée]ration|lieu|implantation)\b",
    re.IGNORECASE,
)


def extract_address_from_text(text: str) -> ProjectAddress | None:
    """Cherche un motif d'adresse postale FR dans un blob de texte.

    Args:
        text: Texte brut (concaténation d'en-têtes Excel, page de garde
            PDF, OCR…).

    Returns:
        :class:`ProjectAddress` avec ``source="doe"`` si trouvé, sinon
        ``None``.
    """
    if not text or not text.strip():
        return None

    cp_match = _CP_PATTERN.search(text)
    if not cp_match:
        return None
    postal_code = cp_match.group(1)
    town = _clean_town(cp_match.group(2))

    # Tente de récupérer un n° + voie. On cherche en priorité dans une
    # fenêtre proche de la mention CP, pour limiter les faux positifs
    # sur des n° type « bât. 12 » au milieu du document.
    cp_pos = cp_match.start()
    window = text[max(0, cp_pos - 200) : cp_pos + 200]
    street_match = _STREET_PATTERN.search(window) or _STREET_PATTERN.search(text)

    address_lines: list[str] = []
    if street_match:
        line = " ".join(
            p.strip()
            for p in (
                street_match.group(1),
                street_match.group(2),
                _clean_town(street_match.group(3)),
            )
            if p
        )
        address_lines.append(line)

    return ProjectAddress(
        source="doe",
        address_lines=address_lines,
        postal_code=postal_code,
        town=town,
    )


def _clean_town(raw: str) -> str:
    """Normalise un libellé commune : supprime queue numérique / suffixe parasite."""
    s = raw.strip().rstrip(",;:.")
    # Coupe au premier saut de ligne / tab / longue espace
    s = re.split(r"[\n\r\t]|  +", s, maxsplit=1)[0].strip()
    return s


# ── Lecture des sources DOE ──────────────────────────────────────────────


def _from_xlsx(path: Path, max_rows: int = 20) -> str:
    """Concatène les ``max_rows`` premières lignes de chaque feuille xlsx."""
    import openpyxl

    chunks: list[str] = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                for cell in row:
                    if cell is not None:
                        chunks.append(str(cell))
    finally:
        wb.close()
    return "\n".join(chunks)


def _from_pdf(path: Path, max_pages: int = 2) -> str:
    """Concatène le texte des ``max_pages`` premières pages d'un PDF natif."""
    try:
        import pdfplumber
    except ImportError:
        return ""
    chunks: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages[:max_pages]:
            text = page.extract_text() or ""
            if text:
                chunks.append(text)
    return "\n".join(chunks)


def _from_image_ocr(path: Path) -> str:
    """Lance Tesseract (français) sur une image. ``""`` si OCR indisponible."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return ""
    return pytesseract.image_to_string(Image.open(path), lang="fra") or ""


def extract_address_from_doe(doe_path: str | Path) -> ProjectAddress | None:
    """Détecte l'adresse postale projet dans un fichier DOE.

    Dispatch par extension :

    - ``.xlsx`` / ``.xlsm`` → 20 premières lignes de chaque feuille.
    - ``.pdf`` → 2 premières pages (texte natif via pdfplumber).
    - ``.png`` / ``.jpg`` / ``.tif`` → OCR Tesseract (français).

    Args:
        doe_path: Chemin du fichier DOE.

    Returns:
        :class:`ProjectAddress` (``source="doe"``) si une adresse est
        détectée, ``None`` sinon (fichier inexistant, format non géré,
        ou regex sans match).
    """
    p = Path(doe_path)
    if not p.exists():
        return None
    ext = p.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        text = _from_xlsx(p)
    elif ext == ".pdf":
        text = _from_pdf(p)
    elif ext in (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"):
        text = _from_image_ocr(p)
    else:
        return None
    return extract_address_from_text(text)
