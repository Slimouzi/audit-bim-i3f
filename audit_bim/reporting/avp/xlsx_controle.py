"""Builder « Contrôle Maquettes » (.xlsx) et l'ensemble de ses helpers.

Grille de contrôle MOA, template vivant openpyxl, statistiques de conformité
dérivées de l'``AuditResult`` et comptage QA de la grille.
"""

from __future__ import annotations

from collections import Counter
from copy import copy
from pathlib import Path

import openpyxl
import xlsxwriter

from ...audit.engine import AuditResult
from ...audit.findings import ErrorType, Theme
from ...extraction.model_data import ModelSnapshot
from ..avp_sources import ControleMaquettesSource, SheetTable
from ..word_report import NOT_AVAILABLE
from ..xlsx_annex import write_safe
from .models import _CONTROLE_STATS_SHEETS, AvpMeta
from .xlsx_common import (
    _cell,
    _dict_text,
    _moa_formats,
    _norm,
    _openpyxl_safe_value,
    _write_moa_grid,
    _write_moa_value,
)


def _build_controle_maquettes_xlsx(
    path, result, sources, meta, snap: ModelSnapshot | None = None
) -> Path:
    src = sources.controle if sources else None
    template_path = getattr(src, "template_path", None)
    if template_path:
        template = Path(template_path)
        if template.exists():
            return _build_controle_maquettes_template_xlsx(path, template, result, src, meta, snap)

    wb = xlsxwriter.Workbook(str(path), {"strings_to_formulas": False})
    fmts = _moa_formats(wb)
    ws = wb.add_worksheet("Grille de contrôle")
    _write_controle_moa_header(ws, fmts, src, meta)

    grille = _audit_controle_table(result)
    if grille is None and src is not None:
        grille = src.grille
    _write_controle_moa_table(ws, fmts, grille)

    # Onglets de stats conformité : synthèse KPI depuis l'audit.
    for name in _CONTROLE_STATS_SHEETS:
        ws_s = wb.add_worksheet(name[:31])
        grid = _controle_grid(name, src) if result is None else None
        if grid and grid.rows:
            _write_moa_grid(ws_s, fmts, grid.rows, start_row=0)
        else:
            rows = _controle_stat_rows(name, _controle_stats(name, result, src), meta)
            _write_moa_grid(ws_s, fmts, rows, start_row=0)
    wb.close()
    return path


def _build_controle_maquettes_template_xlsx(
    path: Path,
    template_path: Path,
    result: AuditResult | None,
    src: ControleMaquettesSource | None,
    meta: AvpMeta,
    snap: ModelSnapshot | None = None,
) -> Path:
    """Réutilise le classeur MOA Contrôle comme template vivant.

    Les onglets statistiques du fichier MOA contiennent des formules, listes de
    conformité et zones de copie. Le générateur ne doit pas les remplacer par un
    squelette : on conserve le template et on injecte uniquement l'identité
    projet + les listes maquette/audit disponibles.
    """
    wb = openpyxl.load_workbook(template_path, data_only=False)
    try:
        _ensure_control_template_sheets(wb)
        _refresh_template_metadata(wb, src, meta)
        effective_snap = (
            snap if snap is not None else (result.snapshot if result is not None else None)
        )
        if effective_snap is not None:
            _clear_template_grid_assessments(wb["Grille de contrôle"])
            _refresh_template_stats_tabs(wb, effective_snap)
        if result is not None:
            _append_audit_summary_to_template_grid(wb, result)
        _force_workbook_recalc(wb)
        wb.save(path)
    finally:
        wb.close()
    return path


def _ensure_control_template_sheets(wb) -> None:
    if "Grille de contrôle" not in wb.sheetnames:
        wb.create_sheet("Grille de contrôle", 0)
    for name in _CONTROLE_STATS_SHEETS:
        if name not in wb.sheetnames:
            wb.create_sheet(name)
    expected = ["Grille de contrôle", *_CONTROLE_STATS_SHEETS]
    ordered = [wb[name] for name in expected]
    ordered.extend(ws for ws in wb.worksheets if ws.title not in expected)
    wb._sheets = ordered


def _force_workbook_recalc(wb) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is None:
        return
    try:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True
    except AttributeError:
        return


def _metadata_cell(ws, label: str, fallback: str):
    wanted = label.strip().lower()
    for row in range(1, min(ws.max_row, 12) + 1):
        cell_label = ws.cell(row, 2).value
        if isinstance(cell_label, str) and cell_label.strip().lower().startswith(wanted):
            return ws.cell(row, 3)
    return ws[fallback]


def _refresh_template_metadata(wb, src: ControleMaquettesSource | None, meta: AvpMeta) -> None:
    if "Grille de contrôle" not in wb.sheetnames:
        return
    ws = wb["Grille de contrôle"]
    header = (src.header if src else {}) or {}
    values = {
        "Projet": meta.project_name or header.get("projet"),
        "ESI": meta.project_code or header.get("esi"),
        "Phase": meta.phase or header.get("phase"),
    }
    for label, value in values.items():
        if value not in (None, ""):
            _metadata_cell(
                ws, label, {"Projet": "C5", "ESI": "C6", "Phase": "C7"}[label]
            ).value = _openpyxl_safe_value(value)
    if meta.date_controle:
        _metadata_cell(ws, "Date d'analyse", "C9").value = _openpyxl_safe_value(meta.date_controle)


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def _copy_template_row_style(ws, src_row: int, dst_row: int, *, max_col: int = 6) -> None:
    for col in range(1, max_col + 1):
        _copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def _find_control_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row, 40) + 1):
        if any(
            isinstance(ws.cell(row, col).value, str)
            and ws.cell(row, col).value.strip().lower() == "points de controle"
            for col in range(1, ws.max_column + 1)
        ):
            return row
    return None


def _clear_template_grid_assessments(ws) -> None:
    header_row = _find_control_header_row(ws)
    if header_row is None:
        return
    for row in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row, 2).value in (None, ""):
            continue
        for col in (5, 6):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.startswith("="):
                continue
            ws.cell(row, col).value = None


def _append_audit_summary_to_template_grid(wb, result: AuditResult) -> None:
    if "Grille de contrôle" not in wb.sheetnames:
        return
    ws = wb["Grille de contrôle"]
    header_row = _find_control_header_row(ws)
    if header_row is None:
        return
    _clear_template_grid_assessments(ws)
    grille = _audit_controle_table(result)
    if grille is None or not grille.rows:
        return
    start = ws.max_row + 2
    _copy_template_row_style(ws, header_row, start)
    ws.cell(start, 1).value = "Synthèse audit MCP"
    for col in range(2, 7):
        ws.cell(start, col).value = None
    headers = [
        "CODE 3F",
        "POINTS DE CONTROLE",
        "EXIGENCE CCH BIM 3F",
        "Outil utilisé",
        "EVALUATION",
        "Commentaires CdP Bim",
    ]
    _copy_template_row_style(ws, header_row, start + 1)
    for col, value in enumerate(headers, start=1):
        ws.cell(start + 1, col).value = value
    data_style_row = header_row + 1
    for idx, row_values in enumerate(_controle_rows_for_moa(grille), start=start + 2):
        _copy_template_row_style(ws, data_style_row, idx)
        for col, value in enumerate(row_values[: len(headers)], start=1):
            ws.cell(idx, col).value = _openpyxl_safe_value(value)


def _copy_zone_start_row(ws) -> int:
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 5) + 1)]
        if any(isinstance(value, str) and "coller ici" in value.lower() for value in values):
            return row + 1
    return 13


def _clear_template_copy_zone(ws, *, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in (1, 2, 3):
            ws.cell(row, col).value = None


def _write_template_counts(
    wb,
    sheet_name: str,
    counts: Counter[str],
    row_label: str,
    *,
    clear_when_empty: bool,
    end_row: int = 200,
) -> None:
    if sheet_name not in wb.sheetnames:
        return
    if not counts and not clear_when_empty:
        return
    ws = wb[sheet_name]
    start = _copy_zone_start_row(ws)
    clear_to = max(end_row, start + len(counts) + 2)
    _clear_template_copy_zone(ws, start_row=start, end_row=min(clear_to, max(ws.max_row, end_row)))
    for idx, (label, count) in enumerate(sorted(counts.items(), key=lambda item: item[0].lower())):
        row = start + idx
        ws.cell(row, 1).value = row_label if idx == 0 else None
        ws.cell(row, 2).value = _openpyxl_safe_value(label)
        ws.cell(row, 3).value = count


def _counter_from_dicts(
    items: list[dict] | None, *keys: str, empty_label: str = "(vide)"
) -> Counter[str]:
    counter: Counter[str] = Counter()
    for item in items or []:
        label = _dict_text(item, *keys) or empty_label
        counter[label] += 1
    return counter


def _refresh_template_stats_tabs(wb, snap: ModelSnapshot) -> None:
    zones = snap.zones or []
    spaces = snap.spaces or []
    elements = snap.elements or []
    _write_template_counts(
        wb,
        "Zones Nommage",
        _counter_from_dicts(zones, "name", "Name", empty_label="(Name vide)"),
        "Zones",
        clear_when_empty=True,
    )
    _write_template_counts(
        wb,
        "Pièces Nommage",
        _counter_from_dicts(
            spaces,
            "longname",
            "long_name",
            "LongName",
            "name",
            "Name",
            empty_label="(Name vide)",
        ),
        "Pièces",
        clear_when_empty=True,
    )
    _write_template_counts(
        wb,
        "Zones ObjectType",
        _counter_from_dicts(
            zones,
            "object_type",
            "objectType",
            "ObjectType",
            "predefined_type",
            "PredefinedType",
            empty_label="(ObjectType vide)",
        ),
        "Zones",
        clear_when_empty=True,
    )
    missing_materials = Counter(
        _dict_text(element, "type", "ifc_type", "ifc_class", "IFCType") or "(classe IFC absente)"
        for element in elements
        if not _has_material(element)
    )
    _write_template_counts(
        wb,
        "ARC bsence de matériau",
        missing_materials,
        "ObjT Pièces",
        clear_when_empty=True,
        end_row=206,
    )
    if "ARC bsence de matériau" in wb.sheetnames:
        ws = wb["ARC bsence de matériau"]
        if ws.max_row >= 7 and ws.max_column >= 7:
            ws.cell(7, 7).value = len(elements)


def _write_controle_moa_header(ws, fmts, src, meta) -> None:
    """Entête proche du classeur MOA « Contrôle Maquettes »."""
    header = (src.header if src else {}) or {}
    legend = (src.legend if src else {}) or {
        0: "Non fourni / non trouvé",
        1: "Insuffisant : à reprendre ou compléter",
        2: "Satisfaisant",
    }
    ws.set_column(0, 0, 12)
    ws.set_column(1, 1, 34)
    ws.set_column(2, 2, 54)
    ws.set_column(3, 3, 22)
    ws.set_column(4, 4, 14)
    ws.set_column(5, 5, 48)
    write_safe(ws, 1, 0, "2.         Grille de contrôle des exigences du CCH BIM 3F", fmts["title"])
    write_safe(
        ws,
        2,
        0,
        "Contrôle automatisé depuis l'audit BIM et les données extraites de la maquette.",
        fmts["note"],
    )
    fallbacks = {"projet": meta.project_name, "esi": meta.project_code, "phase": meta.phase}
    for offset, (label, key) in enumerate(
        (("Projet", "projet"), ("ESI", "esi"), ("Phase", "phase")),
        start=4,
    ):
        val = header.get(key)
        if val in (None, ""):
            val = fallbacks[key]
        write_safe(ws, offset, 1, label, fmts["meta_label"])
        write_safe(ws, offset, 2, _cell(val), fmts["meta_value"])
    for idx, code in enumerate(sorted(legend), start=6):
        write_safe(ws, idx, 4, code, fmts["center"])
        write_safe(ws, idx, 5, legend[code], fmts["note"])


def _write_controle_moa_table(ws, fmts, grille: SheetTable | None) -> None:
    headers = [
        "CODE 3F",
        "POINTS DE CONTROLE",
        "EXIGENCE CCH BIM 3F",
        "Outil utilisé",
        "EVALUATION",
        "Commentaires CdP Bim",
    ]
    start = 12
    _write_moa_grid(ws, fmts, [headers], start_row=start)
    if grille is None or not grille.rows:
        write_safe(ws, start + 1, 0, NOT_AVAILABLE, fmts["data"])
        return
    rows = _controle_rows_for_moa(grille)
    for i, row in enumerate(rows, start=start + 1):
        for c, value in enumerate(row[: len(headers)]):
            fmt = fmts["center"] if c == 4 else fmts["data"]
            _write_moa_value(ws, i, c, value, fmt)


def _controle_rows_for_moa(grille: SheetTable) -> list[list]:
    """Normalise une grille source/audit vers les 6 colonnes MOA."""
    headers = [str(h or "").strip() for h in grille.headers]
    if headers == [
        "CODE 3F",
        "POINTS DE CONTROLE",
        "EXIGENCE CCH BIM 3F",
        "Outil utilisé",
        "EVALUATION",
        "Commentaires CdP Bim",
    ]:
        return grille.rows
    rows: list[list] = []
    for raw in grille.rows:
        point = raw[0] if len(raw) > 0 else ""
        total = raw[1] if len(raw) > 1 else ""
        conformes = raw[2] if len(raw) > 2 else ""
        non_conformes = raw[3] if len(raw) > 3 else ""
        ratio = raw[4] if len(raw) > 4 else ""
        try:
            evaluation = 2 if float(non_conformes or 0) == 0 else 1
        except (TypeError, ValueError):
            evaluation = ""
        rows.append(
            [
                "",
                point,
                "Contrôle automatisé MCP audit-bim-i3f",
                "Audit BIM / IFC OpenShell",
                evaluation,
                (
                    f"Total: {total}; conformes: {conformes}; "
                    f"non conformes: {non_conformes}; taux: {ratio}"
                ),
            ]
        )
    return rows


def _controle_stat_rows(name: str, stats: dict | None, meta) -> list[list]:
    """Onglet statistique proche MOA, alimenté par l'audit quand disponible."""
    rows = [
        [],
        ["", meta.project_name],
        ["", meta.project_code],
        [],
        ["", name],
        [],
        ["", "Indicateur", "Total", "Conforme", "Taux Conforme", "Non Conforme", "Taux"],
    ]
    if not stats:
        rows.append(["", NOT_AVAILABLE])
        return rows
    rows.append(
        [
            "",
            stats.get("label") or name,
            stats.get("total"),
            stats.get("conforme"),
            stats.get("conforme_ratio"),
            stats.get("non_conforme"),
            stats.get("non_conforme_ratio"),
        ]
    )
    return rows


def _controle_grid(name: str, src):
    """Grille détaillée d'un onglet de contrôle (matching nom normalisé)."""
    if not src or not src.stat_grids:
        return None
    for key, grid in src.stat_grids.items():
        if _norm(key).replace("bsence", "absence") == _norm(name):
            return grid
    return None


def _controle_stats(name: str, result: AuditResult | None, src) -> dict | None:
    """Stats conformité d'un onglet — audit, puis source de repli sans snapshot."""
    if result is not None:
        return _audit_stats(name, result)
    if src and src.stats:
        # La clé source peut porter la faute d'origine ("ARC bsence…").
        for key, val in src.stats.items():
            if _norm(key) == _norm(name) or _norm(key).replace("bsence", "absence") == _norm(name):
                if val:
                    return val
    return None


def _zone_finding_kind(f) -> str | None:
    """Classe un finding de nommage d'IfcZone : ``"objecttype"`` ou ``"name"``.

    Le nommage de zone produit **deux contrôles distincts** partageant le même
    thème (``NAMING_ZONE``) : le **Name** (pattern XXXXL-YYYY, présence) et
    l'**ObjectType** (typologie dans la liste I3F, présence). On ne peut donc PAS
    agréger par thème.

    Discrimination **prioritaire** par le champ structuré ``field_path``
    (``"IfcZone.ObjectType"`` / ``"IfcZone.Name"``, bim-core ≥ 0.1.1) — fiable et
    **indépendant du libellé** du finding. Repli sur l'heuristique de wording
    uniquement pour les findings historiques sans ``field_path``.
    """
    if f.ifc_type != "IfcZone" or f.theme != Theme.NAMING_ZONE:
        return None
    # 1) Champ structuré (source de vérité) : un reformulage du wording des
    #    règles ne peut plus fausser silencieusement le classement.
    fp = (getattr(f, "field_path", None) or "").strip().lower()
    if fp.endswith(".objecttype"):
        return "objecttype"
    if fp.endswith(".name"):
        return "name"
    # 2) Repli heuristique (findings sans field_path).
    text = f"{f.recommended_action or ''} {f.expected or ''}".lower()
    if f.error_type == ErrorType.NAMING_NOT_IN_LIST or "objecttype" in text:
        return "objecttype"
    return "name"


def _material_name(item) -> str | None:
    """Nom de matériau exploitable d'un item de ``material_list`` (formes variées :
    ``{"material": {"name": …}}`` BIMData, ``{"name": …}``, ou chaîne)."""
    if isinstance(item, dict):
        mat = item.get("material")
        if isinstance(mat, dict) and mat.get("name") and str(mat["name"]).strip():
            return str(mat["name"]).strip()
        if item.get("name") and str(item["name"]).strip():
            return str(item["name"]).strip()
    elif isinstance(item, str) and item.strip():
        return item.strip()
    return None


def _has_material(e: dict) -> bool:
    """Vrai si l'élément porte un **vrai nom de matériau**.

    Lit ``material_list`` (forme produite par ``bimdata-read`` :
    ``[{"material": {"name": "Béton"}}]``) avec repli ``materials``. La simple
    présence de la clé ne suffit pas : il faut un nom non vide.
    """
    for key in ("material_list", "materials"):
        for item in e.get(key) or []:
            if _material_name(item):
                return True
    return False


def _naming_stat(total: int, nc: int) -> dict:
    conf = total - nc
    return {
        "label": "Nombre de Noms",
        "total": total,
        "conforme": conf,
        "conforme_ratio": conf / total if total else None,
        "non_conforme": nc,
        "non_conforme_ratio": nc / total if total else None,
    }


def _audit_stats(name: str, result: AuditResult) -> dict | None:
    snap = result.snapshot
    # Nommage zones : Name et ObjectType sont des contrôles SÉPARÉS (mêmes
    # zones, anomalies distinctes) — on ne compte pas le même ensemble deux fois.
    if name == "Zones Nommage":
        total = len(snap.zones or [])
        if total == 0:
            return None
        nc = len(
            {
                f.element_uuid
                for f in result.findings
                if _zone_finding_kind(f) == "name" and f.element_uuid
            }
        )
        return _naming_stat(total, nc)
    if name == "Zones ObjectType":
        total = len(snap.zones or [])
        if total == 0:
            return None
        nc = len(
            {
                f.element_uuid
                for f in result.findings
                if _zone_finding_kind(f) == "objecttype" and f.element_uuid
            }
        )
        return _naming_stat(total, nc)
    if name == "Pièces Nommage":
        total = len(snap.spaces or [])
        if total == 0:
            return None
        nc = len(
            {
                f.element_uuid
                for f in result.findings
                if f.theme == Theme.NAMING_SPACE and f.element_uuid
            }
        )
        return _naming_stat(total, nc)
    if "matériau" in name.lower():
        elements = snap.elements or []
        total = len(elements)
        if total == 0:
            return None
        sans = sum(1 for e in elements if not _has_material(e))
        conf = total - sans
        return {
            "label": "Nombre d'éléments sans matériau",
            "total": total,
            "conforme": conf,
            "conforme_ratio": conf / total if total else None,
            "non_conforme": sans,
            "non_conforme_ratio": sans / total if total else None,
        }
    return None


def _audit_controle_table(result: AuditResult | None) -> SheetTable | None:
    """Grille de contrôle **réelle dérivée de l'AuditResult** (aucune source I3F).

    Une ligne par point de contrôle effectivement évalué (nommage zones/pièces,
    ObjectType zones, matériaux), avec la conformité mesurée sur la maquette.
    Renvoie ``None`` si l'audit n'expose aucun point exploitable — auquel cas
    l'annexe Contrôle est vide et la QA gate lève (livrable non exploitable).
    """
    if result is None:
        return None
    rows: list[list] = []
    for name in _CONTROLE_STATS_SHEETS:
        stats = _audit_stats(name, result)
        if not stats:
            continue
        total = stats.get("total")
        nc = stats.get("non_conforme")
        conf = stats.get("conforme")
        if conf is None and total is not None and nc is not None:
            conf = total - nc
        ratio = stats.get("conforme_ratio")
        eval_value = 2 if isinstance(nc, (int, float)) and nc == 0 else 1
        rows.append(
            [
                "",
                name,
                "Contrôle automatisé MCP audit-bim-i3f",
                "Audit BIM / IFC OpenShell",
                eval_value,
                (
                    f"Total: {total if total is not None else ''}; "
                    f"conformes: {conf if conf is not None else ''}; "
                    f"non conformes: {nc if nc is not None else ''}; "
                    f"taux: {round(ratio, 3) if isinstance(ratio, (int, float)) else ''}"
                ),
            ]
        )
    if not rows:
        return None
    return SheetTable(
        title="Grille de contrôle",
        headers=[
            "CODE 3F",
            "POINTS DE CONTROLE",
            "EXIGENCE CCH BIM 3F",
            "Outil utilisé",
            "EVALUATION",
            "Commentaires CdP Bim",
        ],
        rows=rows,
    )


def _count_controle_rows(path: Path) -> int:
    """Compte les lignes de données sous l'en-tête MOA « POINTS DE CONTROLE ».

    Compteur **propre à l'annexe Contrôle** : contrairement à
    :func:`_count_business_rows`, il ignore l'entête projet, la légende, les
    titres et ``NOT_AVAILABLE`` — il ne compte que les points de contrôle réels
    de la grille. Sert de garde qualité fiable (0 = grille sans contrôle réel).
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return 0
    try:
        if "Grille de contrôle" not in wb.sheetnames:
            return 0
        ws = wb["Grille de contrôle"]
        anchor = None
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(isinstance(c, str) and c.strip().lower() == "points de controle" for c in row):
                anchor = r_idx
                break
        if anchor is None:
            return 0
        total = 0
        for row in ws.iter_rows(min_row=anchor + 1, values_only=True):
            cells = [c for c in row if c not in (None, "")]
            if not cells:
                continue
            if str(cells[0]).strip() == NOT_AVAILABLE:
                continue
            if len(row) > 1 and row[1] not in (None, ""):
                total += 1
            elif row[0] not in (None, ""):
                # Repli pour une vieille grille 5 colonnes éventuelle.
                continue
        return total
    finally:
        wb.close()
