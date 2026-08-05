"""Inventaire du gabarit de grille de contrôle MRN.

Ce module **lit** le classeur modèle et le décrit. Il ne contrôle rien, ne
remplit rien, et n'ouvre aucune connexion : le remplissage depuis la maquette
est un lot distinct, et les mélanger rendrait impossible de dire si un écart
vient du gabarit ou de la donnée.

Deux règles de lecture, tirées du fichier réel plutôt que de son apparence :

**La section d'une ligne se déduit de son identifiant, jamais de sa position.**
Dans le gabarit livré, la section ``2.13`` apparaît ligne 31, au milieu du
chapitre 1, et la suite n'est pas monotone — ``2.5, 2.6, 2.7, 2.8, 2.4, 2.9``.
Un parcours séquentiel qui mémorise « la dernière section vue » rattacherait
donc des contrôles à la mauvaise section, silencieusement.

**Le désordre est conservé.** Il appartient au document du maître d'ouvrage ;
le réordonner produirait un livrable qui ne correspond plus à la grille qu'il
a fournie.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

__all__ = [
    "MODEL_COLUMNS",
    "NON_MODEL_COLUMN",
    "TOOL_MARKER_COLUMN",
    "MRNControlRow",
    "MRNTemplate",
    "parse_mrn_template",
]

CONTROL_SHEET = "1. Informations générales"
CATEGORY_SHEET = "CAT_SSCAT"

#: Colonne des marqueurs d'outillage historique (Solibri, ITO…). Ce ne sont pas
#: des données client : elles disent avec quel logiciel le contrôle était fait
#: avant. Le lot de génération devra les neutraliser sans supprimer la colonne.
TOOL_MARKER_COLUMN = 6  # F

#: Colonne « NON MN » — contrôles hors maquette numérique. Elle porte la même
#: validation que les colonnes maquette mais n'en est pas une : rien n'y est
#: écrit tant que la source est l'API BIMData.
NON_MODEL_COLUMN = 7  # G

#: Colonnes de statut par maquette. Le gabarit en prévoit deux ; la V1 MCP ne
#: remplit que la première. La seconde reste disponible sans être inventée.
MODEL_COLUMNS = (8, 9)  # H, I

LABEL_COLUMN = 5  # E
NUMBER_COLUMN = 4  # D
PHASE_COLUMN = 2  # B
DATE_COLUMN = 3  # C
COMMENT_COLUMN = 10  # J

_CONTROL_RE = re.compile(r"^(\d+)\.(\d+)\.(\d+)$")
_SECTION_RE = re.compile(r"^(\d+)\.(\d+)$")
_CHAPTER_RE = re.compile(r"^(\d+)\.$")


@dataclass(frozen=True)
class MRNControlRow:
    """Une ligne de contrôle du gabarit, telle qu'elle s'y trouve.

    ``chapter_id`` et ``section_id`` sont **dérivés de** ``control_id``, pas de
    la position : c'est la seule lecture qui résiste au désordre des sections.
    """

    control_id: str
    section_id: str
    chapter_id: str
    label: str
    row: int
    phase: str | None = None
    control_date: Any = None
    tool_marker: str | None = None

    @classmethod
    def from_number(cls, number: str, **kwargs) -> MRNControlRow:
        chapter, section, _ = _CONTROL_RE.match(number).groups()
        return cls(
            control_id=number,
            section_id=f"{chapter}.{section}",
            chapter_id=chapter,
            **kwargs,
        )


@dataclass(frozen=True)
class MRNTemplate:
    """Description du gabarit : ce qu'il contient, et où."""

    path: Path
    sheet_names: list[str]
    controls: list[MRNControlRow]
    sections: dict[str, str]
    chapter_header_rows: list[int]
    last_control_row: int
    tool_markers: dict[int, str]
    trailing_rows: list[int]
    status_values: list[str]
    categories: list[dict[str, Any]] = field(default_factory=list)

    @property
    def distinct_chapter_ids(self) -> list[str]:
        """Chapitres réellement couverts, déduits de la numérotation.

        Distinct de ``chapter_header_rows`` : le gabarit ne porte qu'**une**
        ligne d'en-tête de chapitre alors qu'il couvre deux chapitres. Compter
        les en-têtes ferait croire à un document mono-chapitre.
        """
        return sorted({control.chapter_id for control in self.controls})

    @property
    def distinct_chapters(self) -> int:
        """Nombre de chapitres couverts. Distinct de ``chapter_header_rows``.

        Deux noms pour deux faits : le gabarit porte **une** ligne d'en-tête et
        couvre **deux** chapitres. Exposer une liste sous le nom d'un compteur
        rendait le contrat ambigu.
        """
        return len(self.distinct_chapter_ids)

    @property
    def sections_are_ordered(self) -> bool:
        """Les sections apparaissent-elles dans l'ordre de leur numéro ?

        ``False`` sur le gabarit livré. Exposé plutôt que corrigé : un parseur
        qui suppose l'ordre se trompe, et le savoir est ce qui empêche de le
        supposer.
        """
        seen = list(self.sections)
        keys = [tuple(int(part) for part in ref.split(".")) for ref in seen]
        return keys == sorted(keys)

    def summary(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "sheets": list(self.sheet_names),
            "control_rows": len(self.controls),
            "section_rows": len(self.sections),
            "chapter_header_rows": len(self.chapter_header_rows),
            "distinct_chapters": self.distinct_chapters,
            "distinct_chapter_ids": self.distinct_chapter_ids,
            "last_control_row": self.last_control_row,
            "sections_are_ordered": self.sections_are_ordered,
            "tool_marker_rows": len(self.tool_markers),
            "trailing_rows": list(self.trailing_rows),
            "status_values": list(self.status_values),
            "non_model_column": NON_MODEL_COLUMN,
            "model_columns": list(MODEL_COLUMNS),
        }


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _covers_status_columns(validation) -> bool:
    """La validation porte-t-elle sur les colonnes de statut ?

    Choisir « la première liste trouvée » marchait sur le gabarit actuel, qui
    n'en a qu'une. Une liste ajoutée ailleurs — une phase, une catégorie —
    prendrait sa place et le parseur servirait de faux statuts, sans rien
    signaler.
    """
    letters = {chr(64 + column) for column in (NON_MODEL_COLUMN, *MODEL_COLUMNS)}
    cells = str(validation.sqref or "")
    return any(letter in cells for letter in letters)


def _status_values(sheet) -> list[str]:
    """Valeurs de statut admises, lues dans la validation qui couvre G:I.

    Les recopier en dur ici les figerait à côté du document : si le maître
    d'ouvrage en ajoute une, le gabarit l'accepterait et le parseur l'ignorerait.
    """
    for validation in sheet.data_validations.dataValidation:
        formula = (validation.formula1 or "").strip().strip('"')
        if validation.type == "list" and formula and _covers_status_columns(validation):
            return [part.strip() for part in formula.split(",") if part.strip()]
    return []


def parse_mrn_template(path: str | Path) -> MRNTemplate:
    """Lit le gabarit MRN et en décrit la structure.

    Args:
        path: chemin du classeur modèle.

    Raises:
        FileNotFoundError: si le fichier n'existe pas.
        ValueError: si l'onglet de contrôle est absent — mieux vaut refuser que
            produire un inventaire vide qui passerait pour un gabarit sans
            contrôles.
    """
    import openpyxl

    source = Path(path).expanduser()
    if not source.is_file():
        raise FileNotFoundError(f"Gabarit MRN introuvable : {source}")

    workbook = openpyxl.load_workbook(source, data_only=False)
    if CONTROL_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Onglet {CONTROL_SHEET!r} absent de {source.name} "
            f"(onglets présents : {workbook.sheetnames})."
        )

    sheet = workbook[CONTROL_SHEET]
    controls: list[MRNControlRow] = []
    sections: dict[str, str] = {}
    chapter_headers: list[int] = []
    tool_markers: dict[int, str] = {}

    for row in range(1, sheet.max_row + 1):
        number = _text(sheet.cell(row, NUMBER_COLUMN).value)
        label = _text(sheet.cell(row, LABEL_COLUMN).value)
        marker = _text(sheet.cell(row, TOOL_MARKER_COLUMN).value)
        if marker:
            tool_markers[row] = marker

        if _CONTROL_RE.match(number):
            controls.append(
                MRNControlRow.from_number(
                    number,
                    label=label,
                    row=row,
                    phase=_text(sheet.cell(row, PHASE_COLUMN).value) or None,
                    control_date=sheet.cell(row, DATE_COLUMN).value,
                    tool_marker=marker or None,
                )
            )
        elif _SECTION_RE.match(number):
            sections[number] = label
        elif _CHAPTER_RE.match(number):
            chapter_headers.append(row)

    last_control_row = max((control.row for control in controls), default=0)
    trailing = [
        row
        for row in range(last_control_row + 1, sheet.max_row + 1)
        if any(_text(sheet.cell(row, column).value) for column in range(1, sheet.max_column + 1))
    ]

    categories: list[dict[str, Any]] = []
    if CATEGORY_SHEET in workbook.sheetnames:
        cat_sheet = workbook[CATEGORY_SHEET]
        headers = [_text(cat_sheet.cell(1, c).value) for c in range(1, cat_sheet.max_column + 1)]
        for row in range(2, cat_sheet.max_row + 1):
            values = [_text(cat_sheet.cell(row, c).value) for c in range(1, len(headers) + 1)]
            if any(values):
                categories.append(dict(zip(headers, values, strict=False)))

    workbook.close()
    return MRNTemplate(
        path=source,
        sheet_names=list(workbook.sheetnames),
        controls=controls,
        sections=sections,
        chapter_header_rows=chapter_headers,
        last_control_row=last_control_row,
        tool_markers=tool_markers,
        trailing_rows=trailing,
        status_values=_status_values(sheet),
        categories=categories,
    )
