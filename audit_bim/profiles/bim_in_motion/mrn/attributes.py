"""Inventaire de la table des attributs MRN — exigences par feuille.

Le classeur mélange **plusieurs grilles** : les quatre feuilles d'exigences
n'ont ni les mêmes colonnes, ni la même ligne d'en-tête. Un parseur unique s'y
trompe silencieusement — il lirait la colonne des propriétés d'une feuille dans
la colonne des types d'objet d'une autre.

Le layout est donc **déclaré par feuille**, dans une table testée, et jamais
deviné. Deux erreurs mesurées ont conduit à cette forme :

- lire la colonne ``D`` partout donnait 178 exigences au lieu de 1 013 ; les
  feuilles techniques portent la propriété en ``G``, ``D`` étant leur
  ``IfcTypeObject``, souvent fusionné ou vide ;
- détecter le bloc des phases par mot-clé attribuait à ``VRD-Extérieur`` trois
  colonnes de trop : « Nom de la **pro**priété » et « **Exe**mple de valeur »
  contiennent des libellés de phase par coïncidence. Le bloc est donc repéré
  par sa **position déclarée**, et ses en-têtes sont vérifiés.

Une exigence atomique est **une ligne de propriété attendue**. Les phases et
les maquettes porteuses sont des dimensions attachées à cette ligne, pas des
multiplicateurs : compter les croix donnerait 4 623 « exigences » qui n'en sont
pas.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

__all__ = [
    "ApplicabilityMark",
    "EXPECTED_PHASES",
    "SHEET_LAYOUTS",
    "MRNAttributeRequirement",
    "MRNAttributeTable",
    "SheetLayout",
    "parse_mrn_attribute_table",
]

#: En-têtes attendus du bloc de phases, dans l'ordre. Vérifiés à la lecture :
#: un bloc qui ne les porte pas n'est pas un bloc de phases, et le supposer
#: ferait compter des colonnes métier comme de l'applicabilité.
EXPECTED_PHASES = (
    "ESQ/\nDIAG",
    "APS",
    "APD",
    "PRO",
    "DCE",
    "EXE",
    "DOE",
    "EXPL \nMAINT",
)


@dataclass(frozen=True)
class SheetLayout:
    """Disposition d'une feuille d'exigences. Déclarée, jamais inférée."""

    header_row: int
    property_column: int
    phase_columns: tuple[int, int]
    carrier_columns: tuple[int, int] | None = None
    object_column: int = 1
    ifc_object_column: int | None = None
    ifc_type_object_column: int | None = None
    uniformat_column: int | None = None
    property_label_column: int | None = None
    pset_column: int | None = None
    value_type_column: int | None = None
    example_column: int | None = None

    def phase_range(self) -> range:
        return range(self.phase_columns[0], self.phase_columns[1] + 1)

    def carrier_range(self) -> range:
        if self.carrier_columns is None:
            return range(0)
        return range(self.carrier_columns[0], self.carrier_columns[1] + 1)


#: Une entrée par feuille. Les colonnes sont des index 1-based, comme openpyxl.
SHEET_LAYOUTS: dict[str, SheetLayout] = {
    "Généralités": SheetLayout(
        header_row=3,
        property_column=4,  # D
        carrier_columns=(8, 14),  # H:N
        phase_columns=(15, 22),  # O:V
        object_column=1,
        ifc_object_column=3,
        pset_column=5,
        value_type_column=6,
        example_column=7,
        property_label_column=2,
    ),
    "Gros Oeuvre - CEA": SheetLayout(
        header_row=3,
        property_column=7,  # G
        phase_columns=(10, 17),  # J:Q
        ifc_object_column=3,
        ifc_type_object_column=4,
        uniformat_column=5,
        property_label_column=6,
        pset_column=8,
        example_column=9,
    ),
    "CVC-PLB-SSI-ELEC": SheetLayout(
        header_row=3,
        property_column=7,
        phase_columns=(10, 17),
        ifc_object_column=3,
        ifc_type_object_column=4,
        uniformat_column=5,
        property_label_column=6,
        pset_column=8,
        example_column=9,
    ),
    "VRD-Extérieur": SheetLayout(
        # Même bloc de phases que ses deux sœurs, mais en-têtes une ligne plus
        # haut. C'est ce décalage qui rend une détection générique dangereuse.
        header_row=2,
        property_column=7,
        phase_columns=(10, 17),
        ifc_object_column=3,
        ifc_type_object_column=4,
        uniformat_column=5,
        property_label_column=6,
        pset_column=8,
        example_column=9,
    ),
}

#: Colonnes propagées vers le bas quand elles sont fusionnées : elles décrivent
#: un groupe de lignes, et les laisser vides ferait perdre l'objet auquel une
#: exigence se rattache.
_FORWARD_FILLED = (
    "object_column",
    "ifc_object_column",
    "ifc_type_object_column",
    "uniformat_column",
)


@dataclass(frozen=True)
class ApplicabilityMark:
    """Une croix, avec sa nuance et l'axe auquel elle appartient.

    Les listes ``applicable_phases`` et ``carrier_models`` restent pratiques
    pour un filtrage, mais elles aplatissent : une fois réduites à des libellés,
    plus rien ne dit quelle exigence portait ``x+``. Or c'est précisément la
    nuance qu'on a décidé de conserver.
    """

    axis: str  # "phase" | "carrier"
    label: str
    marker: str
    marker_kind: str


@dataclass(frozen=True)
class MRNAttributeRequirement:
    """Une **ligne de propriété attendue**, avec ses dimensions d'applicabilité."""

    sheet: str
    row: int
    property_name: str
    object_name: str = ""
    ifc_object: str = ""
    ifc_type_object: str = ""
    uniformat: str = ""
    property_label: str = ""
    pset: str = ""
    value_type: str = ""
    example: str = ""
    carrier_models: list[str] = field(default_factory=list)
    applicable_phases: list[str] = field(default_factory=list)
    applicability: list[ApplicabilityMark] = field(default_factory=list)
    carrier_scope: str = "non_specifie"

    @property
    def nuanced_marks(self) -> list[ApplicabilityMark]:
        """Croix portant une nuance — ``x+`` aujourd'hui."""
        return [m for m in self.applicability if m.marker_kind != "applicable"]

    @property
    def applicability_cells(self) -> int:
        """Croix portées par cette ligne. Diagnostic, **jamais** un nombre d'exigences."""
        return len(self.carrier_models) + len(self.applicable_phases)


@dataclass(frozen=True)
class MRNAttributeTable:
    """Inventaire complet : les exigences, et de quelles feuilles elles viennent."""

    path: Path
    requirements: list[MRNAttributeRequirement]
    sheet_names: list[str]
    hidden_sheets: list[str]
    marker_variants: dict[str, int] = field(default_factory=dict)
    formula_applicability_cells: int = 0

    def by_sheet(self, name: str) -> list[MRNAttributeRequirement]:
        return [req for req in self.requirements if req.sheet == name]

    def summary(self) -> dict[str, Any]:
        per_sheet = {}
        for name in SHEET_LAYOUTS:
            rows = self.by_sheet(name)
            per_sheet[name] = {
                "requirement_rows": len(rows),
                "applicability_cells": sum(req.applicability_cells for req in rows),
            }
        return {
            "path": str(self.path),
            "requirement_rows": len(self.requirements),
            "applicability_cells_effective": sum(
                req.applicability_cells for req in self.requirements
            ),
            "applicability_marker_variants": dict(self.marker_variants),
            "formula_applicability_cells": self.formula_applicability_cells,
            "per_sheet": per_sheet,
            "requirement_sheets": list(SHEET_LAYOUTS),
            "hidden_sheets": list(self.hidden_sheets),
        }


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


#: Marqueurs d'applicabilite reconnus, et leur nature. ``x+`` est une croix
#: nuancee : la compter comme un ``x`` ordinaire aplatirait l'information, ne
#: pas la compter perdrait une exigence visible dans le classeur.
APPLICABILITY_MARKERS = {
    "x": "applicable",
    "x+": "applicable_with_note",
}


def _marker_kind(value: Any) -> str | None:
    return APPLICABILITY_MARKERS.get(_text(value).lower())


def _marked(value: Any) -> bool:
    return _marker_kind(value) is not None


def _verify_phase_headers(sheet, layout: SheetLayout, name: str) -> list[str]:
    """Lit les en-têtes du bloc de phases et refuse s'ils ne correspondent pas.

    Sans ce refus, un décalage de colonnes ferait compter des libellés métier
    comme de l'applicabilité — c'est exactement ce qui est arrivé sur
    ``VRD-Extérieur`` avec « Nom de la propriété » et « Exemple de valeur ».
    """
    headers = [_text(sheet.cell(layout.header_row, col).value) for col in layout.phase_range()]
    expected = [phase.strip() for phase in EXPECTED_PHASES]
    if [h.strip() for h in headers] != expected:
        raise ValueError(
            f"Feuille {name!r} : le bloc de phases déclaré ne porte pas les en-têtes "
            f"attendus.\n  attendu : {expected}\n  lu      : {[h.strip() for h in headers]}\n"
            f"Vérifier header_row et phase_columns avant de compter quoi que ce soit."
        )
    return headers


def parse_mrn_attribute_table(path: str | Path) -> MRNAttributeTable:
    """Lit la table des attributs MRN et en extrait les exigences.

    Raises:
        FileNotFoundError: fichier absent.
        ValueError: une feuille d'exigences manque, ou son bloc de phases ne
            porte pas les en-têtes attendus.
    """
    import openpyxl

    source = Path(path).expanduser()
    if not source.is_file():
        raise FileNotFoundError(f"Table des attributs MRN introuvable : {source}")

    # Deux lectures : les VALEURS font foi — un livrable opposable doit refleter
    # ce qu'un lecteur voit dans Excel — et les FORMULES servent au diagnostic.
    # Sans la seconde, une cellule dont l'applicabilite vient d'un calcul sans
    # valeur mise en cache disparaitrait sans bruit.
    workbook = openpyxl.load_workbook(source, data_only=True)
    formulas = openpyxl.load_workbook(source, data_only=False)
    missing = [name for name in SHEET_LAYOUTS if name not in workbook.sheetnames]
    if missing:
        raise ValueError(
            f"Feuilles d'exigences absentes de {source.name} : {missing}. "
            f"Un inventaire partiel passerait pour une table incomplète."
        )

    requirements: list[MRNAttributeRequirement] = []
    marker_variants: dict[str, int] = {}
    formula_cells = 0
    missing_values: list[str] = []

    for name, layout in SHEET_LAYOUTS.items():
        sheet = workbook[name]
        formula_sheet = formulas[name]
        phase_headers = _verify_phase_headers(sheet, layout, name)
        carrier_headers = [
            _text(sheet.cell(layout.header_row, col).value) for col in layout.carrier_range()
        ]
        carried: dict[str, str] = {}

        for row in range(layout.header_row + 1, sheet.max_row + 1):
            for attr in _FORWARD_FILLED:
                column = getattr(layout, attr)
                if column:
                    value = _text(sheet.cell(row, column).value)
                    if value:
                        carried[attr] = value

            property_name = _text(sheet.cell(row, layout.property_column).value)
            if not property_name:
                continue

            phases, carriers, marks = [], [], []
            for headers, columns, bucket in (
                (phase_headers, layout.phase_range(), phases),
                (carrier_headers, layout.carrier_range(), carriers),
            ):
                for index, col in enumerate(columns):
                    value = sheet.cell(row, col).value
                    raw = _text(formula_sheet.cell(row, col).value)
                    if raw.startswith("="):
                        formula_cells += 1
                        if _text(value) == "":
                            missing_values.append(f"{name}!{chr(64 + col)}{row}")
                            continue
                    marker = _text(value).lower()
                    kind = _marker_kind(value)
                    if kind:
                        marker_variants[marker] = marker_variants.get(marker, 0) + 1
                        bucket.append(headers[index])
                        marks.append(
                            ApplicabilityMark(
                                axis="phase" if bucket is phases else "carrier",
                                label=headers[index],
                                marker=marker,
                                marker_kind=kind,
                            )
                        )

            # Liaison explicite des variables de boucle : une fermeture qui les
            # capture par référence lirait la dernière feuille parcourue, pas
            # celle en cours — un défaut qui ne se voit qu'au dernier tour.
            def _cell(attr: str, _sheet=sheet, _layout=layout, _row=row) -> str:
                column = getattr(_layout, attr)
                return _text(_sheet.cell(_row, column).value) if column else ""

            requirements.append(
                MRNAttributeRequirement(
                    sheet=name,
                    row=row,
                    property_name=property_name,
                    object_name=carried.get("object_column", ""),
                    ifc_object=carried.get("ifc_object_column", ""),
                    ifc_type_object=carried.get("ifc_type_object_column", ""),
                    uniformat=carried.get("uniformat_column", ""),
                    property_label=_cell("property_label_column"),
                    pset=_cell("pset_column"),
                    value_type=_cell("value_type_column"),
                    example=_cell("example_column"),
                    carrier_models=carriers,
                    applicable_phases=phases,
                    applicability=marks,
                    # Aucun héritage depuis Généralités : une feuille sans
                    # colonnes maquette ne dit rien de ses porteurs, et
                    # l'inventer produirait une exigence attribuée à tort.
                    carrier_scope=("declare" if layout.carrier_columns else "non_specifie"),
                )
            )

    if missing_values:
        raise ValueError(
            "formula_value_missing : des cellules d'applicabilite portent une "
            f"formule sans valeur calculee lisible : {missing_values[:10]}. "
            "Rouvrir et enregistrer le classeur dans Excel, ou fournir un export "
            "avec valeurs — les compter comme vides fausserait l'inventaire."
        )

    hidden = [ws.title for ws in workbook.worksheets if ws.sheet_state != "visible"]
    sheet_names = list(workbook.sheetnames)
    workbook.close()
    formulas.close()
    return MRNAttributeTable(
        path=source,
        requirements=requirements,
        sheet_names=sheet_names,
        hidden_sheets=hidden,
        marker_variants=marker_variants,
        formula_applicability_cells=formula_cells,
    )
