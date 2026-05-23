"""Lecture des classifications validées par l'auditeur dans un XLSX d'audit.

L'auditeur télécharge l'annexe ``audit_*_annexes.xlsx``, édite l'onglet
*Classifications suggérées* en colonne « Suggestion 1 — code » :

- garder la valeur suggérée  → la classification sera appliquée
- modifier le code           → on applique le code corrigé
- effacer la cellule         → ligne ignorée

Le reader extrait la liste des items ``{uuid, code, label, system}`` à
appliquer en aval via ``apply_classifications``.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from ..requirements._openpyxl_compat import patch_openpyxl

patch_openpyxl()

SHEET_NAME = "Classifications suggérées"


def read_classifications_from_xlsx(xlsx_path: str | Path) -> list[dict]:
    """Lit l'onglet *Classifications suggérées* et retourne les choix retenus.

    Args:
        xlsx_path: chemin de l'annexe XLSX d'audit (potentiellement modifiée
            par l'auditeur).

    Returns:
        Liste de dicts ``{uuid, code, label, system, ifc_type, name,
        confidence}`` — ne contient que les lignes où la colonne
        *Suggestion 1 — code* est non vide.

    Raises:
        FileNotFoundError: si le fichier n'existe pas.
        ValueError: si l'onglet ou une colonne attendue est introuvable.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"Onglet '{SHEET_NAME}' introuvable dans {path.name}. "
                f"Onglets présents : {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]

        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        # En-têtes attendues (cf. xlsx_annex._write_classification_suggestions)
        def _idx(label: str) -> int:
            try:
                return header.index(label)
            except ValueError:
                raise ValueError(
                    f"Colonne '{label}' absente dans l'onglet "
                    f"'{SHEET_NAME}'. En-têtes lues : {header}"
                )

        i_uuid = _idx("UUID")
        i_class = _idx("Classe IFC")
        i_name = _idx("Nom")
        i_code = _idx("Suggestion 1 — code")
        i_label = _idx("Sug. 1 — libellé")
        i_conf = _idx("Conf. 1")

        items: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            uuid = row[i_uuid] if i_uuid < len(row) else None
            code = row[i_code] if i_code < len(row) else None
            if not uuid or not code:
                continue
            code_str = str(code).strip()
            if not code_str:
                continue
            items.append(
                {
                    "uuid": str(uuid).strip(),
                    "ifc_type": (str(row[i_class]).strip() if row[i_class] else ""),
                    "name": (str(row[i_name]).strip() if row[i_name] else ""),
                    "code": code_str,
                    "label": (str(row[i_label]).strip() if row[i_label] else code_str),
                    "system": "UniFormat II",
                    "confidence": (
                        float(row[i_conf]) if row[i_conf] is not None else None
                    ),
                }
            )
        return items
    finally:
        wb.close()
