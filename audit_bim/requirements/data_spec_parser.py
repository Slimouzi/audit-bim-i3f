"""Parseur de l'annexe « Spécification des données » (xlsx).

L'annexe est structurée comme suit (1 feuille « ANNEXE 2 ») :

| col | contenu                                                                |
|-----|------------------------------------------------------------------------|
| A   | Thème (Générale, Cloisons, CVC, Électricité…)                          |
| B   | Objet (Projet, Site, Bâtiment, Étage, Zone, Pièce, Mur, Porte…)        |
| C   | Définition de l'objet                                                  |
| D   | Classe IFC (IfcProject, IfcSite, IfcSpace…)                            |
| E   | Propriété / Document attendu                                           |
| F   | Pset ou attribut IFC porteur (Pset_SpaceCommon, IfcName, …)            |
| G..M| Phases BIM (APS, AVP, PRO, DCE, EXE, DOE, GESTION) — "X" si requise    |
| N   | Niveau de détail géométrique / commentaire                             |
| O   | Précision usage 3F                                                     |
| P   | Outils 3F                                                              |
| Q   | Req 3F                                                                 |

Les cellules de la colonne A et B sont *mergées* sur plusieurs lignes : on
forward-fill ces colonnes pour reconstituer le thème/objet courant.

Les sous-en-têtes « Documents », « Propriétés » réinitialisent la *catégorie*
de la ligne suivante (document attendu vs propriété IFC).
"""
from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

import openpyxl

from ..audit.ifc_hierarchy import normalize_catalog_class
from ._openpyxl_compat import patch_openpyxl
from .models import BIMPhase, PropertySpec

patch_openpyxl()

# Index de colonnes (0-based) — robustes au schéma observé V3.7
COL_THEME = 0
COL_OBJET = 1
COL_DEFINITION = 2
COL_IFC_CLASS = 3
COL_PROPERTY = 4
COL_PSET = 5
COL_PHASES = list(range(6, 13))  # G..M
COL_COMMENT = 13
COL_USAGE_3F = 14

PHASE_ORDER = BIMPhase.ordered()

# Marqueurs de sous-en-tête (réinitialisent la "kind" de la ligne suivante)
KIND_HEADERS = {
    "documents": "document",
    "propriétés": "property",
    "proprietes": "property",
}


def _iter_rows(xlsx_path: Path) -> Iterator[tuple]:
    # Mode non read_only : certains fichiers I3F ont des AutoFilter custom que
    # le parseur streaming d'openpyxl refuse de charger.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        yield from ws.iter_rows(values_only=True)
    finally:
        wb.close()


def _is_header_row(row: tuple) -> bool:
    """Détecte la ligne d'en-tête principale (« Objet | Définition | Classe IFC »)."""
    cells = [str(c).strip().lower() if c is not None else "" for c in row[:6]]
    return "objet" in cells and "classe ifc" in cells


def _detect_phase_columns(row: tuple) -> dict[BIMPhase, int] | None:
    """Repère sur quelles colonnes se trouvent APS/AVP/…/GESTION.

    L'ordre peut varier selon versions ; on remappe sur les labels.
    """
    cells = {
        i: str(c).strip().upper() if c is not None else "" for i, c in enumerate(row)
    }
    mapping: dict[BIMPhase, int] = {}
    for phase in PHASE_ORDER:
        token_main = phase.value
        token_alt = f"BIM {phase.value}"
        for i, txt in cells.items():
            if txt == token_main or txt == token_alt:
                mapping[phase] = i
                break
    return mapping if mapping else None


def _required_phases(row: tuple, phase_cols: dict[BIMPhase, int]) -> list[BIMPhase]:
    out: list[BIMPhase] = []
    for phase, col in phase_cols.items():
        if col >= len(row):
            continue
        v = row[col]
        if v is None:
            continue
        s = str(v).strip().upper()
        # « X » est le marqueur explicite ; certains chiffres (LOD) servent
        # aussi à indiquer que la donnée géométrique est attendue.
        if s == "X" or s.startswith("X"):
            out.append(phase)
        elif s.isdigit() and int(s) >= 1:
            out.append(phase)
    return out


def parse_data_spec(xlsx_path: str | Path) -> list[PropertySpec]:
    """Parse l'annexe Spécification des données.

    Args:
        xlsx_path: Chemin vers le fichier xlsx.

    Returns:
        Liste de PropertySpec, une par ligne exploitable.
    """
    xlsx_path = Path(xlsx_path)
    specs: list[PropertySpec] = []

    current_theme: str = ""
    current_objet: str = ""
    current_ifc_class: str = ""
    current_kind: str = "property"
    phase_cols: dict[BIMPhase, int] | None = None
    header_seen = False

    for row in _iter_rows(xlsx_path):
        if not row or not any(row):
            continue

        # Détection de l'en-tête (et ré-essai si non trouvé sur les sous-headers)
        if not header_seen and _is_header_row(row):
            header_seen = True
            phase_cols = _detect_phase_columns(row) or phase_cols
            continue
        if phase_cols is None:
            # Cherche une ligne portant les noms de phases (juste après l'en-tête)
            phase_cols = _detect_phase_columns(row) or None
            if phase_cols:
                continue

        # Forward-fill thème et objet
        theme = row[COL_THEME] if COL_THEME < len(row) else None
        objet = row[COL_OBJET] if COL_OBJET < len(row) else None
        ifc_class = row[COL_IFC_CLASS] if COL_IFC_CLASS < len(row) else None
        prop_name = row[COL_PROPERTY] if COL_PROPERTY < len(row) else None
        pset = row[COL_PSET] if COL_PSET < len(row) else None
        comment = row[COL_COMMENT] if COL_COMMENT < len(row) else None
        usage = row[COL_USAGE_3F] if COL_USAGE_3F < len(row) else None

        if theme:
            current_theme = str(theme).strip()
        if objet:
            current_objet = str(objet).strip()
        if ifc_class:
            current_ifc_class = str(ifc_class).strip()

        # Sous-en-tête « Documents » / « Propriétés »
        if prop_name and not pset:
            tok = str(prop_name).strip().lower()
            if tok in KIND_HEADERS:
                current_kind = KIND_HEADERS[tok]
                continue

        # Une ligne n'est exploitable que si on a au moins une classe IFC + une
        # propriété (le pset peut être vide pour les documents).
        if not current_ifc_class or not prop_name:
            continue
        if not str(current_ifc_class).lower().startswith("ifc"):
            # Ligne d'intro / commentaire générique
            continue

        required = (
            _required_phases(row, phase_cols) if phase_cols else []
        )

        # Une ligne du CCH peut référencer plusieurs classes IFC en une
        # seule cellule (ex: "IfcDuctFittingType\nIfcDuctSegmentType") ou
        # contenir un suffixe métier (« IfcCovering_CEILING »). On déplie
        # en une PropertySpec par classe IFC normalisée.
        for ifc_class in normalize_catalog_class(str(current_ifc_class)):
            specs.append(
                PropertySpec(
                    theme=current_theme or "Générale",
                    objet=current_objet or "",
                    ifc_class=ifc_class,
                    property_name=str(prop_name).strip(),
                    pset_or_attribute=str(pset).strip() if pset else None,
                    kind="document" if current_kind == "document" else "property",
                    required_phases=required,
                    comment=str(comment).strip() if comment else None,
                    usage_3f=str(usage).strip() if usage else None,
                    ref_cch="Chap 6.2",
                )
            )

    return specs
