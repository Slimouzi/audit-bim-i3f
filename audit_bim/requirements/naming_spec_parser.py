"""Parseur de l'annexe « Nommage » (xlsx).

Deux feuilles :
1. *Annexe 6,3,1 Site bat étage* — règles pour IfcProject, IfcSite, IfcBuilding,
   IfcBuildingStorey + **liste fermée des noms d'étages** admis.
2. *Annexe 6,3,2 zones et pièces* — règles pour IfcZone, IfcSpace +
   **liste fermée des types de zones** et **liste des noms de pièces**.

Le format n'est pas tabulaire pur (cellules mergées, sous-tables, exemples
graphiques). On extrait :

- les *règles* (NamingRule) sur les propriétés à renseigner pour chaque objet,
- les *listes fermées* (StoreyName, ZoneSpec, RoomSpec) qui servent de
  référentiel de validation lors de l'audit.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl

from .models import NamingRule, RoomSpec, StoreyName, ZoneSpec
from ._openpyxl_compat import patch_openpyxl

patch_openpyxl()

# Patterns I3F codifiés (CCH chap 6.3.1)
SITE_NAME_PATTERN = r"^\d{4}[LP]$"  # 1802L (logements) | 1802P (parkings)
BUILDING_NAME_PATTERN = r"^\d{4}[LP]-[A-Z]([0-9]+)?$"  # 1802L-A, 1802L-A2…
ZONE_LOGEMENT_NAME_PATTERN = r"^\d{4}[LP]-\d{3,4}$"  # 1802L-1101


def _open(xlsx_path: Path) -> openpyxl.Workbook:
    # Mode non read_only : voir data_spec_parser._iter_rows pour la raison.
    return openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)


def _extract_storey_names(ws) -> list[StoreyName]:
    """Liste les noms d'étages depuis la feuille 6,3,1.

    Stratégie tolérante : on collecte toutes les cellules dont la valeur
    textuelle matche un libellé d'étage typique (sous-sol, RDC, étages, combles,
    toiture). Les étages numérotés à compteur (TOITURE 01…) sont représentés
    par un pattern.
    """
    canonical = {
        "REZ-DE-CHAUSSEE",
        "REZ-DE-JARDIN",
        "COMBLES",
        "TOITURE",
    }
    storey_re = re.compile(
        r"^\s*("
        r"\d{1,2}[EÈ]ME\s+SOUS-SOL"
        r"|1ER\s+SOUS-SOL"
        r"|REZ-DE-CHAUSSEE|REZ-DE-JARDIN"
        r"|ENTRESOL(\s+\d{1,2})?"
        r"|1ER\s+ETAGE"
        r"|\d{1,2}[EÈ]ME\s+ETAGE"
        r"|COMBLES|TOITURE(\s+\d{1,2})?"
        r")\s*$",
        re.IGNORECASE,
    )
    seen: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            for line in cell.splitlines():
                line = line.strip().upper()
                if storey_re.fullmatch(line):
                    if line not in seen:
                        seen.append(line)
                elif line in canonical and line not in seen:
                    seen.append(line)
    out = [
        StoreyName(
            name=n,
            pattern=(
                r"^TOITURE(\s+\d{1,2})?$"
                if n == "TOITURE"
                else r"^ENTRESOL(\s+\d{1,2})?$" if n == "ENTRESOL" else None
            ),
        )
        for n in seen
    ]
    return out


def _extract_zone_and_room_specs(ws) -> tuple[list[ZoneSpec], list[RoomSpec]]:
    """Extrait les listes Zones / Pièces depuis la feuille 6,3,2.

    Le tableau de référence commence à la ligne 27 environ :
        col A : Liste noms zones
        col B : Liste types zones
        col C : PP | PC
        col D : Définition

        col F : Liste noms pièces
        col G : Liste types pièces
        col H : PP | PC
        col I : Définition
        col K (variable) : Type surface (SHAB/SU)
    """
    zones: list[ZoneSpec] = []
    rooms: list[RoomSpec] = []
    in_table = False
    type_surface_col: Optional[int] = None

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        # Détection de la ligne d'en-tête du tableau
        joined = " | ".join(str(c) for c in cells if c)
        joined_l = joined.lower()
        if "liste des types de zones" in joined_l or "liste de types des zones" in joined_l:
            in_table = True
            # Repérer la colonne « Type de surface » si présente
            for i, c in enumerate(cells):
                if c and "type de surface" in str(c).lower():
                    type_surface_col = i
            continue
        if not in_table:
            continue

        # Tableau ZONES (cols ~A..D)
        z_name = cells[0] if len(cells) > 0 else None
        z_type = cells[1] if len(cells) > 1 else None
        z_loc = cells[2] if len(cells) > 2 else None
        z_def = cells[3] if len(cells) > 3 else None
        if z_type and isinstance(z_type, str) and z_type.strip().lower().startswith("zone"):
            zones.append(
                ZoneSpec(
                    name=str(z_name).strip() if z_name else None,
                    type_label=z_type.strip(),
                    localisation=(str(z_loc).strip().upper() if z_loc else "PP"),
                    definition=str(z_def).strip() if z_def else None,
                )
            )

        # Tableau PIÈCES (à droite, cols ~F..I + K)
        # On scanne les colonnes 5..10 pour résister aux décalages
        for start in range(5, min(12, len(cells))):
            r_name = cells[start]
            r_type = cells[start + 1] if start + 1 < len(cells) else None
            r_loc = cells[start + 2] if start + 2 < len(cells) else None
            r_def = cells[start + 3] if start + 3 < len(cells) else None
            if (
                r_name
                and r_type
                and isinstance(r_name, str)
                and isinstance(r_type, str)
                and r_name.strip().isupper()
                and r_loc in ("PP", "PC")
            ):
                surf = None
                if type_surface_col is not None and type_surface_col < len(cells):
                    sv = cells[type_surface_col]
                    if sv and isinstance(sv, str):
                        surf = sv.strip()
                rooms.append(
                    RoomSpec(
                        name=r_name.strip(),
                        type_label=r_type.strip(),
                        localisation=r_loc,
                        surface_type=surf,
                        definition=str(r_def).strip() if r_def else None,
                    )
                )
                break  # une seule pièce par ligne

    # Dédoublonnage stable
    def _dedup(items, key):
        seen, out = set(), []
        for it in items:
            k = key(it)
            if k in seen:
                continue
            seen.add(k)
            out.append(it)
        return out

    return _dedup(zones, lambda z: (z.name, z.type_label)), _dedup(
        rooms, lambda r: (r.name, r.localisation)
    )


def parse_naming_spec(
    xlsx_path: str | Path,
) -> tuple[list[NamingRule], list[StoreyName], list[ZoneSpec], list[RoomSpec]]:
    """Parse l'annexe Nommage.

    Returns:
        (naming_rules, storey_names, zone_specs, room_specs)
    """
    xlsx_path = Path(xlsx_path)
    wb = _open(xlsx_path)
    try:
        storey_names: list[StoreyName] = []
        zone_specs: list[ZoneSpec] = []
        room_specs: list[RoomSpec] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lname = sheet_name.lower()
            if "site" in lname or "bat" in lname or "étage" in lname or "etage" in lname:
                storey_names = _extract_storey_names(ws)
            if "zone" in lname or "pi" in lname:
                z, r = _extract_zone_and_room_specs(ws)
                zone_specs = z
                room_specs = r
    finally:
        wb.close()

    naming_rules: list[NamingRule] = [
        NamingRule(
            objet="Projet",
            ifc_class="IfcProject",
            ifc_attribute="LongName",
            pattern=None,
            max_length=30,
            comment="Nom détaillé du programme — fourni par 3F.",
            ref_cch="Chap 6.3.1",
        ),
        NamingRule(
            objet="Site",
            ifc_class="IfcSite",
            ifc_attribute="Name",
            pattern=SITE_NAME_PATTERN,
            comment="Codification I3F : 4 chiffres + L (logement) ou P (parking).",
            ref_cch="Chap 6.3.1",
        ),
        NamingRule(
            objet="Bâtiment",
            ifc_class="IfcBuilding",
            ifc_attribute="Name",
            pattern=BUILDING_NAME_PATTERN,
            max_length=30,
            comment="Codification I3F + tiret + lettre du bâtiment (A, B, C…).",
            ref_cch="Chap 6.3.1",
        ),
        NamingRule(
            objet="Étage",
            ifc_class="IfcBuildingStorey",
            ifc_attribute="Name",
            allowed_values=[s.name for s in storey_names],
            comment="Liste fermée du CCH chap 6.3.1.",
            ref_cch="Chap 6.3.1",
        ),
        NamingRule(
            objet="Zone (logement)",
            ifc_class="IfcZone",
            ifc_attribute="Name",
            pattern=ZONE_LOGEMENT_NAME_PATTERN,
            comment="Nom usuel du logement (Exemple : 1802L-1101).",
            ref_cch="Chap 6.3.2.1",
        ),
        NamingRule(
            objet="Zone — type",
            ifc_class="IfcZone",
            ifc_attribute="ObjectType",
            allowed_values=[z.type_label for z in zone_specs],
            comment="Type de zone obligatoire (Zone Logement T2, Zone Bureaux…).",
            ref_cch="Chap 6.3.2",
        ),
        NamingRule(
            objet="Pièce",
            ifc_class="IfcSpace",
            ifc_attribute="LongName",
            allowed_values=sorted({r.name for r in room_specs}),
            case_sensitive=True,
            comment=(
                "Nom de pièce en majuscules, conforme à la liste du CCH "
                "(BALCON, CHAMBRE, CUISINE…). Suffixe numérique autorisé "
                "(CHAMBRE 01, CHAMBRE 02)."
            ),
            ref_cch="Chap 6.3.2",
        ),
    ]
    return naming_rules, storey_names, zone_specs, room_specs
