"""Parseur de l'annexe « Spécification des données » — format CCH 2026 (Vbéta2).

Le format 2026 réorganise entièrement les colonnes par rapport au V3.7
(cf. ``data_spec_parser``) :

| col | contenu                                                            |
|-----|---------------------------------------------------------------------|
| B   | Objet (Site, Bâtiment, Zone, Pièce, Murs, …) — cellules mergées     |
| C   | Classe IFC (IfcSite, IfcSpace, …) — mergée, parfois multi-classes   |
| D..J| Phases APS / AVP / PRO / DCE / EXE / DOE / GESTION — « X » si requis |
| K   | « Attendu à partir de » (APS / PC, APD / AVP, PRO, DOE…)            |
| M   | Propriété de l'objet ou document attendu                            |
| N   | Localisation dans l'IFC (IfcName, IfcLongName, Pset_3F, …)          |
| O   | Précision de l'usage 3F                                             |
| P   | Précision sur la propriété (liste de données CCH…)                   |
| Q   | Outils 3F                                                            |
| S   | Req 3F (« X »)                                                       |

Particularités gérées :
- sections « Documents » (col M) → kind="document" ;
- lignes sans « X » de phase mais avec « Attendu à partir de » → l'exigence
  s'applique de cette phase jusqu'à GESTION ;
- localisations normalisées vers les attributs natifs (IfcName → Name,
  IfcLongName → LongName, Type Name → ObjectType) et les Psets
  (« Pset I3F » / « Pset 3F » → Pset_3F) ;
- fichier avec AutoFilter custom qui plante openpyxl : le patch
  ``_openpyxl_compat`` est appliqué comme pour le V3.7.
"""

from __future__ import annotations

import re
from collections.abc import Iterator
from pathlib import Path

import openpyxl

from ..audit.ifc_hierarchy import normalize_catalog_class
from ._openpyxl_compat import patch_openpyxl
from .models import BIMPhase, PropertySpec

patch_openpyxl()

# Index de colonnes (0-based) — schéma observé sur « CCH 2026 Vbéta2 »
COL_OBJET = 1
COL_IFC_CLASS = 2
COL_PHASES_START = 3  # D..J dans l'ordre BIMPhase.ordered()
COL_ATTENDU = 10
COL_PROPERTY = 12
COL_LOCALISATION = 13
COL_USAGE_3F = 14
COL_PRECISION = 15

PHASE_ORDER = BIMPhase.ordered()

# Lignes de M qui sont des en-têtes de section, pas des exigences
_SECTION_HEADERS = {
    "propriétés de l'objet",
    "propriétés des objets",
    "localisation dans l'ifc",
    "autres propriétés particulières",
    "autres propriétés particulière",
}

# « Attendu à partir de » → première phase d'exigence
_ATTENDU_PHASE = {
    "aps": BIMPhase.APS,
    "pc": BIMPhase.APS,
    "permis": BIMPhase.APS,
    "avp": BIMPhase.AVP,
    "apd": BIMPhase.AVP,
    "pro": BIMPhase.PRO,
    "dce": BIMPhase.DCE,
    "exe": BIMPhase.EXE,
    "doe": BIMPhase.DOE,
    "gestion": BIMPhase.GESTION,
}


def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\n", " ")).strip()


def _iter_rows(xlsx_path: Path) -> Iterator[tuple]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        yield from ws.iter_rows(values_only=True)
    finally:
        wb.close()


def is_2026_format(xlsx_path: str | Path) -> bool:
    """Détecte le format 2026 : ligne d'en-tête avec « Objet » en colonne B
    et « Classe IFC » en colonne C (le V3.7 les porte en B et D)."""
    for row in _iter_rows(Path(xlsx_path)):
        cells = [_clean(c).lower() for c in (list(row) + [None] * 4)[:4]]
        if cells[1] == "objet" and cells[2] == "classe ifc":
            return True
        if cells[1] == "objet" and cells[3] == "classe ifc":
            return False
    return False


def _phases_from_marks(row: tuple) -> list[BIMPhase]:
    out: list[BIMPhase] = []
    for i, phase in enumerate(PHASE_ORDER):
        col = COL_PHASES_START + i
        if col < len(row) and _clean(row[col]).upper().startswith("X"):
            out.append(phase)
    return out


def _phases_from_attendu(attendu: str) -> list[BIMPhase]:
    low = attendu.lower()
    for token, phase in _ATTENDU_PHASE.items():
        if token in low:
            idx = PHASE_ORDER.index(phase)
            return PHASE_ORDER[idx:]
    return []


def _normalize_pset(loc: str) -> tuple[str | None, str]:
    """(pset_or_attribute normalisé, kind) depuis la colonne Localisation."""
    if not loc:
        return None, "property"
    low = loc.lower()
    if "basequantit" in low or re.search(r"\bqto\b", low):
        # Conserve le chemin complet (« BaseQuantites/NetArea » →
        # « BaseQuantities/NetArea ») : resolve_value le résout en
        # Pset/Property avec repli ArchiCAD sur les surfaces.
        m = re.search(r"basequantit[ei]s?\s*/\s*([A-Za-z/]+)", loc, re.I)
        if m:
            return f"BaseQuantities/{m.group(1).strip()}", "quantity"
        return "BaseQuantities", "quantity"
    if "longname" in low:
        return "LongName", "property"
    if "ifcname" in low or low == "name":
        return "Name", "property"
    if "type name" in low or "typename" in low:
        return "ObjectType", "property"
    if "ifcmaterial" in low or "materi" in low:
        return "IfcMaterial", "property"
    # « Pset I3F », « Pset 3F », « Pset_3F » → Pset_3F (pset propriétaire)
    if re.search(r"pset[ _]?i?3f", low):
        return "Pset_3F", "property"
    m = re.search(r"(Pset_[A-Za-z0-9]+)", loc)
    if m:
        return m.group(1), "property"
    return loc or None, "property"


def parse_data_spec_2026(xlsx_path: str | Path) -> list[PropertySpec]:
    """Parse l'annexe Spécification des données au format CCH 2026 (Vbéta2).

    Args:
        xlsx_path: Chemin vers le fichier xlsx.

    Returns:
        Liste de PropertySpec (une par ligne exploitable et par classe IFC).
    """
    specs: list[PropertySpec] = []
    current_theme = "Générale"
    current_objet = ""
    current_ifc_raw = ""
    in_documents = False

    for row in _iter_rows(Path(xlsx_path)):
        if not row or not any(row):
            continue
        row = tuple(row) + (None,) * (20 - len(row))

        objet = _clean(row[COL_OBJET])
        # La cellule Classe IFC garde ses sauts de ligne : c'est le séparateur
        # multi-classes attendu par normalize_catalog_class.
        ifc_raw = str(row[COL_IFC_CLASS]).strip() if row[COL_IFC_CLASS] else ""
        prop = _clean(row[COL_PROPERTY])
        loc = _clean(row[COL_LOCALISATION])
        attendu = _clean(row[COL_ATTENDU])
        usage = _clean(row[COL_USAGE_3F])
        precision = _clean(row[COL_PRECISION])

        # Thème éventuel en colonne A (rare dans ce format : « Générale »)
        theme = _clean(row[0])
        if theme and len(theme) < 40 and not theme.lower().startswith("certains"):
            current_theme = theme

        low_prop = prop.lower()
        if low_prop == "documents":
            in_documents = True
            continue

        # Nouveau bloc objet (les en-têtes de tableau portent « Classe IFC »)
        if objet and ifc_raw and "classe ifc" not in ifc_raw.lower():
            current_objet = objet
            current_ifc_raw = ifc_raw
            in_documents = False
        elif objet and not ifc_raw:
            current_objet = objet

        if not prop or low_prop in _SECTION_HEADERS or low_prop.startswith("propriété"):
            continue

        # Documents : section projet initiale (rattachée à IfcProject) ou
        # sous-section « Documents » d'un bloc objet (fiches techniques,
        # spécifications de pose… rattachées à la classe de l'objet courant).
        if in_documents or (current_objet == "Projet" and not loc):
            phases = _phases_from_marks(row) or _phases_from_attendu(attendu)
            doc_classes = (normalize_catalog_class(current_ifc_raw) if current_ifc_raw else []) or [
                "IfcProject"
            ]
            for doc_class in doc_classes:
                specs.append(
                    PropertySpec(
                        theme=current_theme,
                        objet=current_objet or "Projet",
                        ifc_class=doc_class,
                        property_name=prop,
                        pset_or_attribute=None,
                        kind="document",
                        required_phases=phases,
                        comment=precision or None,
                        usage_3f=usage or None,
                        ref_cch="CCH 2026 — Annexe 2",
                    )
                )
            continue

        if not current_ifc_raw:
            continue

        phases = _phases_from_marks(row) or _phases_from_attendu(attendu)
        pset, kind = _normalize_pset(loc)

        for ifc_class in normalize_catalog_class(current_ifc_raw):
            specs.append(
                PropertySpec(
                    theme=current_theme,
                    objet=current_objet,
                    ifc_class=ifc_class,
                    property_name=prop,
                    pset_or_attribute=pset,
                    kind=kind,
                    required_phases=phases,
                    comment=precision or None,
                    usage_3f=usage or None,
                    ref_cch="CCH 2026 — Annexe 2",
                )
            )

    return specs
