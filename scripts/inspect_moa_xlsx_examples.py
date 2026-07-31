from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

FILES = [
    Path("/Users/stani/code/MCP/Documents maître d'ouvrage/260130 Tarare export Menuiseries.xlsx"),
    Path(
        "/Users/stani/code/MCP/Documents maître d'ouvrage/260130 Tarare Export Zones et Espaces.xlsx"
    ),
    Path(
        "/Users/stani/code/MCP/Documents maître d'ouvrage/260130 Tarare Extraction surface enveloppe.xlsx"
    ),
    Path(
        "/Users/stani/code/MCP/Documents maître d'ouvrage/260201 Tatare 0546L AVP - export SHAB maquette.xlsx"
    ),
    Path(
        "/Users/stani/code/MCP/Documents maître d'ouvrage/260203 Tatare 0546L AVP - export plancher.xlsx"
    ),
]


def non_empty_bounds(ws):
    rows = []
    cols = []
    count = 0
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                rows.append(cell.row)
                cols.append(cell.column)
                count += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.coordinate)
    if not rows:
        return None
    return {
        "min_row": min(rows),
        "max_row": max(rows),
        "min_col": min(cols),
        "max_col": max(cols),
        "non_empty_cells": count,
        "formula_cells": formulas[:30],
        "n_formula_cells": len(formulas),
    }


def row_values(ws, row_idx, max_col):
    vals = []
    for col in range(1, max_col + 1):
        v = ws.cell(row_idx, col).value
        if hasattr(v, "isoformat"):
            v = v.isoformat()
        vals.append(v)
    return vals


def likely_header_rows(ws, bounds):
    if not bounds:
        return []
    out = []
    for r in range(bounds["min_row"], min(bounds["max_row"], bounds["min_row"] + 25) + 1):
        vals = row_values(ws, r, bounds["max_col"])
        n_text = sum(1 for v in vals if isinstance(v, str) and v.strip())
        n_non_empty = sum(1 for v in vals if v not in (None, ""))
        if n_non_empty >= 2 and n_text >= max(1, n_non_empty // 2):
            out.append({"row": r, "values": vals})
    return out[:10]


def style_sample(ws, bounds):
    if not bounds:
        return []
    samples = []
    for r in range(bounds["min_row"], min(bounds["max_row"], bounds["min_row"] + 12) + 1):
        for c in range(bounds["min_col"], min(bounds["max_col"], bounds["min_col"] + 10) + 1):
            cell = ws.cell(r, c)
            if cell.value not in (None, ""):
                fill = str(
                    cell.fill.fgColor.rgb or cell.fill.fgColor.indexed or cell.fill.fgColor.type
                )
                font_color = (
                    str(cell.font.color.rgb)
                    if cell.font.color and cell.font.color.type == "rgb"
                    else None
                )
                samples.append(
                    {
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:80],
                        "font": {
                            "name": cell.font.name,
                            "size": cell.font.sz,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "color": font_color,
                        },
                        "fill": fill,
                        "number_format": cell.number_format,
                        "alignment": {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text,
                        },
                    }
                )
    return samples[:15]


def workbook_summary(path):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        sheets = []
        for ws in wb.worksheets:
            bounds = non_empty_bounds(ws)
            max_col = bounds["max_col"] if bounds else 0
            sheets.append(
                {
                    "name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "bounds": bounds,
                    "merged_ranges": [str(r) for r in ws.merged_cells.ranges][:20],
                    "tables": list(ws.tables.keys()),
                    "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                    "column_widths": {
                        col: ws.column_dimensions[col].width
                        for col in list(ws.column_dimensions)[:20]
                        if ws.column_dimensions[col].width
                    },
                    "row_heights": {
                        str(idx): dim.height
                        for idx, dim in list(ws.row_dimensions.items())[:20]
                        if dim.height
                    },
                    "header_candidates": likely_header_rows(ws, bounds),
                    "style_samples": style_sample(ws, bounds),
                    "first_rows": [
                        {"row": r, "values": row_values(ws, r, max_col)}
                        for r in range(1, min(ws.max_row, 12) + 1)
                    ]
                    if bounds
                    else [],
                }
            )
        return {
            "file": str(path),
            "size": path.stat().st_size,
            "sheetnames": wb.sheetnames,
            "defined_names": sorted(d.name for d in wb.defined_names.values()),
            "sheets": sheets,
        }
    finally:
        wb.close()


def _short_value(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _compact_sheet_summary(ws):
    bounds = non_empty_bounds(ws)
    if not bounds:
        return {
            "name": ws.title,
            "bounds": None,
            "headers": [],
            "formulas": [],
            "summary_rows": [],
        }
    max_col = bounds["max_col"]
    header_rows = likely_header_rows(ws, bounds)[:4]
    formulas = []
    summary_rows = []
    for row in ws.iter_rows():
        values = [_short_value(cell.value) for cell in row[:max_col]]
        if any(isinstance(v, str) and v.startswith("=") for v in values):
            for cell in row[:max_col]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
        text = " ".join(str(v) for v in values if v not in (None, ""))
        low = text.lower()
        if any(
            token in low
            for token in (
                "total général",
                "nombre de types",
                "superficie des",
                "ratio fac",
                "seuil 3f",
                "shab",
                "ecart",
                "écart",
            )
        ):
            summary_rows.append({"row": row[0].row, "values": values})
    return {
        "name": ws.title,
        "bounds": bounds,
        "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        "headers": header_rows,
        "formulas": formulas[:24],
        "formula_count": len(formulas),
        "summary_rows": summary_rows[:16],
        "column_widths": {
            col: ws.column_dimensions[col].width
            for col in ws.column_dimensions
            if ws.column_dimensions[col].width
        },
        "row_heights": {
            str(idx): dim.height for idx, dim in ws.row_dimensions.items() if dim.height
        },
    }


def compact_workbook_summary(path):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        return {
            "file": str(path),
            "sheetnames": wb.sheetnames,
            "sheets": [_compact_sheet_summary(ws) for ws in wb.worksheets],
        }
    finally:
        wb.close()


if __name__ == "__main__":
    fn = compact_workbook_summary if "--compact" in sys.argv else workbook_summary
    print(json.dumps([fn(p) for p in FILES], ensure_ascii=False, indent=2))
