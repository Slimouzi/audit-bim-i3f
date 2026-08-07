"""Inventaire de la liste de contrôle Domofrance — description, pas conformité.

Ce script **décrit** le classeur du maître d'ouvrage. Il ne lit aucune maquette,
n'émet aucun statut de conformité, et ne dit jamais qu'un contrôle est
« évaluable » : ce verdict-là suppose une preuve géométrique qui n'existe pas
encore (cf. ``docs/scope-domofrance-controls.md``).

Deux sorties :

- ``--summary`` (défaut) : les compteurs figés dans le scope, tous mesurés ici.
- ``--csv`` : une ligne par contrôle, avec sa ligne source et ses signaux
  lexicaux, pour relecture humaine.

Les signaux lexicaux sont une **hypothèse outillable**, pas une classification
validée : ils disent quels mots porte le texte du contrôle, pas ce qu'il faut
mesurer. Ils sont non exclusifs — un contrôle peut n'en porter aucun, ou
plusieurs. Aucun compteur de ce script ne doit être présenté au client comme un
taux de couverture.

Usage::

    python scripts/inventory_domofrance_controls.py <Liste de contrôle.xlsx>
    python scripts/inventory_domofrance_controls.py <fichier.xlsx> --csv > out.csv
"""

from __future__ import annotations

import csv
import re
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass, field

CONTROL_SHEET = "LISTE DE CONTROLE"
CONTROL_HEADER_ROW = 3
CONTROL_FIRST_ROW = 4
CONTROL_COLUMNS = ("type_logement", "zone", "element", "verification", "description")
CONTROL_COLUMN_INDEXES = (3, 4, 5, 6, 7)

SURFACE_SHEET = "SURFACE"
SURFACE_HEADER_ROW = 5
SURFACE_FIRST_ROW = 6
SURFACE_TOTAL_ROW = 19
SURFACE_CAPTION_ROW = 20

#: Familles de signaux lexicaux. Les motifs sont cherchés dans le texte
#: **normalisé** (cf. :func:`_normalize`), qui est borné par des espaces : c'est
#: ce qui permet à ``" cm "`` ou ``" m2"`` de ne pas capturer un mot plus long.
SIGNALS: dict[str, tuple[str, ...]] = {
    "needs_bbox": (
        "largeur",
        "hauteur",
        "longueur",
        "profondeur",
        "dimension",
        "surface",
        "superficie",
        "epaisseur",
        "diametre",
        "emprise",
        " metre",
        " cm ",
        " m2",
        "hauteur sous",
    ),
    "needs_collision": (
        "obstacle",
        "encombre",
        "degagement",
        "recoin",
        "saillie",
        "ressaut",
        "libre de tout",
        "sans gene",
        "chevauch",
        "empiete",
    ),
    "needs_space_context": (
        "presence",
        "situe",
        "positionn",
        "emplacement",
        "localisation",
        "proximite",
        "a proximite",
        "depuis",
        "dans le local",
        "dans la piece",
        "acces",
        "cheminement",
        "circulation",
        "attenant",
        "contigu",
    ),
    "manual_only": (
        "recommand",
        "a proscrire",
        "souhaitable",
        "de preference",
        "preferentiel",
        "si possible",
        "dans la mesure du possible",
        "optimiser",
        "aise",
        "adapte",
        "qualite",
        "confort",
        "esthetique",
        "harmonis",
        "vigilance",
        "veiller",
        "pertinen",
        "suffisant",
        "convenable",
        "eviter",
    ),
}

#: Les trois familles qui supposent une mesure sur la maquette. ``manual_only``
#: en est exclu : c'est un signal d'appréciation humaine, pas de géométrie.
GEOMETRIC_SIGNALS = ("needs_bbox", "needs_collision", "needs_space_context")

#: Un seuil est un nombre **porteur d'une unité**. « 30% des boîtes » ou
#: « format A2 » sont des nombres, pas des seuils dimensionnels.
NUMERIC_THRESHOLD = re.compile(r"\d+[.,]?\d*\s*(?:m2|m\b|cm\b|mm\b|metre)")


def _normalize(value: object) -> str:
    """Minuscules, sans accents, ponctuation réduite à des espaces.

    L'ordre des deux opérations compte. Passer par ``encode("ascii", "ignore")``
    avant de remplacer la ponctuation supprime l'apostrophe typographique
    **sans laisser d'espace** : « rampes d'accès » devenait ``rampesdacces``
    accolé, et aucun motif contenant « d accès » ne pouvait plus le voir. On
    retire donc d'abord les accents (marques combinantes), puis on réduit tout
    ce qui n'est pas alphanumérique à un espace.

    Le résultat est **borné par des espaces** pour que les motifs qui portent
    leur propre délimiteur (``" cm "``) se comportent en début et en fin de
    texte comme au milieu.
    """
    if not isinstance(value, str):
        value = "" if value is None else str(value)
    decomposed = unicodedata.normalize("NFD", value)
    without_accents = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    collapsed = re.sub(r"[^0-9A-Za-z]+", " ", without_accents).lower().strip()
    return f" {collapsed} "


def _cell(value: object) -> str:
    """Texte d'une cellule, vide si elle ne porte rien."""
    if value is None:
        return ""
    return str(value).strip() if isinstance(value, str) else str(value)


@dataclass(frozen=True)
class Control:
    """Une ligne de la feuille de contrôles, telle qu'écrite par le client."""

    row: int
    type_logement: str
    zone: str
    element: str
    verification: str
    description: str
    signals: frozenset[str] = field(default_factory=frozenset)
    has_numeric_threshold: bool = False

    @property
    def identity(self) -> tuple[str, ...]:
        """Les cinq colonnes. Deux lignes de même identité sont un doublon."""
        return (
            self.type_logement,
            self.zone,
            self.element,
            self.verification,
            self.description,
        )

    @property
    def identity_without_type(self) -> tuple[str, ...]:
        """Les quatre colonnes hors TYPE DE LOGEMENT.

        Un même contrôle peut être écrit deux fois, une fois par type de
        logement. Le compter une seule fois est une lecture légitime — mais
        c'est une lecture différente, donc un compteur différent.
        """
        return self.identity[1:]

    @property
    def needs_geometry(self) -> bool:
        """Le texte porte au moins un signal géométrique.

        **Ce n'est pas un verdict d'évaluabilité.** Voir le scope : ce dérivé
        sature à 80 % et n'a aucune valeur de couverture.
        """
        return any(signal in self.signals for signal in GEOMETRIC_SIGNALS)


def _detect_signals(control_text: str) -> frozenset[str]:
    """Familles dont au moins un motif apparaît dans le texte normalisé."""
    return frozenset(
        name for name, patterns in SIGNALS.items() if any(p in control_text for p in patterns)
    )


def parse_controls(path: str) -> list[Control]:
    """Lit la feuille de contrôles. Aucune ligne n'est fusionnée ni dédoublonnée.

    Les doublons appartiennent au document du client : les écraser ici ferait
    disparaître un fait à lui rapporter.
    """
    import openpyxl

    sheet = openpyxl.load_workbook(path, data_only=True)[CONTROL_SHEET]
    controls: list[Control] = []
    for row in range(CONTROL_FIRST_ROW, sheet.max_row + 1):
        values = [_cell(sheet.cell(row, col).value) for col in CONTROL_COLUMN_INDEXES]
        if not any(values):
            continue
        control_text = _normalize(f"{values[3]} {values[4]}")
        controls.append(
            Control(
                row=row,
                type_logement=values[0],
                zone=values[1],
                element=values[2],
                verification=values[3],
                description=values[4],
                signals=_detect_signals(control_text),
                has_numeric_threshold=bool(NUMERIC_THRESHOLD.search(control_text)),
            )
        )
    return controls


@dataclass(frozen=True)
class SurfaceTable:
    """Une des deux tables de surfaces minimales (collectif / individuel)."""

    label: str
    caption: str
    typologies: tuple[str, ...]
    room_types: tuple[str, ...]
    has_width_column: bool
    numeric_cells: int
    non_numeric_cells: int
    total_row: int | None


def _is_numeric(text: str) -> bool:
    """Une valeur chiffrée, l'astérisque de renvoi mis à part.

    ``"2,5*"`` reste une surface ; ``"1,20/0,9"`` est une paire de largeurs
    écrite dans une seule cellule, et n'est pas un nombre.
    """
    candidate = text.replace("*", "").replace(",", ".").strip()
    if not candidate:
        return False
    try:
        float(candidate)
    except ValueError:
        return False
    return True


def _surface_table(sheet, label_col: int, label: str) -> SurfaceTable:
    """Décrit la table dont la colonne des types de pièces est ``label_col``."""
    typologies: list[str] = []
    has_width_column = False
    col = label_col + 1
    while col <= sheet.max_column:
        header = _cell(sheet.cell(SURFACE_HEADER_ROW, col).value)
        if not header:
            break
        if header.strip().upper().startswith("LARGEUR"):
            has_width_column = True
            break
        typologies.append(header)
        col += 1
    last_value_col = col - 1 if has_width_column else col

    room_types: list[str] = []
    total_row: int | None = None
    numeric_cells = 0
    non_numeric_cells = 0
    for row in range(SURFACE_FIRST_ROW, SURFACE_TOTAL_ROW + 1):
        name = _cell(sheet.cell(row, label_col).value)
        if not name:
            continue
        if name.strip().lower().startswith("total"):
            total_row = row
            continue
        room_types.append(name)
        for value_col in range(label_col + 1, last_value_col + 1):
            text = _cell(sheet.cell(row, value_col).value)
            if not text:
                continue
            if _is_numeric(text):
                numeric_cells += 1
            else:
                non_numeric_cells += 1

    return SurfaceTable(
        label=label,
        caption=_cell(sheet.cell(SURFACE_CAPTION_ROW, label_col).value),
        typologies=tuple(typologies),
        room_types=tuple(dict.fromkeys(room_types)),
        has_width_column=has_width_column,
        numeric_cells=numeric_cells,
        non_numeric_cells=non_numeric_cells,
        total_row=total_row,
    )


def parse_surface_tables(path: str) -> list[SurfaceTable]:
    """Les deux tables de la feuille SURFACE, repérées par leur titre."""
    import openpyxl

    sheet = openpyxl.load_workbook(path, data_only=True)[SURFACE_SHEET]
    tables: list[SurfaceTable] = []
    for label_col, expected in ((2, "LOGEMENT COLLECTIF"), (12, "LOGEMENT INDIVIDUEL")):
        title = _cell(sheet.cell(3, label_col).value)
        if title.strip().upper() == expected:
            tables.append(_surface_table(sheet, label_col, title))
    return tables


def distinct_controls(controls: list[Control]) -> list[Control]:
    """Un représentant par identité, dans l'ordre du classeur."""
    seen: dict[tuple[str, ...], Control] = {}
    for control in controls:
        seen.setdefault(control.identity, control)
    return list(seen.values())


def tooling_core(controls: list[Control]) -> list[Control]:
    """Les contrôles qu'un outil pourrait trancher — conjonction stricte.

    Trois conditions **cumulatives** : une grandeur nommée (``needs_bbox``), un
    seuil chiffré avec unité, et aucun vocabulaire d'appréciation. C'est cette
    conjonction qui fait chuter le compte de 331 à 30 : la route lexicale seule
    sature, parce que « présence » et « accès » attrapent presque tout le
    classeur sans rien rendre mesurable.

    Appeler cette fonction sur les 413 lignes plutôt que sur les distincts
    gonflerait le noyau des doublons du client.
    """
    return [
        c
        for c in controls
        if "needs_bbox" in c.signals and c.has_numeric_threshold and "manual_only" not in c.signals
    ]


def print_summary(controls: list[Control], tables: list[SurfaceTable]) -> None:
    """Les compteurs que le scope fige, tous recalculés à chaque exécution."""
    print(f"STRUCTURE — feuille « {CONTROL_SHEET} »")
    print(f"  lignes de contrôle                  : {len(controls)}")
    print(f"  première / dernière ligne du fichier: {controls[0].row} / {controls[-1].row}")
    print(f"  lignes distinctes (5 colonnes)      : {len({c.identity for c in controls})}")
    print(
        "  distinctes hors TYPE DE LOGEMENT    : "
        f"{len({c.identity_without_type for c in controls})}"
    )
    print(f"  zones distinctes                    : {len({c.zone for c in controls})}")
    print(f"  éléments distincts                  : {len({c.element for c in controls})}")
    print(f"  libellés de vérification distincts  : {len({c.verification for c in controls})}")
    print(f"  descriptions distinctes             : {len({c.description for c in controls})}")

    print("  répartition par type de logement :")
    for type_logement, count in Counter(c.type_logement for c in controls).most_common():
        print(f"    {type_logement or '(vide)':32} {count:4}")

    groups = {identity: n for identity, n in Counter(c.identity for c in controls).items() if n > 1}
    # Deux lectures, volontairement toutes les deux publiées : elles ne
    # répondent pas à la même question, et n'en donner qu'une invite à citer le
    # mauvais nombre. Une ligne présente 7 fois, c'est 7 lignes concernées et
    # 6 répétitions.
    print(f"  lignes impliquées dans un doublon   : {sum(groups.values())}")
    print(f"  dont répétitions (hors 1re occurr.) : {sum(n - 1 for n in groups.values())}")
    print(f"  groupes de doublons                 : {len(groups)}")
    print(f"  plus grand groupe                   : {max(groups.values(), default=0)}")

    print()
    print("SIGNAUX LEXICAUX — hypothèse d'outillage, PAS une couverture")
    for name in SIGNALS:
        print(f"  {name:22} {sum(1 for c in controls if name in c.signals):4}")
    print(f"  {'needs_geometry (dérivé)':22} {sum(1 for c in controls if c.needs_geometry):4}")
    print(
        f"  {'seuil chiffré dans le texte':22} "
        f"{sum(1 for c in controls if c.has_numeric_threshold):4}"
    )
    print(f"  {'aucun signal':22} {sum(1 for c in controls if not c.signals):4}")
    print(
        f"  {'signal géométrique ET manual_only':22} "
        f"{sum(1 for c in controls if c.needs_geometry and 'manual_only' in c.signals):4}"
    )

    print()
    print("NOYAU OUTILLABLE — sur les contrôles DISTINCTS, pas sur les 413 lignes")
    distinct = distinct_controls(controls)
    core = tooling_core(distinct)
    print(f"  contrôles distincts                 : {len(distinct)}")
    print(f"  noyau outillable                    : {len(core)}")
    print(f"  soit                                : {100 * len(core) / len(distinct):.1f} %")

    print()
    print(f"TABLES DE SURFACES — feuille « {SURFACE_SHEET} »")
    for table in tables:
        print(f"  {table.label}")
        print(f"    typologies            : {len(table.typologies)} {list(table.typologies)}")
        print(f"    types de pièces       : {len(table.room_types)}")
        print(f"    colonne largeur mini  : {table.has_width_column}")
        print(f"    cellules numériques   : {table.numeric_cells}")
        print(f"    cellules non numériques: {table.non_numeric_cells}")
        print(f"    ligne de total        : {table.total_row}")
        print(f"    légende               : {table.caption}")


def print_csv(controls: list[Control]) -> None:
    """Une ligne par contrôle, pour relecture humaine."""
    writer = csv.writer(sys.stdout)
    writer.writerow(["row", *CONTROL_COLUMNS, *SIGNALS, "needs_geometry", "has_numeric_threshold"])
    for control in controls:
        writer.writerow(
            [
                control.row,
                control.type_logement,
                control.zone,
                control.element,
                control.verification,
                control.description,
                *(int(name in control.signals) for name in SIGNALS),
                int(control.needs_geometry),
                int(control.has_numeric_threshold),
            ]
        )


def main(argv: list[str]) -> int:
    if len(argv) < 2 or argv[1] in frozenset({"--help", "-h"}):
        print(__doc__)
        return 0
    path = argv[1]
    controls = parse_controls(path)
    if len(argv) > 2 and argv[2] == "--csv":
        print_csv(controls)
        return 0
    print_summary(controls, parse_surface_tables(path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
